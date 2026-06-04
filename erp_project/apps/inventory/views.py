"""
Inventory Views - Categories, Warehouses, Items, Stock
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.views.generic import ListView, CreateView, UpdateView, DetailView
from django.urls import reverse, reverse_lazy
from django.db.models import Q, Sum, F, Value, DecimalField, DateField, Count, Avg, Prefetch
from django.db import models as db_models
from django.db.models.functions import Coalesce, TruncDate
from django.db import transaction
from django.utils.safestring import mark_safe
import csv
import json

from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.csrf import ensure_csrf_cookie
from django.utils.decorators import method_decorator
from decimal import Decimal

from .models import (
    Category,
    Warehouse,
    StorageLocation,
    Item,
    ItemGroup,
    ItemBaseGroup,
    ItemGroupMembership,
    ItemSerialNumber,
    Stock,
    StockMovement,
    ConsumableRequest,
    ConsumableRequestItem,
    ConsumableRequestAttachment,
    ConditionLog,
)
from .consumable_inventory_reports import REPORT_BUILDERS, build_report
from .consumable_report_export import export_report_pdf, export_report_xlsx
from apps.purchase.models import ItemPurchaseReceiptHistory
from .consumable_project_lines import sync_consumable_request_to_project_item_lines
from .serial_stock import annotate_item_available_stock, unregistered_on_hand_count, register_on_hand_model_numbers
from .forms import (
    CategoryForm, WarehouseForm, ItemForm, StockAdjustmentForm,
    ConsumableRequestForm, ConsumableRequestItemFormSet,
    ConsumableRequestApproveForm, ConsumableRequestRejectForm,
    StockTransferForm, ItemConditionForm,
)
from apps.core.mixins import PermissionRequiredMixin, CreatePermissionMixin, UpdatePermissionMixin
from apps.core.utils import PermissionChecker


def _consumable_with_consumption_date():
    """Dispensed requests annotated with calendar date of consumption (dispense date, else request)."""
    return ConsumableRequest.objects.filter(
        is_active=True,
        status='dispensed',
    ).annotate(
        _consumption_date=Coalesce(
            TruncDate('dispensed_date'),
            F('request_date'),
            output_field=DateField(),
        )
    )


def consumable_requests_in_consumption_period(month_start, month_end):
    """Consumption date in [month_start, month_end)."""
    return _consumable_with_consumption_date().filter(
        _consumption_date__gte=month_start,
        _consumption_date__lt=month_end,
    )


# ============ CATEGORY VIEWS ============

class CategoryListView(PermissionRequiredMixin, ListView):
    model = Category
    template_name = 'inventory/category_list.html'
    context_object_name = 'categories'
    module_name = 'inventory'
    permission_type = 'view'
    
    def get_queryset(self):
        queryset = Category.objects.filter(is_active=True)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | Q(code__icontains=search)
            )
        return queryset.order_by('-created_at', '-pk')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Categories'
        context['form'] = CategoryForm()
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'inventory', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'inventory', 'edit')
        context['can_delete'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'inventory', 'delete')
        return context
    
    def post(self, request, *args, **kwargs):
        if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'create')):
            messages.error(request, 'Permission denied.')
            return redirect('inventory:category_list')
        
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Category {category.name} created.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
        return redirect('inventory:category_list')


class CategoryUpdateView(UpdatePermissionMixin, UpdateView):
    model = Category
    form_class = CategoryForm
    template_name = 'inventory/category_form.html'
    success_url = reverse_lazy('inventory:category_list')
    module_name = 'inventory'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Category: {self.object.name}'
        return context
    
    def form_valid(self, form):
        messages.success(self.request, f'Category {form.instance.name} updated.')
        return super().form_valid(form)


@login_required
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'delete'):
        category.is_active = False
        category.save()
        messages.success(request, f'Category {category.name} deleted.')
    else:
        messages.error(request, 'Permission denied.')
    return redirect('inventory:category_list')


# ============ WAREHOUSE VIEWS ============

class WarehouseListView(PermissionRequiredMixin, ListView):
    model = Warehouse
    template_name = 'inventory/warehouse_list.html'
    context_object_name = 'warehouses'
    module_name = 'inventory'
    permission_type = 'view'
    
    def get_queryset(self):
        queryset = Warehouse.objects.filter(is_active=True)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | Q(code__icontains=search)
            )
        return queryset.order_by('-created_at', '-pk')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Warehouses'
        context['form'] = WarehouseForm()
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'inventory', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'inventory', 'edit')
        context['can_delete'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'inventory', 'delete')
        return context
    
    def post(self, request, *args, **kwargs):
        if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'create')):
            messages.error(request, 'Permission denied.')
            return redirect('inventory:warehouse_list')
        
        form = WarehouseForm(request.POST)
        if form.is_valid():
            warehouse = form.save()
            messages.success(request, f'Warehouse {warehouse.name} created.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
        return redirect('inventory:warehouse_list')


class WarehouseUpdateView(UpdatePermissionMixin, UpdateView):
    model = Warehouse
    form_class = WarehouseForm
    template_name = 'inventory/warehouse_form.html'
    success_url = reverse_lazy('inventory:warehouse_list')
    module_name = 'inventory'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Warehouse: {self.object.name}'
        return context
    
    def form_valid(self, form):
        messages.success(self.request, f'Warehouse {form.instance.name} updated.')
        return super().form_valid(form)


@login_required
def warehouse_delete(request, pk):
    warehouse = get_object_or_404(Warehouse, pk=pk)
    if request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'delete'):
        warehouse.is_active = False
        warehouse.save()
        messages.success(request, f'Warehouse {warehouse.name} deleted.')
    else:
        messages.error(request, 'Permission denied.')
    return redirect('inventory:warehouse_list')


# ============ ITEM VIEWS ============

@method_decorator(ensure_csrf_cookie, name='dispatch')
class ItemListView(PermissionRequiredMixin, ListView):
    model = Item
    template_name = 'inventory/item_list.html'
    context_object_name = 'items'
    module_name = 'inventory'
    permission_type = 'view'
    paginate_by = 25
    
    def get_queryset(self):
        # Annotate total_stock at database level to ensure fresh data
        queryset = Item.objects.filter(is_active=True).select_related('category').prefetch_related(
            'item_groups',
            Prefetch(
                'stock_records',
                queryset=Stock.objects.filter(
                    warehouse__is_active=True,
                    quantity__gt=0
                ).select_related('warehouse'),
                to_attr='active_stock_records'
            )
        )
        queryset = annotate_item_available_stock(queryset)
        queryset = annotate_item_available_stock(queryset)
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(item_code__icontains=search) |
                Q(name__icontains=search)
            )
        
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)
        
        item_type = self.request.GET.get('item_type')
        if item_type:
            queryset = queryset.filter(item_type=item_type)

        group = (self.request.GET.get('group') or '').strip()
        if group == '__none__':
            queryset = queryset.annotate(_gc=Count('item_groups')).filter(_gc=0)
        elif group.isdigit():
            queryset = queryset.filter(item_groups__pk=int(group)).distinct()

        return queryset.order_by('-created_at', '-pk')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Items'
        context['categories'] = Category.objects.filter(is_active=True).order_by('name')
        context['type_choices'] = Item.TYPE_CHOICES
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'inventory', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'inventory', 'edit')
        context['can_delete'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'inventory', 'delete')
        
        # Stats
        items = self.get_queryset()
        context['total_items'] = items.count()
        # Use annotation for low stock check
        context['low_stock_count'] = sum(
            1 for item in items 
            if item.item_type == 'product' 
            and (item.total_stock_calc or Decimal('0.00')) < item.minimum_stock
        )

        context['item_groups'] = list(ItemGroup.objects.all().order_by('name'))
        q = self.request.GET.copy()
        q.pop('page', None)
        context['filter_querystring'] = q.urlencode()
        
        return context


@login_required
def item_export_csv(request):
    """Export all active inventory items as CSV (reference for estimate line-item import)."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'view')):
        return HttpResponseForbidden('Permission denied.')

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="inventory_items.csv"'
    response.write('\ufeff')
    w = csv.writer(response)
    w.writerow(
        [
            'item_code',
            'name',
            'description',
            'category_code',
            'groups',
            'item_type',
            'unit',
            'purchase_price',
            'selling_price',
            'minimum_selling_price',
            'maximum_selling_price',
            'tax_code',
            'vat_rate',
            'minimum_stock',
            'status',
            'barcode',
            'storage_location',
            'condition_status',
        ]
    )
    qs = (
        Item.objects.filter(is_active=True)
        .select_related('category', 'tax_code', 'storage_location_master')
        .prefetch_related('item_groups')
        .order_by('item_code')
    )
    for item in qs:
        group_names = ' | '.join(item.item_groups.order_by('name').values_list('name', flat=True))
        w.writerow(
            [
                item.item_code,
                item.name,
                (item.description or '').replace('\n', ' ').replace('\r', ' ')[:2000],
                item.category.code if item.category and item.category.code else (item.category.name if item.category else ''),
                group_names,
                item.item_type,
                item.unit,
                item.purchase_price,
                item.selling_price,
                item.minimum_selling_price,
                item.maximum_selling_price,
                item.tax_code.code if item.tax_code_id else '',
                item.vat_rate,
                item.minimum_stock,
                item.status,
                item.barcode or '',
                item.get_storage_shelf_label(),
                item.condition_status,
            ]
        )
    return response


def _ordered_base_group_sub_ids(sub_ids, post=None):
    """Sort sub-group PKs by optional sub_group_order_<pk> fields from POST."""
    ordered = [int(x) for x in sub_ids if str(x).isdigit()]
    if not post:
        return ordered
    orders = {}
    for pk in ordered:
        raw = post.get(f'sub_group_order_{pk}')
        try:
            orders[pk] = int(str(raw).strip())
        except (TypeError, ValueError):
            orders[pk] = 999999
    return [
        ordered[i]
        for i in sorted(
            range(len(ordered)),
            key=lambda idx: (orders.get(ordered[idx], 999999), idx),
        )
    ]


def _apply_base_group_sub_groups(base_group, sub_ids, post=None):
    """Link sub-groups to a base group preserving explicit order from POST / UI."""
    ordered = _ordered_base_group_sub_ids(sub_ids, post)
    ItemGroup.objects.filter(base_group=base_group).exclude(pk__in=ordered).update(
        base_group=None,
        base_group_sort_order=0,
    )
    for order, pk in enumerate(ordered):
        ItemGroup.objects.filter(pk=pk).update(
            base_group=base_group,
            base_group_sort_order=order,
        )


def item_group_manage(request):
    """Manage item sub-groups and base groups: members, default estimate qty, rename, PDF settings."""
    can_edit = request.user.is_superuser or PermissionChecker.has_permission(
        request.user, 'inventory', 'edit'
    )
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('inventory:item_list')

    active_tab = (request.GET.get('tab') or 'base').strip().lower()
    if active_tab not in ('sub', 'base'):
        active_tab = 'base'

    def _redirect(group=None, tab=None, base=None):
        tab = tab or active_tab
        url = reverse('inventory:item_group_manage')
        params = [f'tab={tab}']
        if tab == 'base':
            if base:
                params.append(f'base={base.pk}')
        elif group:
            params.append(f'group={group.pk}')
        return redirect(f'{url}?{"&".join(params)}')

    if request.method == 'POST':
        if not can_edit:
            messages.error(request, 'Permission denied.')
            return redirect('inventory:item_list')

        action = (request.POST.get('action') or '').strip()
        post_tab = (request.POST.get('tab') or active_tab).strip().lower()
        if post_tab not in ('sub', 'base'):
            post_tab = active_tab

        if action == 'create_base_group':
            name = (request.POST.get('new_base_group_name') or '').strip()[:200]
            if not name:
                messages.warning(request, 'Enter a base group name.')
                return _redirect(tab='base')
            if ItemBaseGroup.objects.filter(name__iexact=name).exists():
                messages.error(request, 'A base group with that name already exists.')
                return _redirect(tab='base')
            bg = ItemBaseGroup.objects.create(name=name)
            sub_ids = request.POST.getlist('sub_group_ids')
            _apply_base_group_sub_groups(bg, sub_ids, request.POST)
            if sub_ids:
                messages.success(
                    request,
                    f'Base group "{bg.name}" created with {len(sub_ids)} sub-group(s).',
                )
            else:
                messages.success(request, f'Base group "{bg.name}" created.')
            return _redirect(tab='base', base=bg)

        base_pk = request.POST.get('base_id')

        if action == 'save_base_group':
            if not base_pk or not str(base_pk).isdigit():
                messages.warning(request, 'Choose a base group first.')
                return _redirect(tab='base')
            base_group = get_object_or_404(ItemBaseGroup, pk=int(base_pk))
            new_name = (request.POST.get('new_base_group_name') or '').strip()[:200]
            if not new_name:
                messages.warning(request, 'Enter a base group name.')
                return _redirect(tab='base', base=base_group)
            if ItemBaseGroup.objects.filter(name__iexact=new_name).exclude(pk=base_group.pk).exists():
                messages.error(request, 'Another base group already uses that name.')
                return _redirect(tab='base', base=base_group)
            sub_ids = request.POST.getlist('sub_group_ids')
            base_group.name = new_name
            base_group.save(update_fields=['name'])
            _apply_base_group_sub_groups(base_group, sub_ids, request.POST)
            messages.success(request, f'Base group "{base_group.name}" saved.')
            return _redirect(tab='base', base=base_group)

        if action in ('delete_base', 'remove_sub_group'):
            if not base_pk or not str(base_pk).isdigit():
                messages.warning(request, 'Choose a base group first.')
                return _redirect(tab='base')
            base_group = get_object_or_404(ItemBaseGroup, pk=int(base_pk))

            if action == 'delete_base':
                name = base_group.name
                base_group.delete()
                messages.success(request, f'Base group "{name}" deleted. Sub-groups are now unassigned.')
                return _redirect(tab='base')

            if action == 'remove_sub_group':
                sub_pk = request.POST.get('sub_group_id')
                if sub_pk and str(sub_pk).isdigit():
                    sub = ItemGroup.objects.filter(pk=int(sub_pk), base_group=base_group).first()
                    if sub:
                        sub.base_group = None
                        sub.save(update_fields=['base_group'])
                        messages.success(request, f'Removed "{sub.name}" from base group.')
                    else:
                        messages.warning(request, 'Sub-group not found in this base group.')
                return _redirect(tab='base', base=base_group)

        group_pk = request.POST.get('group_id')

        if action == 'create_group':
            name = (request.POST.get('new_group_name') or '').strip()[:200]
            if not name:
                messages.warning(request, 'Enter a group name.')
                return _redirect(tab='sub')
            if ItemGroup.objects.filter(name__iexact=name).exists():
                messages.error(request, 'A group with that name already exists.')
                return _redirect(tab='sub')
            g = ItemGroup.objects.create(name=name)
            messages.success(request, f'Group "{g.name}" created.')
            return _redirect(group=g, tab='sub')

        if not group_pk or not str(group_pk).isdigit():
            messages.warning(request, 'Choose a group first.')
            return _redirect(tab='sub')
        group = get_object_or_404(ItemGroup, pk=int(group_pk))

        if action == 'rename':
            new_name = (request.POST.get('rename_new_name') or '').strip()[:200]
            if not new_name:
                messages.warning(request, 'Enter a new name.')
                return _redirect(group=group, tab='sub')
            if ItemGroup.objects.filter(name__iexact=new_name).exclude(pk=group.pk).exists():
                messages.error(request, 'Another group already uses that name.')
                return _redirect(group=group, tab='sub')
            group.name = new_name
            group.save(update_fields=['name'])
            messages.success(request, 'Group renamed.')
            return _redirect(group=group, tab='sub')

        if action == 'set_pdf_hide':
            hide = request.POST.get('hide_items_on_pdf') == 'on'
            group.hide_items_on_pdf = hide
            group.save(update_fields=['hide_items_on_pdf'])
            if hide:
                messages.success(
                    request,
                    f'"{group.name}" will show as one consolidated line on quotation PDFs.',
                )
            else:
                messages.success(
                    request,
                    f'"{group.name}" will show individual items on quotation PDFs.',
                )
            return _redirect(group=group, tab='sub')

        if action == 'save_quantities':
            updated = 0
            for key, val in request.POST.items():
                if not key.startswith('qty_'):
                    continue
                mid = key[4:]
                if not str(mid).isdigit():
                    continue
                membership = ItemGroupMembership.objects.filter(pk=int(mid), group=group).first()
                if not membership:
                    continue
                try:
                    qty = Decimal(str(val).strip().replace(',', ''))
                except Exception:
                    continue
                if qty < Decimal('1'):
                    qty = Decimal('1')
                membership.default_quantity = qty.quantize(Decimal('0.01'))
                membership.save(update_fields=['default_quantity'])
                updated += 1
            order_updated = 0
            for key, val in request.POST.items():
                if not key.startswith('order_'):
                    continue
                mid = key[6:]
                if not str(mid).isdigit():
                    continue
                membership = ItemGroupMembership.objects.filter(pk=int(mid), group=group).first()
                if not membership:
                    continue
                try:
                    order = int(str(val).strip())
                except (TypeError, ValueError):
                    continue
                if order < 0:
                    order = 0
                membership.sort_order = order
                membership.save(update_fields=['sort_order'])
                order_updated += 1
            if order_updated:
                messages.success(
                    request,
                    f'Saved quantities for {updated} item(s) and order for {order_updated} item(s).',
                )
            else:
                messages.success(request, f'Saved quantities for {updated} item(s).')
            return _redirect(group=group, tab='sub')

        if action == 'add_item':
            item_id = request.POST.get('add_item_id')
            if not item_id or not str(item_id).isdigit():
                messages.warning(request, 'Choose an item to add.')
                return _redirect(group=group, tab='sub')
            item = get_object_or_404(Item, pk=int(item_id), is_active=True)
            try:
                qty = Decimal(str(request.POST.get('add_item_qty') or '1').strip().replace(',', ''))
            except Exception:
                qty = Decimal('1')
            if qty < Decimal('1'):
                qty = Decimal('1')
            next_order = (
                ItemGroupMembership.objects.filter(group=group)
                .order_by('-sort_order')
                .values_list('sort_order', flat=True)
                .first()
            )
            next_order = (next_order or 0) + 1
            membership, created = ItemGroupMembership.objects.get_or_create(
                group=group,
                item=item,
                defaults={
                    'default_quantity': qty.quantize(Decimal('0.01')),
                    'sort_order': next_order,
                },
            )
            if not created:
                messages.info(request, f'{item.item_code} is already in this group.')
            else:
                messages.success(request, f'Added {item.item_code} to "{group.name}".')
            return _redirect(group=group, tab='sub')

        if action == 'remove_member':
            mid = request.POST.get('membership_id')
            if mid and str(mid).isdigit():
                deleted, _ = ItemGroupMembership.objects.filter(pk=int(mid), group=group).delete()
                if deleted:
                    messages.success(request, 'Item removed from group.')
                else:
                    messages.warning(request, 'Item not found in this group.')
            return _redirect(group=group, tab='sub')

        if action == 'delete_group':
            name = group.name
            group.delete()
            messages.success(request, f'Group "{name}" deleted.')
            return _redirect(tab='sub')

        messages.error(request, 'Invalid action.')
        return _redirect(group=group, tab='sub')

    groups = (
        ItemGroup.objects.annotate(member_count=Count('memberships'))
        .select_related('base_group')
        .order_by('name')
    )
    selected = None
    memberships = []
    available_items = Item.objects.none()
    group_param = (request.GET.get('group') or '').strip()
    if group_param.isdigit():
        selected = ItemGroup.objects.filter(pk=int(group_param)).first()
    if not selected and groups.exists() and active_tab == 'sub':
        selected = groups.first()

    if selected:
        memberships = (
            ItemGroupMembership.objects.filter(group=selected)
            .select_related('item')
            .order_by('sort_order', 'item__item_code', 'pk')
        )
        member_ids = memberships.values_list('item_id', flat=True)
        available_items = (
            Item.objects.filter(is_active=True, status='active')
            .exclude(pk__in=member_ids)
            .order_by('item_code', 'name')[:500]
        )

    base_groups = (
        ItemBaseGroup.objects.annotate(sub_group_count=Count('sub_groups'))
        .order_by('name')
    )
    selected_base = None
    base_sub_groups = ItemGroup.objects.none()
    all_sub_groups = (
        ItemGroup.objects.annotate(member_count=Count('memberships'))
        .select_related('base_group')
        .order_by('name')
    )
    base_param = (request.GET.get('base') or '').strip()
    if base_param.isdigit():
        selected_base = ItemBaseGroup.objects.filter(pk=int(base_param)).first()
    elif active_tab == 'base' and not request.GET.get('new') and base_groups.exists():
        selected_base = base_groups.first()

    if selected_base:
        base_sub_groups = (
            ItemGroup.objects.filter(base_group=selected_base)
            .annotate(member_count=Count('memberships'))
            .order_by('base_group_sort_order', 'name')
        )

    sub_group_picker_data = {
        'all': [],
        'selected': [],
    }
    for sg in all_sub_groups:
        other_base = None
        if sg.base_group_id:
            if not selected_base or request.GET.get('new') or sg.base_group_id != selected_base.pk:
                other_base = sg.base_group.name if sg.base_group_id else None
        sub_group_picker_data['all'].append({
            'id': sg.pk,
            'name': sg.name,
            'member_count': sg.member_count,
            'other_base': other_base,
        })
    if selected_base and not request.GET.get('new'):
        sub_group_picker_data['selected'] = [
            {
                'id': sg.pk,
                'order': (sg.base_group_sort_order or 0) + 1,
            }
            for sg in ItemGroup.objects.filter(base_group=selected_base)
            .order_by('base_group_sort_order', 'name')
        ]

    return render(
        request,
        'inventory/item_group_manage.html',
        {
            'title': 'Item groups',
            'active_tab': active_tab,
            'groups': groups,
            'selected_group': selected,
            'memberships': memberships,
            'available_items': available_items,
            'base_groups': base_groups,
            'selected_base_group': selected_base,
            'base_sub_groups': base_sub_groups,
            'all_sub_groups': all_sub_groups,
            'sub_group_picker_data': sub_group_picker_data,
            'can_edit': can_edit,
        },
    )


class ItemCreateView(CreatePermissionMixin, CreateView):
    model = Item
    form_class = ItemForm
    template_name = 'inventory/item_form.html'
    success_url = reverse_lazy('inventory:item_list')
    module_name = 'inventory'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Item'
        context['storage_locations'] = StorageLocation.objects.filter(is_active=True).order_by('name')
        context['item_serial_numbers'] = []
        return context
    
    def form_valid(self, form):
        messages.success(self.request, f'Item {form.instance.name} created.')
        return super().form_valid(form)


class ItemUpdateView(UpdatePermissionMixin, UpdateView):
    model = Item
    form_class = ItemForm
    template_name = 'inventory/item_form.html'
    success_url = reverse_lazy('inventory:item_list')
    module_name = 'inventory'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Item: {self.object.name}'
        context['storage_locations'] = StorageLocation.objects.filter(is_active=True).order_by('name')
        context['item_serial_numbers'] = (
            ItemSerialNumber.objects.filter(item=self.object, is_active=True)
            .select_related('assigned_project', 'warehouse')
            .order_by('-date_received', 'model_number')[:50]
        )
        return context
    
    def form_valid(self, form):
        messages.success(self.request, f'Item {form.instance.name} updated.')
        return super().form_valid(form)


class ItemDetailView(PermissionRequiredMixin, DetailView):
    model = Item
    template_name = 'inventory/item_detail.html'
    context_object_name = 'item'
    module_name = 'inventory'
    permission_type = 'view'
    
    def get_queryset(self):
        return annotate_item_available_stock(
            Item.objects.prefetch_related('item_groups')
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Item: {self.object.name}'
        context['stock_records'] = Stock.objects.filter(
            item=self.object,
            warehouse__is_active=True
        ).select_related('warehouse')
        context['movements'] = StockMovement.objects.filter(
            item=self.object
        ).select_related('warehouse', 'to_warehouse')[:50]
        context['purchase_receipt_history'] = (
            ItemPurchaseReceiptHistory.objects.filter(item=self.object, is_active=True)
            .select_related(
                'vendor',
                'purchase_order',
                'receipt',
                'receipt__created_by',
                'stock_movement',
            )
            .order_by('-created_at')[:100]
        )
        context['condition_logs'] = ConditionLog.objects.filter(
            item=self.object
        ).select_related('changed_by')[:20]
        context['condition_form'] = ItemConditionForm(initial={
            'condition_status': self.object.condition_status
        })
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'inventory', 'edit')
        context['item_serial_numbers'] = (
            ItemSerialNumber.objects.filter(item=self.object, is_active=True)
            .select_related('assigned_project', 'warehouse', 'delivered_by')
            .order_by('-date_received', 'model_number')
        )
        context['unregistered_on_hand_count'] = unregistered_on_hand_count(self.object)
        # Transfer form
        context['transfer_form'] = StockTransferForm(initial={'item': self.object.pk})
        context['warehouses'] = Warehouse.objects.filter(is_active=True, status='active').order_by('name')
        return context


@login_required
@require_POST
def item_register_on_hand_serials(request, pk):
    """Register model numbers for units received before serial tracking was enabled."""
    item = get_object_or_404(Item, pk=pk, is_active=True)
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('inventory:item_detail', pk=pk)

    bulk = (request.POST.get('bulk_model_numbers') or '').strip()
    if bulk:
        model_numbers = [ln.strip() for ln in bulk.splitlines() if ln.strip()]
    else:
        model_numbers = request.POST.getlist('model_number')

    try:
        registered = register_on_hand_model_numbers(item, model_numbers, request.user)
        messages.success(
            request,
            f'Registered {len(registered)} model number(s) for on-hand stock.',
        )
    except ValidationError as exc:
        msgs = exc.messages if hasattr(exc, 'messages') else [str(exc)]
        for msg in msgs:
            messages.error(request, msg)

    return redirect('inventory:item_detail', pk=pk)


@login_required
def item_delete(request, pk):
    item = get_object_or_404(Item, pk=pk)
    if request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'delete'):
        item.is_active = False
        item.save()
        messages.success(request, f'Item {item.name} deleted.')
    else:
        messages.error(request, 'Permission denied.')
    return redirect('inventory:item_list')


# ============ STOCK VIEWS ============

class StockListView(PermissionRequiredMixin, ListView):
    model = Stock
    template_name = 'inventory/stock_list.html'
    context_object_name = 'stocks'
    module_name = 'inventory'
    permission_type = 'view'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = Stock.objects.filter(
            item__is_active=True,
            warehouse__is_active=True
        ).select_related('item', 'warehouse')
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(item__name__icontains=search) |
                Q(item__item_code__icontains=search)
            )
        
        warehouse = self.request.GET.get('warehouse')
        if warehouse:
            queryset = queryset.filter(warehouse_id=warehouse)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Stock Levels'
        context['warehouses'] = Warehouse.objects.filter(is_active=True, status='active').order_by('name')
        context['can_adjust'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'inventory', 'edit')
        return context


@login_required
def stock_adjustment(request):
    """Stock adjustment view."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('inventory:stock_list')
    
    if request.method == 'POST':
        form = StockAdjustmentForm(request.POST)
        if form.is_valid():
            item = form.cleaned_data['item']
            warehouse = form.cleaned_data['warehouse']
            quantity = Decimal(str(form.cleaned_data['quantity']))
            movement_type = form.cleaned_data['movement_type']
            adjustment_reason = form.cleaned_data.get('adjustment_reason', '')
            reference = form.cleaned_data['reference']
            notes = form.cleaned_data['notes']

            try:
                with transaction.atomic():
                    from datetime import date as dt_date

                    if movement_type in ('out', 'adjustment_minus'):
                        stock_record = Stock.objects.filter(item=item, warehouse=warehouse).first()
                        available = stock_record.quantity if stock_record else Decimal('0.00')
                        if available < quantity:
                            messages.error(request, f'Insufficient stock. Available: {available}, Requested: {quantity}')
                            items = Item.objects.filter(is_active=True).order_by('name')
                            warehouses = Warehouse.objects.filter(is_active=True, status='active').order_by('name')
                            return render(request, 'inventory/stock_adjustment.html', {
                                'title': 'Stock Adjustment',
                                'form': form,
                                'items': items,
                                'warehouses': warehouses,
                            })

                    old_quantity = Stock.objects.filter(
                        item=item, warehouse=warehouse
                    ).values_list('quantity', flat=True).first() or Decimal('0.00')

                    movement = StockMovement.objects.create(
                        item=item,
                        warehouse=warehouse,
                        movement_type=movement_type,
                        adjustment_reason=adjustment_reason,
                        source='manual',
                        quantity=quantity,
                        unit_cost=item.purchase_price or Decimal('0.00'),
                        reference=reference,
                        notes=notes,
                        movement_date=dt_date.today(),
                        created_by=request.user,
                    )

                    # Atomic: update quantity + post GL together
                    movement.execute(user=request.user)

                    new_quantity = Stock.objects.filter(
                        item=item, warehouse=warehouse
                    ).values_list('quantity', flat=True).first() or Decimal('0.00')

                    messages.success(request, f'Stock adjusted for {item.name} at {warehouse.name}. Quantity: {old_quantity} → {new_quantity}')
                    return redirect('inventory:stock_list')
                    
            except Exception as e:
                messages.error(request, f'Error updating stock: {str(e)}')
                items = Item.objects.filter(is_active=True).order_by('name')
                warehouses = Warehouse.objects.filter(is_active=True, status='active').order_by('name')
                return render(request, 'inventory/stock_adjustment.html', {
                    'title': 'Stock Adjustment',
                    'form': form,
                    'items': items,
                    'warehouses': warehouses,
                })
    else:
        form = StockAdjustmentForm()
    
    # Get items and warehouses for template context
    items = Item.objects.filter(is_active=True).order_by('name')
    warehouses = Warehouse.objects.filter(is_active=True, status='active').order_by('name')
    
    return render(request, 'inventory/stock_adjustment.html', {
        'title': 'Stock Adjustment',
        'form': form,
        'items': items,
        'warehouses': warehouses,
    })


class MovementListView(PermissionRequiredMixin, ListView):
    model = StockMovement
    template_name = 'inventory/movement_list.html'
    context_object_name = 'movements'
    module_name = 'inventory'
    permission_type = 'view'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = StockMovement.objects.filter(
            item__is_active=True
        ).select_related('item', 'warehouse', 'journal_entry')
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(item__name__icontains=search) |
                Q(reference__icontains=search) |
                Q(movement_number__icontains=search)
            )
        
        movement_type = self.request.GET.get('type')
        if movement_type:
            queryset = queryset.filter(movement_type=movement_type)
        
        posted = self.request.GET.get('posted')
        if posted == '1':
            queryset = queryset.filter(posted=True)
        elif posted == '0':
            queryset = queryset.filter(posted=False)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Stock Movements'
        context['type_choices'] = StockMovement.MOVEMENT_TYPE_CHOICES
        context['can_post'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'inventory', 'edit')
        
        # Calculate metrics
        all_movements = StockMovement.objects.filter(item__is_active=True)
        context['total_movements'] = all_movements.count()
        context['posted_movements'] = all_movements.filter(posted=True).count()
        context['unposted_movements'] = all_movements.filter(posted=False, total_cost__gt=0).count()
        context['total_value'] = all_movements.filter(posted=True).aggregate(Sum('total_cost'))['total_cost__sum'] or Decimal('0.00')
        
        return context


@login_required
def movement_post_to_accounting(request, pk):
    """Post a stock movement to accounting."""
    movement = get_object_or_404(StockMovement, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('inventory:movement_list')
    
    if movement.posted:
        messages.warning(request, f'Movement {movement.movement_number} already posted to accounting.')
        return redirect('inventory:movement_list')
    
    if movement.total_cost <= 0:
        messages.error(request, f'Movement {movement.movement_number} has no cost value. Update cost before posting.')
        return redirect('inventory:movement_list')
    
    try:
        movement.post_to_accounting(user=request.user)
        messages.success(request, f'Movement {movement.movement_number} posted to accounting. Journal Entry: {movement.journal_entry.entry_number}')
    except Exception as e:
        messages.error(request, f'Error posting to accounting: {str(e)}')
    
    return redirect('inventory:movement_list')


@login_required
def movement_export_excel(request):
    """Export stock movements to a formatted Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    queryset = StockMovement.objects.filter(
        item__is_active=True
    ).select_related('item', 'warehouse', 'to_warehouse', 'journal_entry')

    search = request.GET.get('search')
    if search:
        queryset = queryset.filter(
            Q(item__name__icontains=search) |
            Q(reference__icontains=search) |
            Q(movement_number__icontains=search)
        )

    movement_type = request.GET.get('type')
    if movement_type:
        queryset = queryset.filter(movement_type=movement_type)

    posted = request.GET.get('posted')
    if posted == '1':
        queryset = queryset.filter(posted=True)
    elif posted == '0':
        queryset = queryset.filter(posted=False)

    movements = queryset.order_by('-movement_date', '-id')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Stock Movements'

    headers = [
        'Movement #', 'Date', 'Item Code', 'Item Name', 'Warehouse',
        'Type', 'Source', 'Quantity', 'Unit Cost', 'Total Cost',
        'Reference', 'To Warehouse', 'Adjustment Reason',
        'GL Status', 'Journal #', 'Notes',
    ]
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    type_display = dict(StockMovement.MOVEMENT_TYPE_CHOICES)
    source_display = dict(StockMovement.SOURCE_CHOICES)
    reason_display = dict(StockMovement.ADJUSTMENT_REASON_CHOICES)

    posted_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    pending_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    number_fmt = '#,##0.00'

    for row_idx, m in enumerate(movements, 2):
        row_data = [
            m.movement_number,
            m.movement_date,
            m.item.item_code,
            m.item.name,
            m.warehouse.name,
            type_display.get(m.movement_type, m.movement_type),
            source_display.get(m.source, m.source),
            float(m.quantity),
            float(m.unit_cost),
            float(m.total_cost),
            m.reference,
            m.to_warehouse.name if m.to_warehouse else '',
            reason_display.get(m.adjustment_reason, m.adjustment_reason) if m.adjustment_reason else '',
            'Posted' if m.posted else ('Pending' if m.total_cost > 0 else 'No Cost'),
            m.journal_entry.entry_number if m.journal_entry else '',
            m.notes,
        ]
        row_fill = posted_fill if m.posted else (pending_fill if m.total_cost > 0 else None)
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
            if col_idx in (8, 9, 10):
                cell.number_format = number_fmt
                cell.alignment = Alignment(horizontal='right')
            if col_idx == 2:
                cell.number_format = 'DD/MM/YYYY'

    col_widths = [18, 14, 14, 28, 20, 16, 16, 12, 14, 16, 22, 20, 22, 12, 18, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Summary sheet
    ws_summary = wb.create_sheet('Summary')
    ws_summary.sheet_properties.tabColor = '10B981'
    summary_header_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')

    summary_title = ws_summary.cell(row=1, column=1, value='Stock Movement Summary')
    summary_title.font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:D1')

    all_movements = StockMovement.objects.filter(item__is_active=True)
    summary_data = [
        ('Total Movements', all_movements.count()),
        ('Posted to GL', all_movements.filter(posted=True).count()),
        ('Pending Posting', all_movements.filter(posted=False, total_cost__gt=0).count()),
        ('Total Posted Value', float(all_movements.filter(posted=True).aggregate(Sum('total_cost'))['total_cost__sum'] or 0)),
        ('', ''),
        ('By Type', ''),
    ]
    for choice_val, choice_label in StockMovement.MOVEMENT_TYPE_CHOICES:
        count = all_movements.filter(movement_type=choice_val).count()
        value = float(all_movements.filter(movement_type=choice_val).aggregate(Sum('total_cost'))['total_cost__sum'] or 0)
        summary_data.append((f'  {choice_label}', f'{count} movements — {value:,.2f} value'))

    for row_idx, (label, value) in enumerate(summary_data, 3):
        ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=2, value=value)
    ws_summary.column_dimensions['A'].width = 24
    ws_summary.column_dimensions['B'].width = 36

    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'
    ws.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="stock_movements_export.xlsx"'
    return response


@login_required
def movement_detail(request, pk):
    """View stock movement detail."""
    movement = get_object_or_404(
        StockMovement.objects.select_related('item', 'warehouse', 'to_warehouse', 'journal_entry'),
        pk=pk
    )
    
    context = {
        'title': f'Movement: {movement.movement_number}',
        'movement': movement,
        'can_post': not movement.posted and movement.total_cost > 0 and (
            request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'edit')
        ),
    }
    
    if movement.journal_entry:
        context['journal_lines'] = movement.journal_entry.lines.all().select_related('account')
    
    return render(request, 'inventory/movement_detail.html', context)


# ============ STOCK TRANSFER VIEW ============

@login_required
def stock_transfer(request):
    """Manual stock transfer between warehouses."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('inventory:movement_list')
    
    if request.method == 'POST':
        form = StockTransferForm(request.POST)
        if form.is_valid():
            item = form.cleaned_data['item']
            from_warehouse = form.cleaned_data['from_warehouse']
            to_warehouse = form.cleaned_data['to_warehouse']
            quantity = form.cleaned_data['quantity']
            reference = form.cleaned_data['reference']
            notes = form.cleaned_data['notes']
            
            try:
                with transaction.atomic():
                    from datetime import date
                    movement = StockMovement.objects.create(
                        item=item,
                        warehouse=from_warehouse,
                        to_warehouse=to_warehouse,
                        movement_type='transfer',
                        source='manual',
                        quantity=quantity,
                        unit_cost=item.purchase_price or Decimal('0.00'),
                        reference=reference or f'Manual transfer to {to_warehouse.name}',
                        notes=notes,
                        movement_date=date.today(),
                        created_by=request.user,
                    )

                    # Atomic: update quantity + post GL together
                    movement.execute(user=request.user)

                    messages.success(
                        request,
                        f'Successfully transferred {quantity} {item.unit} of {item.name} '
                        f'from {from_warehouse.name} to {to_warehouse.name}. '
                        f'Movement: {movement.movement_number}'
                    )
                    return redirect('inventory:movement_detail', pk=movement.pk)
                    
            except Exception as e:
                messages.error(request, f'Transfer failed: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, error)
                    else:
                        messages.error(request, f'{form.fields[field].label or field}: {error}')
    else:
        initial = {}
        item_id = request.GET.get('item')
        if item_id:
            initial['item'] = item_id
        form = StockTransferForm(initial=initial)
    
    return render(request, 'inventory/stock_transfer.html', {
        'title': 'Manual Stock Transfer',
        'form': form,
    })


@login_required
def item_change_condition(request, pk):
    """Change an item's condition status (in_store, in_use, repair, damaged)."""
    item = get_object_or_404(Item, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('inventory:item_detail', pk=pk)
    
    if request.method == 'POST':
        form = ItemConditionForm(request.POST)
        if form.is_valid():
            new_status = form.cleaned_data['condition_status']
            notes = form.cleaned_data['condition_notes']
            old_display = item.get_condition_status_display()
            
            item.change_condition(new_status, notes, user=request.user)
            
            new_display = item.get_condition_status_display()
            messages.success(
                request,
                f'Item condition updated: {old_display} → {new_display}'
            )
        else:
            messages.error(request, 'Invalid form data.')
    
    return redirect('inventory:item_detail', pk=pk)


# ============ CONSUMABLE REQUEST VIEWS ============

class ConsumableRequestListView(PermissionRequiredMixin, ListView):
    """
    List view for consumable requests.
    - Nurses see their own requests
    - Admin/Inventory see all requests
    """
    model = ConsumableRequest
    template_name = 'inventory/consumable_request_list.html'
    context_object_name = 'requests'
    module_name = 'inventory'
    permission_type = 'view'
    paginate_by = 25
    
    def get_queryset(self):
        user = self.request.user
        queryset = ConsumableRequest.objects.filter(is_active=True).select_related(
            'item',
            'requested_by',
            'warehouse',
            'approved_by',
            'dispensed_by',
            'department',
            'project',
        ).prefetch_related('items')
        
        # Non-admins only see their own requests
        if not user.is_superuser and not PermissionChecker.has_permission(user, 'inventory', 'edit'):
            queryset = queryset.filter(requested_by=user)
        
        # Filters
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(request_number__icontains=search) |
                Q(item__name__icontains=search) |
                Q(items__item__name__icontains=search) |
                Q(requested_by__first_name__icontains=search) |
                Q(requested_by__last_name__icontains=search) |
                Q(project__name__icontains=search) |
                Q(project__project_code__icontains=search)
            ).distinct()
        
        project_id = self.request.GET.get('project')
        if project_id and str(project_id).isdigit():
            queryset = queryset.filter(project_id=int(project_id))
        
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        is_admin = user.is_superuser or PermissionChecker.has_permission(user, 'inventory', 'edit')
        
        context['title'] = 'Consumable Requests'
        context['status_choices'] = ConsumableRequest.STATUS_CHOICES
        context['is_admin'] = is_admin
        context['filter_project_id'] = self.request.GET.get('project', '')
        
        # Stats (for admins)
        if is_admin:
            all_requests = ConsumableRequest.objects.filter(is_active=True)
            context['pending_count'] = all_requests.filter(status='pending').count()
            context['approved_count'] = all_requests.filter(status='approved').count()
            context['dispensed_count'] = all_requests.filter(status='dispensed').count()
        
        return context


@login_required
def consumable_request_create(request):
    """
    Full-page form for creating consumable requests with multiple line items.
    """
    from datetime import date
    
    if request.method == 'POST':
        form = ConsumableRequestForm(request.POST)
        items_formset = ConsumableRequestItemFormSet(request.POST)
        if form.is_valid() and items_formset.is_valid():
            consumable_request = form.save(commit=False)
            consumable_request.requested_by = request.user
            consumable_request.save()
            items_formset.instance = consumable_request
            items_formset.save()
            consumable_request.recalculate_total()
            # Save attachments
            for f in request.FILES.getlist('attachments'):
                ConsumableRequestAttachment.objects.create(
                    consumable_request=consumable_request,
                    file=f,
                    filename=f.name,
                    uploaded_by=request.user
                )
            messages.success(request, f'Request {consumable_request.request_number} submitted!')
            if consumable_request.project_id:
                return redirect('projects:project_detail', pk=consumable_request.project_id)
            return redirect('inventory:consumable_request_list')
    else:
        initial = {}
        raw_project = (request.GET.get('project') or '').strip()
        if raw_project.isdigit():
            from apps.projects.models import Project

            proj = Project.objects.filter(pk=int(raw_project), is_active=True).first()
            if proj:
                initial['project'] = proj.pk
        form = ConsumableRequestForm(initial=initial)
        items_formset = ConsumableRequestItemFormSet()
    
    return render(request, 'inventory/consumable_request_form.html', {
        'title': 'Request items',
        'form': form,
        'items_formset': items_formset,
        'today': date.today().isoformat(),
    })


def _consumable_request_redirect(consumable_request):
    if consumable_request.project_id:
        return redirect('projects:project_detail', pk=consumable_request.project_id)
    return redirect('inventory:consumable_request_detail', pk=consumable_request.pk)


@login_required
def consumable_request_edit(request, pk):
    """Edit a pending consumable request (requester only)."""
    consumable_request = get_object_or_404(
        ConsumableRequest.objects.prefetch_related('items'),
        pk=pk,
        is_active=True,
    )

    if consumable_request.requested_by != request.user:
        messages.error(request, 'Only the person who submitted this request can edit it.')
        return _consumable_request_redirect(consumable_request)

    if consumable_request.status != 'pending':
        messages.warning(request, 'Only pending requests can be edited.')
        return redirect('inventory:consumable_request_detail', pk=pk)

    from datetime import date

    if request.method == 'POST':
        form = ConsumableRequestForm(request.POST, instance=consumable_request)
        items_formset = ConsumableRequestItemFormSet(request.POST, instance=consumable_request)
        if form.is_valid() and items_formset.is_valid():
            form.save()
            items_formset.save()
            consumable_request.recalculate_total()
            for f in request.FILES.getlist('attachments'):
                ConsumableRequestAttachment.objects.create(
                    consumable_request=consumable_request,
                    file=f,
                    filename=f.name,
                    uploaded_by=request.user,
                )
            messages.success(request, f'Request {consumable_request.request_number} updated.')
            return _consumable_request_redirect(consumable_request)
    else:
        form = ConsumableRequestForm(instance=consumable_request)
        items_formset = ConsumableRequestItemFormSet(instance=consumable_request)

    return render(request, 'inventory/consumable_request_form.html', {
        'title': f'Edit request {consumable_request.request_number}',
        'form': form,
        'items_formset': items_formset,
        'request_obj': consumable_request,
        'is_edit': True,
        'today': date.today().isoformat(),
    })


@login_required
@require_POST
def consumable_request_delete(request, pk):
    """Soft-delete a pending consumable request (requester only)."""
    consumable_request = get_object_or_404(ConsumableRequest, pk=pk, is_active=True)

    if consumable_request.requested_by != request.user:
        messages.error(request, 'Only the person who submitted this request can delete it.')
        return _consumable_request_redirect(consumable_request)

    if consumable_request.status != 'pending':
        messages.warning(request, 'Only pending requests can be deleted.')
        return redirect('inventory:consumable_request_detail', pk=pk)

    request_number = consumable_request.request_number
    project_id = consumable_request.project_id
    consumable_request.is_active = False
    consumable_request.save(update_fields=['is_active', 'updated_at'])
    messages.success(request, f'Request {request_number} deleted.')

    if project_id:
        return redirect('projects:project_detail', pk=project_id)
    return redirect('inventory:consumable_request_list')


@login_required
def consumable_request_detail(request, pk):
    """View request details."""
    consumable_request = get_object_or_404(
        ConsumableRequest.objects.select_related(
            'item',
            'requested_by',
            'warehouse',
            'approved_by',
            'dispensed_by',
            'stock_movement',
            'department',
            'project',
        ).prefetch_related('items', 'items__item', 'attachments'),
        pk=pk
    )
    
    user = request.user
    is_admin = user.is_superuser or PermissionChecker.has_permission(user, 'inventory', 'edit')
    
    # Non-admins can only view their own requests
    if not is_admin and consumable_request.requested_by != user:
        messages.error(request, 'Permission denied.')
        return redirect('inventory:consumable_request_list')
    
    context = {
        'title': f'Request: {consumable_request.request_number}',
        'request_obj': consumable_request,
        'is_admin': is_admin,
        'can_edit_request': (
            consumable_request.status == 'pending'
            and consumable_request.requested_by == user
        ),
        'requires_project_delivery': consumable_request.uses_project_item_flow(),
    }
    
    # For admin: show approve/dispense forms
    if is_admin and consumable_request.status in ['pending', 'approved']:
        context['approve_form'] = ConsumableRequestApproveForm(
            consumable_request=consumable_request
        )
        context['reject_form'] = ConsumableRequestRejectForm()
    
    return render(request, 'inventory/consumable_request_detail.html', context)


@login_required
def consumable_request_approve(request, pk):
    """Admin approves a request."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('inventory:consumable_request_list')
    
    consumable_request = get_object_or_404(ConsumableRequest, pk=pk)
    
    if consumable_request.status != 'pending':
        messages.warning(request, f'Request {consumable_request.request_number} is not pending.')
        return redirect('inventory:consumable_request_detail', pk=pk)
    
    if request.method == 'POST':
        form = ConsumableRequestApproveForm(
            request.POST,
            consumable_request=consumable_request
        )
        if form.is_valid():
            try:
                warehouse = form.cleaned_data['warehouse']
                admin_notes = form.cleaned_data.get('admin_notes', '')
                consumable_request.admin_notes = admin_notes
                with transaction.atomic():
                    consumable_request.approve(request.user, warehouse)
                    n_lines = sync_consumable_request_to_project_item_lines(consumable_request)
                msg = f'Request {consumable_request.request_number} approved.'
                if n_lines:
                    msg += f' {n_lines} line(s) added to project items.'
                messages.success(request, msg)
            except Exception as e:
                messages.error(request, f'Error approving request: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    
    return redirect('inventory:consumable_request_detail', pk=pk)


@login_required
def consumable_request_dispense(request, pk):
    """Admin dispenses the consumable (reduces stock)."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('inventory:consumable_request_list')
    
    consumable_request = get_object_or_404(ConsumableRequest, pk=pk)
    
    if consumable_request.status not in ['pending', 'approved']:
        messages.warning(request, f'Request {consumable_request.request_number} cannot be dispensed.')
        return redirect('inventory:consumable_request_detail', pk=pk)
    
    if request.method == 'POST':
        form = ConsumableRequestApproveForm(
            request.POST,
            consumable_request=consumable_request
        )
        if form.is_valid():
            try:
                warehouse = form.cleaned_data['warehouse']
                with transaction.atomic():
                    consumable_request.dispense(request.user, warehouse)
                    n_lines = sync_consumable_request_to_project_item_lines(consumable_request)
                items_dispensed = consumable_request.get_items_for_dispense()
                msg = f'Request {consumable_request.request_number} dispensed.'
                if consumable_request.project_id:
                    msg += ' Items are on the project Items list — deliver or return stock from the project page.'
                elif items_dispensed:
                    msg += f' Stock reduced for {len(items_dispensed)} item(s).'
                if n_lines:
                    msg += f' {n_lines} line(s) added to project items.'
                messages.success(request, msg)
            except Exception as e:
                messages.error(request, f'Error dispensing: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    
    return redirect('inventory:consumable_request_detail', pk=pk)


@login_required
def consumable_request_reject(request, pk):
    """Admin rejects a request."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('inventory:consumable_request_list')
    
    consumable_request = get_object_or_404(ConsumableRequest, pk=pk)
    
    if consumable_request.status not in ['pending', 'approved']:
        messages.warning(request, f'Request {consumable_request.request_number} cannot be rejected.')
        return redirect('inventory:consumable_request_detail', pk=pk)
    
    if request.method == 'POST':
        form = ConsumableRequestRejectForm(request.POST)
        if form.is_valid():
            reason = form.cleaned_data['reason']
            consumable_request.reject(request.user, reason)
            messages.success(request, f'Request {consumable_request.request_number} rejected.')
        else:
            messages.error(request, 'Please provide a rejection reason.')
    
    return redirect('inventory:consumable_request_detail', pk=pk)


# ============ CONSUMABLE REPORTS ============

@login_required
def consumable_dashboard(request):
    """
    Dashboard for consumables showing:
    - Total requests this month
    - Total quantity consumed
    - Total cost
    - Low stock alerts
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    from django.utils import timezone
    from datetime import timedelta
    
    today = timezone.localdate()
    month_start = today.replace(day=1)
    
    # This month's requests
    month_requests = ConsumableRequest.objects.filter(
        is_active=True,
        request_date__gte=month_start
    )
    
    # Stats
    total_requests = month_requests.count()
    dispensed_requests = month_requests.filter(status='dispensed')
    total_quantity = dispensed_requests.aggregate(Sum('quantity'))['quantity__sum'] or Decimal('0')
    total_cost = dispensed_requests.aggregate(Sum('total_cost'))['total_cost__sum'] or Decimal('0')
    
    # Low stock consumables
    low_stock_items = []
    consumable_items = Item.objects.filter(
        is_active=True,
        item_type='product',
        status='active'
    )
    for item in consumable_items:
        total_stock = item.total_stock
        if total_stock < item.minimum_stock:
            low_stock_items.append({
                'item': item,
                'current_stock': total_stock,
                'minimum_stock': item.minimum_stock,
                'shortfall': item.minimum_stock - total_stock
            })
    
    # Recent requests
    recent_requests = ConsumableRequest.objects.filter(
        is_active=True
    ).select_related('item', 'requested_by').order_by('-created_at')[:10]
    
    # Top requested items this month
    top_items = dispensed_requests.values('item__name').annotate(
        total_qty=Sum('quantity'),
        total_cost=Sum('total_cost')
    ).order_by('-total_qty')[:5]
    
    context = {
        'title': 'Consumables Dashboard',
        'total_requests': total_requests,
        'pending_requests': month_requests.filter(status='pending').count(),
        'total_quantity': total_quantity,
        'total_cost': total_cost,
        'low_stock_items': low_stock_items,
        'low_stock_count': len(low_stock_items),
        'recent_requests': recent_requests,
        'top_items': top_items,
        'month_name': today.strftime('%B %Y'),
    }
    
    return render(request, 'inventory/consumable_dashboard.html', context)


@login_required
def consumable_monthly_request_report(request):
    """Monthly Request Report - per nurse & total."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    from django.utils import timezone
    from datetime import date
    
    # Get month from query params
    year = int(request.GET.get('year', timezone.localdate().year))
    month = int(request.GET.get('month', timezone.localdate().month))
    
    month_start = date(year, month, 1)
    if month == 12:
        month_end = date(year + 1, 1, 1)
    else:
        month_end = date(year, month + 1, 1)
    
    # Requests for the month
    requests = ConsumableRequest.objects.filter(
        is_active=True,
        request_date__gte=month_start,
        request_date__lt=month_end
    ).select_related('item', 'requested_by')
    
    # Group by nurse
    nurse_summary = requests.values(
        'requested_by__id',
        'requested_by__first_name',
        'requested_by__last_name',
        'requested_by__username'
    ).annotate(
        total_requests=Count('id'),
        total_quantity=Sum('quantity'),
        total_cost=Sum('total_cost'),
        pending=Count('id', filter=Q(status='pending')),
        approved=Count('id', filter=Q(status='approved')),
        dispensed=Count('id', filter=Q(status='dispensed')),
        rejected=Count('id', filter=Q(status='rejected')),
    ).order_by('-total_requests')
    
    # Totals
    totals = requests.aggregate(
        total_requests=Count('id'),
        total_quantity=Sum('quantity'),
        total_cost=Sum('total_cost'),
    )
    
    context = {
        'title': f'Monthly Request Report - {month_start.strftime("%B %Y")}',
        'nurse_summary': nurse_summary,
        'totals': totals,
        'year': year,
        'month': month,
        'month_name': month_start.strftime('%B %Y'),
        'years': range(2024, timezone.localdate().year + 2),
        'months': [(i, date(2000, i, 1).strftime('%B')) for i in range(1, 13)],
    }
    
    return render(request, 'inventory/consumable_monthly_request_report.html', context)


@login_required
def consumable_monthly_consumption_report(request):
    """Monthly Consumption Report - item-wise quantity used with analytics."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    from django.utils import timezone
    from datetime import date
    from dateutil.relativedelta import relativedelta
    
    # Get month from query params
    year = int(request.GET.get('year', timezone.localdate().year))
    month = int(request.GET.get('month', timezone.localdate().month))
    
    month_start = date(year, month, 1)
    if month == 12:
        month_end = date(year + 1, 1, 1)
    else:
        month_end = date(year, month + 1, 1)
    
    # Only dispensed requests; attribute to dispense date (not request_date) for trends
    period_qs = consumable_requests_in_consumption_period(month_start, month_end)
    consumption = period_qs.values(
        'item__id',
        'item__item_code',
        'item__name',
        'item__unit'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_cost=Sum('total_cost'),
        request_count=Count('id')
    ).order_by('-total_quantity')
    
    # Totals
    totals = period_qs.aggregate(
        total_quantity=Sum('quantity'),
        total_cost=Sum('total_cost'),
        total_requests=Count('id'),
    )
    
    # ===== CHART DATA =====
    
    def _consumption_item_label(row, max_len):
        """Legacy FK `item` may be null on multi-line requests; ORM can return null names."""
        name = row.get('item__name')
        code = row.get('item__item_code')
        label = (name or '').strip() or (code or '').strip() or 'Unknown'
        return label[:max_len]
    
    # 1. Top 10 Most Consumed Items (for forecast)
    top_items = list(consumption[:10])
    top_items_labels = [_consumption_item_label(c, 20) for c in top_items]
    top_items_data = [float(c['total_quantity'] or 0) for c in top_items]
    
    # 2. Consumption by User (who orders more/less)
    user_consumption = period_qs.values(
        'requested_by__username',
        'requested_by__first_name',
        'requested_by__last_name'
    ).annotate(
        total_requests=Count('id'),
        total_quantity=Sum('quantity'),
        total_cost=Sum('total_cost')
    ).order_by('-total_requests')[:10]
    
    user_labels = [f"{u['requested_by__first_name'] or ''} {u['requested_by__last_name'] or ''}".strip() or u['requested_by__username'] for u in user_consumption]
    user_data = [u['total_requests'] for u in user_consumption]
    
    # 3. Monthly Cost Trend (last 6 months)
    monthly_costs = []
    monthly_labels = []
    for i in range(5, -1, -1):
        m_date = month_start - relativedelta(months=i)
        m_end = m_date + relativedelta(months=1)
        cost = consumable_requests_in_consumption_period(m_date, m_end).aggregate(
            total=Sum('total_cost')
        )['total'] or 0
        monthly_costs.append(float(cost))
        monthly_labels.append(m_date.strftime('%b %Y'))
    
    # 4. Items needing refill (high consumption vs current stock)
    from apps.inventory.models import Stock
    refill_items = []
    for item_data in consumption[:20]:
        item_id = item_data['item__id']
        if item_id is None:
            continue
        monthly_consumption = float(item_data['total_quantity'] or 0)
        current_stock = Stock.objects.filter(item_id=item_id).aggregate(total=Sum('quantity'))['total'] or 0
        # If current stock < 2 months of consumption, flag for refill
        if current_stock < (monthly_consumption * 2):
            refill_items.append({
                'name': _consumption_item_label(item_data, 200),
                'monthly_consumption': monthly_consumption,
                'current_stock': float(current_stock),
                'months_left': round(float(current_stock) / monthly_consumption, 1) if monthly_consumption > 0 else 0
            })
    
    # 5. Inactive/Rarely Used Items (items with no consumption in last 3 months)
    three_months_ago = month_start - relativedelta(months=3)
    consumed_item_ids = _consumable_with_consumption_date().filter(
        _consumption_date__gte=three_months_ago
    ).values_list('item_id', flat=True).distinct()
    
    inactive_items = Item.objects.filter(
        is_active=True,
        item_type='product',
        category__name__icontains='medical'
    ).exclude(id__in=consumed_item_ids).values('name', 'item_code')[:10]
    
    # 6. Cost breakdown by item (pie chart)
    cost_breakdown = list(consumption[:8])
    cost_labels = [_consumption_item_label(c, 15) for c in cost_breakdown]
    cost_data = [float(c['total_cost'] or 0) for c in cost_breakdown]
    
    context = {
        'title': f'Consumption Analytics - {month_start.strftime("%B %Y")}',
        'consumption': consumption,
        'totals': totals,
        'year': year,
        'month': month,
        'month_name': month_start.strftime('%B %Y'),
        'years': range(2024, timezone.localdate().year + 2),
        'months': [(i, date(2000, i, 1).strftime('%B')) for i in range(1, 13)],
        # Chart data (JSON for valid Chart.js / JS parsing)
        'monthly_labels_json': mark_safe(json.dumps(monthly_labels)),
        'monthly_costs_json': mark_safe(json.dumps(monthly_costs)),
        'cost_labels_json': mark_safe(json.dumps(cost_labels)),
        'cost_data_json': mark_safe(json.dumps(cost_data)),
        'top_items_labels_json': mark_safe(json.dumps(top_items_labels)),
        'top_items_data_json': mark_safe(json.dumps(top_items_data)),
        'user_labels_json': mark_safe(json.dumps(user_labels)),
        'user_data_json': mark_safe(json.dumps(user_data)),
        'refill_items': refill_items,
        'inactive_items': inactive_items,
        'user_consumption': user_consumption,
    }
    
    return render(request, 'inventory/consumable_monthly_consumption_report.html', context)


@login_required
def consumable_monthly_cost_report(request):
    """Monthly Financial Cost Report - total consumable cost."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    from django.utils import timezone
    from datetime import date
    
    # Get month from query params
    year = int(request.GET.get('year', timezone.localdate().year))
    month = int(request.GET.get('month', timezone.localdate().month))
    
    month_start = date(year, month, 1)
    if month == 12:
        month_end = date(year + 1, 1, 1)
    else:
        month_end = date(year, month + 1, 1)
    
    period_qs = consumable_requests_in_consumption_period(month_start, month_end)
    # Cost breakdown by item
    cost_breakdown = period_qs.values(
        'item__id',
        'item__item_code',
        'item__name',
        'item__category__name'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_cost=Sum('total_cost'),
        avg_unit_cost=Avg('unit_cost')
    ).order_by('-total_cost')
    
    # Daily cost trend (by consumption date; template-friendly keys)
    _daily_rows = period_qs.values('_consumption_date').annotate(
        daily_cost=Sum('total_cost'),
        daily_qty=Sum('quantity')
    ).order_by('_consumption_date')
    daily_costs = [
        {
            'consumption_day': r['_consumption_date'],
            'daily_cost': r['daily_cost'],
            'daily_qty': r['daily_qty'],
        }
        for r in _daily_rows
    ]
    
    # Totals
    totals = period_qs.aggregate(
        total_cost=Sum('total_cost'),
        total_quantity=Sum('quantity'),
        total_requests=Count('id'),
    )
    
    context = {
        'title': f'Monthly Cost Report - {month_start.strftime("%B %Y")}',
        'cost_breakdown': cost_breakdown,
        'daily_costs': list(daily_costs),
        'totals': totals,
        'year': year,
        'month': month,
        'month_name': month_start.strftime('%B %Y'),
        'years': range(2024, timezone.localdate().year + 2),
        'months': [(i, date(2000, i, 1).strftime('%B')) for i in range(1, 13)],
    }
    
    return render(request, 'inventory/consumable_monthly_cost_report.html', context)


def _consumable_report_perm(request):
    return request.user.is_superuser or PermissionChecker.has_permission(request.user, 'inventory', 'view')


def _parse_iso_date(s, default=None):
    from datetime import datetime as dt
    if not s:
        return default
    try:
        return dt.strptime(s.strip(), '%Y-%m-%d').date()
    except ValueError:
        return default


def _json_safe_floats(obj):
    """Replace NaN/inf floats so JsonResponse never fails."""
    import math

    if isinstance(obj, dict):
        return {k: _json_safe_floats(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_json_safe_floats(v) for v in obj]
    if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
        return None
    return obj


@login_required
def consumable_inventory_reports_page(request):
    if not _consumable_report_perm(request):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    return render(
        request,
        'inventory/consumable_inventory_reports.html',
        {
            'title': 'Consumables — Inventory Reports',
            'report_keys': list(REPORT_BUILDERS.keys()),
        },
    )


@login_required
@require_http_methods(['GET'])
def consumable_inventory_report_api(request):
    if not _consumable_report_perm(request):
        return JsonResponse({'ok': False, 'error': 'Permission denied'}, status=403)
    report_key = request.GET.get('report', 'stock_summary')
    if report_key not in REPORT_BUILDERS:
        return JsonResponse({'ok': False, 'error': 'Unknown report'}, status=400)
    date_from = _parse_iso_date(request.GET.get('date_from'))
    date_to = _parse_iso_date(request.GET.get('date_to'))
    if not date_from or not date_to:
        return JsonResponse({'ok': False, 'error': 'date_from and date_to are required (YYYY-MM-DD)'}, status=400)
    if date_from > date_to:
        return JsonResponse({'ok': False, 'error': 'From date cannot be after To date'}, status=400)
    warranty_filter = request.GET.get('warranty_filter', 'all')
    movement_group = request.GET.get('movement_group', 'date')
    try:
        payload = build_report(report_key, date_from, date_to, warranty_filter=warranty_filter, movement_group=movement_group)
    except Exception as exc:
        return JsonResponse({'ok': False, 'error': str(exc)}, status=500)
    # Strip internal id from row payloads for cleaner API
    for row in payload.get('rows', []):
        row.pop('item_id', None)
    payload['ok'] = True
    payload['date_from'] = date_from.isoformat()
    payload['date_to'] = date_to.isoformat()
    payload['generated_by'] = request.user.get_full_name() or request.user.username
    return JsonResponse(_json_safe_floats(payload))


@login_required
@require_http_methods(['GET'])
def consumable_inventory_report_export_pdf(request):
    if not _consumable_report_perm(request):
        return HttpResponseForbidden('Permission denied')
    report_key = request.GET.get('report', 'stock_summary')
    if report_key not in REPORT_BUILDERS:
        return HttpResponseForbidden('Unknown report')
    date_from = _parse_iso_date(request.GET.get('date_from'))
    date_to = _parse_iso_date(request.GET.get('date_to'))
    if not date_from or not date_to or date_from > date_to:
        return HttpResponseForbidden('Invalid date range')
    warranty_filter = request.GET.get('warranty_filter', 'all')
    movement_group = request.GET.get('movement_group', 'date')
    payload = build_report(report_key, date_from, date_to, warranty_filter=warranty_filter, movement_group=movement_group)
    for row in payload.get('rows', []):
        row.pop('item_id', None)
    pdf_bytes = export_report_pdf(
        payload,
        date_from,
        date_to,
        request.user.get_full_name() or request.user.username,
    )
    resp = HttpResponse(pdf_bytes, content_type='application/pdf')
    safe_name = ''.join(c if c.isalnum() or c in '-_' else '_' for c in report_key)
    resp['Content-Disposition'] = f'attachment; filename="{safe_name}_{date_from}_{date_to}.pdf"'
    return resp


@login_required
@require_http_methods(['GET'])
def consumable_inventory_report_export_xlsx(request):
    if not _consumable_report_perm(request):
        return HttpResponseForbidden('Permission denied')
    report_key = request.GET.get('report', 'stock_summary')
    if report_key not in REPORT_BUILDERS:
        return HttpResponseForbidden('Unknown report')
    date_from = _parse_iso_date(request.GET.get('date_from'))
    date_to = _parse_iso_date(request.GET.get('date_to'))
    if not date_from or not date_to or date_from > date_to:
        return HttpResponseForbidden('Invalid date range')
    warranty_filter = request.GET.get('warranty_filter', 'all')
    movement_group = request.GET.get('movement_group', 'date')
    payload = build_report(report_key, date_from, date_to, warranty_filter=warranty_filter, movement_group=movement_group)
    for row in payload.get('rows', []):
        row.pop('item_id', None)
    xlsx_bytes = export_report_xlsx(
        payload,
        date_from,
        date_to,
        request.user.get_full_name() or request.user.username,
    )
    resp = HttpResponse(
        xlsx_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    safe_name = ''.join(c if c.isalnum() or c in '-_' else '_' for c in report_key)
    resp['Content-Disposition'] = f'attachment; filename="{safe_name}_{date_from}_{date_to}.xlsx"'
    return resp


@login_required
@require_http_methods(['POST'])
def storage_location_create(request):
    if not (
        request.user.is_superuser
        or PermissionChecker.has_permission(request.user, 'inventory', 'create')
        or PermissionChecker.has_permission(request.user, 'inventory', 'edit')
    ):
        return JsonResponse({'ok': False, 'error': 'Permission denied'}, status=403)
    try:
        body = json.loads(request.body.decode() or '{}')
    except json.JSONDecodeError:
        body = {}
    name = (body.get('name') or '').strip()
    description = (body.get('description') or '').strip()
    if not name:
        return JsonResponse({'ok': False, 'error': 'Name is required'}, status=400)
    if StorageLocation.objects.filter(name__iexact=name).exists():
        return JsonResponse({'ok': False, 'error': 'A location with this name already exists'}, status=400)
    loc = StorageLocation.objects.create(name=name, description=description)
    return JsonResponse({'ok': True, 'id': loc.id, 'name': loc.name})

