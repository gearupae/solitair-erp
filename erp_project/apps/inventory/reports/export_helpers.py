"""Generic table export for inventory reports (PDF + Excel)."""
from __future__ import annotations

from datetime import date
from io import BytesIO

from django.utils.html import escape
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def export_table_xlsx(report_payload: dict, generated_by: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (report_payload.get('title') or 'Report')[:31]

    H_FILL = PatternFill('solid', fgColor='1F2937')
    H_FONT = Font(bold=True, color='FFFFFF', size=10)
    BORDER = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    FILL_MAP = {
        'High': PatternFill('solid', fgColor='FEE2E2'),
        'Medium': PatternFill('solid', fgColor='FFEDD5'),
        'Low': PatternFill('solid', fgColor='D1FAE5'),
        'Fast Mover': PatternFill('solid', fgColor='DBEAFE'),
        'Slow': PatternFill('solid', fgColor='FEF9C3'),
        'Dead': PatternFill('solid', fgColor='E5E7EB'),
        'Overstocked': PatternFill('solid', fgColor='F3E8FF'),
        'critical': PatternFill('solid', fgColor='FEE2E2'),
        'warning': PatternFill('solid', fgColor='FEF3C7'),
        'good': PatternFill('solid', fgColor='D1FAE5'),
    }

    ws['A1'] = report_payload.get('title', 'Report')
    ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = f'Generated: {date.today().isoformat()} by {generated_by}'
    ws['A2'].font = Font(size=10, italic=True, color='6B7280')

    cols = report_payload.get('columns', [])
    header_row = 4
    col_keys = [c['key'] for c in cols]
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=header_row, column=ci, value=col['label'])
        cell.fill = H_FILL
        cell.font = H_FONT
        cell.border = BORDER

    is_ai = report_payload.get('report_type') == 'ai_forecast'

    for ri, row in enumerate(report_payload.get('rows', []), header_row + 1):
        for ci, col in enumerate(cols, 1):
            key = col['key']
            val = row.get(key, '')
            if key == 'trend_label' and is_ai:
                icon = row.get('trend_icon', '')
                val = f'{icon} {val}'.strip()
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            if col.get('format') == 'number' and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            if is_ai:
                if key == 'stockout_risk' and val in FILL_MAP:
                    cell.fill = FILL_MAP[val]
                elif key == 'status' and val in FILL_MAP:
                    cell.fill = FILL_MAP[val]
                elif key == 'days_left_display':
                    dl_class = row.get('days_left_class', '')
                    if 'critical' in dl_class:
                        cell.fill = FILL_MAP['critical']
                    elif 'warning' in dl_class:
                        cell.fill = FILL_MAP['warning']
                    elif 'good' in dl_class:
                        cell.fill = FILL_MAP['good']

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_table_pdf(report_payload: dict, generated_by: str) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=12 * mm, rightMargin=12 * mm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"<b>{escape(report_payload.get('title', 'Report'))}</b>", styles['Title']),
        Spacer(1, 6),
        Paragraph(
            f"Generated {date.today().strftime('%d/%m/%Y')} — {escape(generated_by)}",
            styles['Normal'],
        ),
        Spacer(1, 12),
    ]

    cols = report_payload.get('columns', [])
    headers = [c['label'] for c in cols]
    data = [headers]
    for row in report_payload.get('rows', []):
        cells = []
        for c in cols:
            val = row.get(c['key'], '')
            if c['key'] == 'trend_label' and report_payload.get('report_type') == 'ai_forecast':
                val = f"{row.get('trend_icon', '')} {val}".strip()
            cells.append(str(val if val is not None else ''))
        data.append(cells)

    if len(data) == 1:
        data.append(['No data'] + [''] * (len(headers) - 1))

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    return buf.getvalue()
