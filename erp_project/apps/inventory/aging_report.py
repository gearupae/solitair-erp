"""
Inventory Aging Report — data builder + export helpers.

Age = Today − date of last GRN receipt (StockMovement.movement_type='in').
      Falls back to last movement date, then item created_at.

Buckets (UAE warehouse standard):
  0–30    days → Fresh         (success / green)
  31–60   days → Monitor       (warning / yellow)
  61–90   days → Slow Moving   (orange)
  91–180  days → Critical      (danger / red)
  180+    days → Dead Stock    (dark / near-black)

NOTE: This is a standalone file — it does NOT modify or import from
any existing inventory views or consumable_inventory_reports.py.
Only models are imported.
"""
from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Any

from django.db.models import Max, Sum
from django.utils import timezone


# ---------------------------------------------------------------------------
# Bucket helpers
# ---------------------------------------------------------------------------

BUCKET_DEFS = [
    ("0-30",   "0–30 Days",   "Fresh",        "success", 0,   30),
    ("31-60",  "31–60 Days",  "Monitor",      "warning", 31,  60),
    ("61-90",  "61–90 Days",  "Slow Moving",  "orange",  61,  90),
    ("91-180", "91–180 Days", "Critical",     "danger",  91,  180),
    ("180+",   "180+ Days",   "Dead Stock",   "dark",    181, 99999),
]

BUCKET_ORDER = [d[0] for d in BUCKET_DEFS]


def _bucket(age_days: int):
    for key, days_label, status_label, color, lo, hi in BUCKET_DEFS:
        if lo <= age_days <= hi:
            return key, f"{days_label} ({status_label})", color
    return "180+", "180+ Days (Dead Stock)", "dark"


# ---------------------------------------------------------------------------
# Cost helper (WAC / purchase_price)
# ---------------------------------------------------------------------------

def _unit_cost(item) -> Decimal:
    """Weighted average cost from receipts; falls back to purchase_price."""
    from apps.inventory.models import StockMovement
    if item.purchase_price and item.purchase_price > 0:
        return item.purchase_price.quantize(Decimal("0.01"))
    agg = StockMovement.objects.filter(
        item=item, movement_type="in", quantity__gt=0
    ).aggregate(t=Sum("total_cost"), q=Sum("quantity"))
    t, q = agg["t"], agg["q"]
    if t and q and q > 0:
        return (t / q).quantize(Decimal("0.01"))
    return Decimal("0.00")


# ---------------------------------------------------------------------------
# Core data builder
# ---------------------------------------------------------------------------

def build_aging_report(
    as_of_date: date | None = None,
    category_id: int | None = None,
    warehouse_id: int | None = None,
    bucket_filter: str | None = None,
    slow_moving_only: bool = False,
    search: str | None = None,
) -> dict[str, Any]:
    """
    Returns:
        {
          'rows': [...],
          'summary': {...},
          'as_of_date': date,
        }

    Each row dict contains all display-ready fields.
    Values are kept as Decimal for precision; templates call |floatformat.
    """
    from apps.inventory.models import Item, Stock, StockMovement

    today: date = as_of_date or timezone.localdate()

    # -----------------------------------------------------------------------
    # Base items
    # -----------------------------------------------------------------------
    items_qs = (
        Item.objects.filter(is_active=True, item_type="product", status="active")
        .select_related("category", "storage_location_master")
        .order_by("name")
    )

    if category_id:
        items_qs = items_qs.filter(category_id=category_id)

    if search:
        from django.db.models import Q
        items_qs = items_qs.filter(
            Q(name__icontains=search) | Q(item_code__icontains=search)
        )

    # -----------------------------------------------------------------------
    # Pre-aggregate movement data (single DB round-trips)
    # -----------------------------------------------------------------------
    # Last receipt date per item
    last_receipt_map: dict[int, date] = {
        r["item_id"]: r["last_rx"]
        for r in StockMovement.objects.filter(
            movement_type="in",
            movement_date__lte=today,
            item__is_active=True,
        )
        .values("item_id")
        .annotate(last_rx=Max("movement_date"))
    }

    # Last ANY movement date per item
    last_move_map: dict[int, date] = {
        r["item_id"]: r["last_mv"]
        for r in StockMovement.objects.filter(
            movement_date__lte=today,
            item__is_active=True,
        )
        .values("item_id")
        .annotate(last_mv=Max("movement_date"))
    }

    # Last GRN movement id per item (for audit reference number)
    last_grn_id_map: dict[int, int] = {
        r["item_id"]: r["last_id"]
        for r in StockMovement.objects.filter(
            movement_type="in",
            movement_date__lte=today,
            item__is_active=True,
        )
        .values("item_id")
        .annotate(last_id=Max("id"))
    }

    # Fetch actual GRN movements for reference field
    grn_ref_map: dict[int, str] = {}
    if last_grn_id_map:
        for m in StockMovement.objects.filter(id__in=last_grn_id_map.values()):
            ref = m.reference or m.movement_number or ""
            grn_ref_map[m.item_id] = ref

    # -----------------------------------------------------------------------
    # Stock quantities from Stock model (balances)
    # -----------------------------------------------------------------------
    stock_qs = Stock.objects.filter(
        item__is_active=True, quantity__gt=0
    ).select_related("warehouse")
    if warehouse_id:
        stock_qs = stock_qs.filter(warehouse_id=warehouse_id)

    stock_qty_map: dict[int, Decimal] = {}
    warehouse_name_map: dict[int, str] = {}
    warehouse_max_qty: dict[int, Decimal] = {}

    for s in stock_qs:
        iid = s.item_id
        stock_qty_map[iid] = stock_qty_map.get(iid, Decimal("0")) + s.quantity
        # Keep the warehouse name where most stock is held (for display)
        if s.quantity > warehouse_max_qty.get(iid, Decimal("-1")):
            warehouse_max_qty[iid] = s.quantity
            warehouse_name_map[iid] = s.warehouse.name

    # -----------------------------------------------------------------------
    # Build rows
    # -----------------------------------------------------------------------
    rows: list[dict] = []

    for item in items_qs:
        qty = stock_qty_map.get(item.id, Decimal("0"))
        if qty <= 0:
            continue  # no stock — skip

        last_receipt = last_receipt_map.get(item.id)
        last_movement = last_move_map.get(item.id)

        # Aging start: last receipt > last movement > item created_at
        if last_receipt:
            aging_start = last_receipt
        elif last_movement:
            aging_start = last_movement
        else:
            aging_start = item.created_at.date() if item.created_at else today

        age_days = max(0, (today - aging_start).days)
        bucket_key, bucket_label, bucket_color = _bucket(age_days)
        slow_moving = age_days >= 90
        dead_stock = age_days >= 180

        # Apply bucket / slow-moving filters
        if bucket_filter and bucket_filter != bucket_key:
            continue
        if slow_moving_only and not slow_moving:
            continue

        unit_cost = _unit_cost(item)
        total_value = (qty * unit_cost).quantize(Decimal("0.01"))

        rows.append(
            {
                "item_id": item.id,
                "item_code": item.item_code or "",
                "item_name": item.name,
                "category": item.category.name if item.category_id else "—",
                "warehouse": warehouse_name_map.get(item.id, "—"),
                "unit": item.unit,
                "qty_on_hand": qty,
                "unit_cost": unit_cost,
                "total_value": total_value,
                "last_receipt_date": last_receipt,
                "last_movement_date": last_movement,
                "age_days": age_days,
                "bucket_key": bucket_key,
                "bucket_label": bucket_label,
                "bucket_color": bucket_color,
                "slow_moving": slow_moving,
                "dead_stock": dead_stock,
                "grn_ref": grn_ref_map.get(item.id, ""),
            }
        )

    # Default sort: oldest first
    rows.sort(key=lambda r: r["age_days"], reverse=True)

    # -----------------------------------------------------------------------
    # Summary
    # -----------------------------------------------------------------------
    grand_value: Decimal = sum((r["total_value"] for r in rows), Decimal("0"))

    bucket_summary: dict[str, dict] = {
        d[0]: {
            "key": d[0],
            "label": f"{d[1]} ({d[2]})",
            "color": d[3],
            "count": 0,
            "value": Decimal("0"),
        }
        for d in BUCKET_DEFS
    }
    for r in rows:
        bk = r["bucket_key"]
        if bk in bucket_summary:
            bucket_summary[bk]["count"] += 1
            bucket_summary[bk]["value"] += r["total_value"]

    slow_count = sum(1 for r in rows if r["slow_moving"])
    dead_count = sum(1 for r in rows if r["dead_stock"])
    slow_value: Decimal = sum(
        (r["total_value"] for r in rows if r["slow_moving"]), Decimal("0")
    )
    pct_slow = (
        (slow_value / grand_value * 100).quantize(Decimal("0.1"))
        if grand_value
        else Decimal("0.0")
    )

    return {
        "rows": rows,
        "summary": {
            "total_items": len(rows),
            "total_value": grand_value,
            "slow_moving_count": slow_count,
            "dead_stock_count": dead_count,
            "pct_90_plus": pct_slow,
            "buckets": list(bucket_summary.values()),
        },
        "as_of_date": today,
    }


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_aging_xlsx(data: dict, generated_by: str = "") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from apps.settings_app.models import CompanySettings

    cs = CompanySettings.get_settings()
    company = cs.company_name or "Al Najah ERP"

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Aging"

    # ------ Colour palette ------
    H_FILL = PatternFill("solid", fgColor="1F2937")   # dark header
    H_FONT = Font(bold=True, color="FFFFFF", size=10)
    TITLE_FONT = Font(bold=True, size=13)
    BORDER = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    BUCKET_FILLS = {
        "success": PatternFill("solid", fgColor="D1FAE5"),
        "warning": PatternFill("solid", fgColor="FEF3C7"),
        "orange":  PatternFill("solid", fgColor="FFEDD5"),
        "danger":  PatternFill("solid", fgColor="FEE2E2"),
        "dark":    PatternFill("solid", fgColor="374151"),
    }
    BUCKET_FONTS = {
        "success": Font(color="065F46", size=9, bold=True),
        "warning": Font(color="92400E", size=9, bold=True),
        "orange":  Font(color="C2410C", size=9, bold=True),
        "danger":  Font(color="B91C1C", size=9, bold=True),
        "dark":    Font(color="FFFFFF", size=9, bold=True),
    }

    # ------ Title block ------
    ws.merge_cells("A1:N1")
    ws["A1"] = f"Inventory Aging Report — {company}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A2:N2")
    ws["A2"] = f"As of: {data['as_of_date'].strftime('%d/%m/%Y')}    Generated by: {generated_by}"
    ws["A2"].font = Font(size=10, italic=True, color="6B7280")

    # ------ Summary block ------
    s = data["summary"]
    ws["A4"] = "Summary"
    ws["A4"].font = Font(bold=True, size=10)
    ws["A5"] = "Total Items"
    ws["B5"] = s["total_items"]
    ws["A6"] = "Total Stock Value (AED)"
    ws["B6"] = float(s["total_value"])
    ws["B6"].number_format = '#,##0.00'
    ws["A7"] = "Slow Moving (90+ days)"
    ws["B7"] = s["slow_moving_count"]
    ws["A8"] = "Dead Stock (180+ days)"
    ws["B8"] = s["dead_stock_count"]
    ws["A9"] = "% Value aged 90+ days"
    ws["B9"] = float(s["pct_90_plus"])
    ws["B9"].number_format = '0.0"%"'

    # ------ Headers ------
    HEADERS = [
        "Item Code", "Item Name", "Category", "Warehouse", "UOM",
        "Qty on Hand", "Unit Cost (AED)", "Total Value (AED)",
        "Last Receipt Date", "Last Movement Date",
        "Age (Days)", "Aging Bucket", "Slow Moving", "GRN Reference",
    ]
    header_row = 11
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = H_FILL
        cell.font = H_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER

    # ------ Data rows ------
    for ri, r in enumerate(data["rows"], header_row + 1):
        values = [
            r["item_code"],
            r["item_name"],
            r["category"],
            r["warehouse"],
            r["unit"],
            float(r["qty_on_hand"]),
            float(r["unit_cost"]),
            float(r["total_value"]),
            r["last_receipt_date"].strftime("%d/%m/%Y") if r["last_receipt_date"] else "—",
            r["last_movement_date"].strftime("%d/%m/%Y") if r["last_movement_date"] else "—",
            r["age_days"],
            r["bucket_label"],
            "Yes" if r["slow_moving"] else "No",
            r["grn_ref"],
        ]
        bcolor = r["bucket_color"]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.border = BORDER
            c.alignment = Alignment(horizontal="left", vertical="center")
            if col in (6, 7):
                c.number_format = "#,##0.00"
            if col == 8:
                c.number_format = "#,##0.00"
            if col == 11:
                c.alignment = Alignment(horizontal="center")
            if col == 12:
                c.fill = BUCKET_FILLS.get(bcolor, BUCKET_FILLS["success"])
                c.font = BUCKET_FONTS.get(bcolor, BUCKET_FONTS["success"])
                c.alignment = Alignment(horizontal="center")

    # ------ Column widths ------
    COL_WIDTHS = [14, 32, 20, 18, 6, 12, 16, 18, 18, 18, 10, 26, 13, 18]
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[header_row].height = 28
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF export (ReportLab)
# ---------------------------------------------------------------------------

def export_aging_pdf(data: dict, generated_by: str = "") -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )
    from apps.settings_app.models import CompanySettings

    cs = CompanySettings.get_settings()
    company = cs.company_name or "Al Najah ERP"
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=14 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    story = []

    # Title
    story.append(
        Paragraph(
            f"<b>{company} — Inventory Aging Report</b>",
            styles["Title"],
        )
    )
    story.append(
        Paragraph(
            f"As of: {data['as_of_date'].strftime('%d/%m/%Y')} &nbsp;|&nbsp; Generated by: {generated_by}",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 6 * mm))

    # Summary table
    s = data["summary"]
    summary_data = [
        ["Total Items", str(s["total_items"]),
         "Slow Moving (90+d)", str(s["slow_moving_count"]),
         "Dead Stock (180+d)", str(s["dead_stock_count"])],
        ["Total Value (AED)", f"{float(s['total_value']):,.2f}",
         "% Value 90+ days", f"{s['pct_90_plus']}%", "", ""],
    ]
    st = Table(summary_data, colWidths=[45 * mm, 35 * mm, 45 * mm, 25 * mm, 45 * mm, 30 * mm])
    st.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("FONTNAME", (4, 0), (4, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F9FAFB"), colors.white]),
        ])
    )
    story.append(st)
    story.append(Spacer(1, 6 * mm))

    # Data table
    BUCKET_COLORS = {
        "success": colors.HexColor("#D1FAE5"),
        "warning": colors.HexColor("#FEF3C7"),
        "orange":  colors.HexColor("#FFEDD5"),
        "danger":  colors.HexColor("#FEE2E2"),
        "dark":    colors.HexColor("#374151"),
    }

    headers = [
        "Item Code", "Item Name", "Category", "Warehouse",
        "UOM", "Qty", "Unit Cost", "Total Value",
        "Last Receipt", "Last Move", "Age (d)", "Bucket",
    ]
    col_widths = [
        18 * mm, 46 * mm, 28 * mm, 24 * mm,
        10 * mm, 14 * mm, 20 * mm, 22 * mm,
        22 * mm, 22 * mm, 14 * mm, 32 * mm,
    ]

    table_data = [headers]
    bucket_color_per_row: list = []

    for r in data["rows"]:
        table_data.append([
            r["item_code"],
            r["item_name"][:40],
            r["category"][:20],
            r["warehouse"][:16],
            r["unit"],
            f"{float(r['qty_on_hand']):,.2f}",
            f"{float(r['unit_cost']):,.2f}",
            f"{float(r['total_value']):,.2f}",
            r["last_receipt_date"].strftime("%d/%m/%Y") if r["last_receipt_date"] else "—",
            r["last_movement_date"].strftime("%d/%m/%Y") if r["last_movement_date"] else "—",
            str(r["age_days"]),
            r["bucket_label"][:22],
        ])
        bucket_color_per_row.append(BUCKET_COLORS.get(r["bucket_color"], colors.white))

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F9FAFB"), colors.white]),
    ]
    for ri, bc in enumerate(bucket_color_per_row, 1):
        style_cmds.append(("BACKGROUND", (11, ri), (11, ri), bc))

    dt = Table(table_data, colWidths=col_widths, repeatRows=1)
    dt.setStyle(TableStyle(style_cmds))
    story.append(dt)

    doc.build(story)
    return buf.getvalue()
