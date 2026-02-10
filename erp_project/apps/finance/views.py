"""
Finance Views - UAE VAT & Corporate Tax Compliant
Chart of Accounts, Journal Entries, Payments, Reports
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.contrib import messages
from django.views.generic import ListView, CreateView, UpdateView, DetailView, TemplateView
from django.urls import reverse_lazy
from django.db.models import Q, Sum, F
from django.db.models.functions import Coalesce
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from datetime import date, timedelta, datetime
from decimal import Decimal

User = get_user_model()

from .models import (
    Account, AccountType, AccountCategory, FiscalYear, AccountingPeriod, JournalEntry, JournalEntryLine, 
    TaxCode, Payment, BankAccount, ExpenseClaim, ExpenseItem, VATReturn, CorporateTaxComputation,
    Budget, BudgetLine, BankTransfer, BankReconciliation, BankStatement, BankStatementLine,
    ReconciliationItem, OpeningBalanceEntry, OpeningBalanceLine, WriteOff, ExchangeRate,
    AccountMapping, AccountingSettings
)
from .forms import (
    AccountForm, FiscalYearForm, AccountingPeriodForm, JournalEntryForm, JournalEntryLineFormSet, 
    PaymentForm, TaxCodeForm, BankAccountForm, ExpenseClaimForm, ExpenseItemFormSet,
    VATReturnForm, CorporateTaxForm, BudgetForm, BudgetLineFormSet, BankTransferForm,
    BankReconciliationForm, BankStatementForm, BankStatementLineFormSet,
    OpeningBalanceEntryForm, OpeningBalanceLineFormSet, WriteOffForm, ExchangeRateForm
)
from django import forms
from apps.core.mixins import PermissionRequiredMixin, CreatePermissionMixin, UpdatePermissionMixin
from apps.core.utils import PermissionChecker


# ============ CHART OF ACCOUNTS VIEWS ============

class AccountListView(PermissionRequiredMixin, ListView):
    model = Account
    template_name = 'finance/account_list.html'
    context_object_name = 'accounts'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        queryset = Account.objects.filter(is_active=True)
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) | Q(name__icontains=search)
            )
        
        account_type = self.request.GET.get('type')
        if account_type:
            queryset = queryset.filter(account_type=account_type)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Chart of Accounts'
        context['account_types'] = AccountType.choices
        context['form'] = AccountForm()
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        context['can_delete'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'delete')
        
        # Flag abnormal balances
        for account in context['accounts']:
            account.abnormal = account.has_abnormal_balance
        
        return context
    
    def post(self, request, *args, **kwargs):
        if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'create')):
            messages.error(request, 'Permission denied.')
            return redirect('finance:account_list')
        
        form = AccountForm(request.POST)
        if form.is_valid():
            account = form.save()
            messages.success(request, f'Account {account.code} created.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
        return redirect('finance:account_list')


class AccountUpdateView(UpdatePermissionMixin, UpdateView):
    model = Account
    form_class = AccountForm
    template_name = 'finance/account_form.html'
    success_url = reverse_lazy('finance:account_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Account: {self.object.code}'
        return context
    
    def form_valid(self, form):
        messages.success(self.request, f'Account {form.instance.code} updated.')
        return super().form_valid(form)


@login_required
def account_delete(request, pk):
    """Soft delete - system accounts cannot be deleted."""
    account = get_object_or_404(Account, pk=pk)
    if account.is_system:
        messages.error(request, 'System accounts cannot be deleted.')
        return redirect('finance:account_list')
    
    if request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'delete'):
        account.is_active = False
        account.save()
        messages.success(request, f'Account {account.code} deleted.')
    else:
        messages.error(request, 'Permission denied.')
    return redirect('finance:account_list')


# ============ JOURNAL ENTRY VIEWS ============

class JournalEntryListView(PermissionRequiredMixin, ListView):
    model = JournalEntry
    template_name = 'finance/journal_list.html'
    context_object_name = 'entries'
    module_name = 'finance'
    permission_type = 'view'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = JournalEntry.objects.filter(is_active=True)
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(entry_number__icontains=search) | Q(reference__icontains=search)
            )
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Journal Entries'
        context['status_choices'] = JournalEntry.STATUS_CHOICES
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        context['today'] = date.today().isoformat()
        return context


class JournalEntryCreateView(CreatePermissionMixin, CreateView):
    model = JournalEntry
    form_class = JournalEntryForm
    template_name = 'finance/journal_form.html'
    success_url = reverse_lazy('finance:journal_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Journal Entry'
        context['today'] = date.today().isoformat()
        if self.request.POST:
            context['lines_formset'] = JournalEntryLineFormSet(self.request.POST)
        else:
            context['lines_formset'] = JournalEntryLineFormSet()
        context['accounts'] = Account.objects.filter(is_active=True).order_by('code')
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        lines_formset = context['lines_formset']
        
        if lines_formset.is_valid():
            self.object = form.save()
            lines_formset.instance = self.object
            lines_formset.save()
            self.object.calculate_totals()
            
            if not self.object.is_balanced:
                messages.warning(self.request, f'Journal Entry {self.object.entry_number} created but is UNBALANCED. Please correct before posting.')
            else:
                messages.success(self.request, f'Journal Entry {self.object.entry_number} created.')
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))


class JournalEntryUpdateView(UpdatePermissionMixin, UpdateView):
    model = JournalEntry
    form_class = JournalEntryForm
    template_name = 'finance/journal_form.html'
    success_url = reverse_lazy('finance:journal_list')
    module_name = 'finance'
    
    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        # Use the new is_editable property for SAP/Oracle compliant rules
        if not obj.is_editable:
            reason = obj.edit_restriction_reason or 'This journal entry cannot be edited.'
            messages.error(self.request, reason)
            return None
        return obj
    
    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object is None:
            return redirect('finance:journal_list')
        return super().get(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Journal Entry: {self.object.entry_number}'
        context['today'] = date.today().isoformat()
        if self.request.POST:
            context['lines_formset'] = JournalEntryLineFormSet(self.request.POST, instance=self.object)
        else:
            context['lines_formset'] = JournalEntryLineFormSet(instance=self.object)
        context['accounts'] = Account.objects.filter(is_active=True).order_by('code')
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        lines_formset = context['lines_formset']
        
        if lines_formset.is_valid():
            self.object = form.save()
            lines_formset.save()
            self.object.calculate_totals()
            messages.success(self.request, f'Journal Entry {self.object.entry_number} updated.')
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))


class JournalEntryDetailView(PermissionRequiredMixin, DetailView):
    model = JournalEntry
    template_name = 'finance/journal_detail.html'
    context_object_name = 'entry'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_context_data(self, **kwargs):
        from apps.core.audit import get_entity_audit_history
        
        context = super().get_context_data(**kwargs)
        context['title'] = f'Journal Entry: {self.object.entry_number}'
        
        has_edit_permission = (
            self.request.user.is_superuser or 
            PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        )
        
        # Use new SAP/Oracle compliant properties
        context['can_edit'] = has_edit_permission and self.object.is_editable
        context['can_delete'] = has_edit_permission and self.object.is_deletable
        context['can_post'] = has_edit_permission and self.object.is_editable and self.object.is_balanced and self.object.line_count >= 2
        context['can_reverse'] = has_edit_permission and self.object.is_reversible
        
        # Show reason why actions are blocked
        context['edit_restriction_reason'] = self.object.edit_restriction_reason
        context['is_system_generated'] = self.object.is_system_generated
        context['is_locked'] = self.object.is_locked
        
        # Audit History
        context['audit_history'] = get_entity_audit_history('JournalEntry', self.object.pk)
        
        return context


@login_required
def journal_post(request, pk):
    """Post a journal entry - validates balance, min lines, leaf accounts, period."""
    from apps.core.audit import audit_journal_post
    
    entry = get_object_or_404(JournalEntry, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:journal_list')
    
    if entry.status != 'draft':
        messages.error(request, 'Only draft entries can be posted.')
        return redirect('finance:journal_detail', pk=pk)
    
    # Validate before posting
    errors = entry.validate_for_posting()
    if errors:
        for error in errors:
            messages.error(request, error)
        return redirect('finance:journal_detail', pk=pk)
    
    try:
        entry.post(user=request.user)
        # Audit log with IP address
        audit_journal_post(entry, request.user, request=request)
        messages.success(request, f'Journal Entry {entry.entry_number} posted successfully.')
    except ValidationError as e:
        for error in e.messages:
            messages.error(request, error)
    except Exception as e:
        messages.error(request, str(e))
    
    return redirect('finance:journal_detail', pk=pk)


@login_required
def journal_reverse(request, pk):
    """
    Reverse a posted journal entry - creates auto-reversal entry.
    This is the ONLY way to correct posted transactions (SAP/Oracle compliant).
    """
    from apps.core.audit import audit_journal_reverse
    
    entry = get_object_or_404(JournalEntry, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:journal_list')
    
    if not entry.is_reversible:
        if entry.status != 'posted':
            messages.error(request, 'Only posted entries can be reversed.')
        elif entry.period and entry.period.is_locked:
            messages.error(request, f'Cannot reverse - accounting period {entry.period.name} is locked.')
        elif entry.fiscal_year and entry.fiscal_year.is_closed:
            messages.error(request, f'Cannot reverse - fiscal year {entry.fiscal_year.name} is closed.')
        else:
            messages.error(request, 'This journal entry cannot be reversed.')
        return redirect('finance:journal_detail', pk=pk)
    
    reason = request.POST.get('reason', 'User requested reversal')
    
    try:
        reversal = entry.reverse(user=request.user, reason=reason)
        # Audit log with IP address
        audit_journal_reverse(entry, reversal, request.user, reason=reason, request=request)
        messages.success(request, f'Journal Entry {entry.entry_number} reversed. Reversal entry: {reversal.entry_number}')
        return redirect('finance:journal_detail', pk=reversal.pk)
    except ValidationError as e:
        for error in e.messages:
            messages.error(request, error)
    except Exception as e:
        messages.error(request, str(e))
    
    return redirect('finance:journal_detail', pk=pk)


@login_required
def journal_delete(request, pk):
    """
    Delete a draft journal entry.
    Only draft (unposted) entries can be deleted.
    Posted entries must be reversed instead.
    """
    entry = get_object_or_404(JournalEntry, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'delete')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:journal_list')
    
    if entry.status != 'draft':
        messages.error(request, 'Only draft entries can be deleted. Posted entries must be reversed.')
        return redirect('finance:journal_detail', pk=pk)
    
    if entry.is_system_generated:
        messages.error(request, 'System-generated entries cannot be deleted. Cancel the source document instead.')
        return redirect('finance:journal_detail', pk=pk)
    
    entry_number = entry.entry_number
    entry.delete()
    messages.success(request, f'Journal Entry {entry_number} deleted successfully.')
    
    return redirect('finance:journal_list')


# ============ PAYMENT VIEWS ============

class PaymentListView(PermissionRequiredMixin, ListView):
    model = Payment
    template_name = 'finance/payment_list.html'
    context_object_name = 'payments'
    module_name = 'finance'
    permission_type = 'view'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = Payment.objects.filter(is_active=True)
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(payment_number__icontains=search) |
                Q(party_name__icontains=search) |
                Q(reference__icontains=search)
            )
        
        payment_type = self.request.GET.get('type')
        if payment_type:
            queryset = queryset.filter(payment_type=payment_type)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Payments'
        context['type_choices'] = Payment.PAYMENT_TYPE_CHOICES
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        context['can_delete'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'delete')
        context['today'] = date.today().isoformat()
        
        # Summary
        payments = self.get_queryset()
        context['total_received'] = payments.filter(payment_type='received').aggregate(Sum('amount'))['amount__sum'] or 0
        context['total_made'] = payments.filter(payment_type='made').aggregate(Sum('amount'))['amount__sum'] or 0
        return context


class PaymentCreateView(CreatePermissionMixin, CreateView):
    model = Payment
    form_class = PaymentForm
    template_name = 'finance/payment_form.html'
    success_url = reverse_lazy('finance:payment_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Payment'
        context['today'] = date.today().isoformat()
        return context
    
    def form_valid(self, form):
        form.instance.party_type = 'customer' if form.instance.payment_type == 'received' else 'vendor'
        form.instance.party_id = 0
        messages.success(self.request, 'Payment created.')
        return super().form_valid(form)


class PaymentDetailView(PermissionRequiredMixin, DetailView):
    """View payment details with audit history."""
    model = Payment
    template_name = 'finance/payment_detail.html'
    context_object_name = 'payment'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_context_data(self, **kwargs):
        from apps.core.audit import get_entity_audit_history
        
        context = super().get_context_data(**kwargs)
        context['title'] = f'Payment: {self.object.payment_number}'
        
        has_permission = (
            self.request.user.is_superuser or 
            PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        )
        context['can_edit'] = has_permission and self.object.status == 'draft'
        context['can_cancel'] = has_permission and self.object.status == 'posted'
        
        # Audit History
        context['audit_history'] = get_entity_audit_history('Payment', self.object.pk)
        
        return context


class PaymentUpdateView(UpdatePermissionMixin, UpdateView):
    model = Payment
    form_class = PaymentForm
    template_name = 'finance/payment_form.html'
    success_url = reverse_lazy('finance:payment_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Payment: {self.object.payment_number}'
        context['today'] = date.today().isoformat()
        return context
    
    def form_valid(self, form):
        messages.success(self.request, f'Payment {form.instance.payment_number} updated.')
        return super().form_valid(form)


@login_required
def payment_cancel(request, pk):
    """Cancel a payment - creates auto-reversal journal."""
    payment = get_object_or_404(Payment, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:payment_list')
    
    if payment.status == 'cancelled':
        messages.error(request, 'Payment is already cancelled.')
        return redirect('finance:payment_list')
    
    from django.utils import timezone
    payment.status = 'cancelled'
    payment.cancelled_date = timezone.now()
    payment.cancellation_reason = request.POST.get('reason', 'User requested cancellation')
    payment.save()
    
    # If there's a linked journal entry, reverse it
    if payment.journal_entry and payment.journal_entry.status == 'posted':
        try:
            reversal = payment.journal_entry.reverse(user=request.user, reason=f'Payment {payment.payment_number} cancelled')
            payment.reversal_entry = reversal
            payment.save()
        except Exception as e:
            messages.warning(request, f'Could not reverse journal entry: {e}')
    
    messages.success(request, f'Payment {payment.payment_number} cancelled.')
    return redirect('finance:payment_list')


@login_required
def payment_delete(request, pk):
    """Delete a draft payment (soft delete)."""
    payment = get_object_or_404(Payment, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'delete')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:payment_list')
    
    if payment.status != 'draft':
        messages.error(request, 'Only draft payments can be deleted.')
        return redirect('finance:payment_list')
    
    # Soft delete
    payment.is_active = False
    payment.save()
    
    messages.success(request, f'Payment {payment.payment_number} deleted.')
    return redirect('finance:payment_list')


# ============ TAX CODE VIEWS ============

class TaxCodeListView(PermissionRequiredMixin, ListView):
    model = TaxCode
    template_name = 'finance/taxcode_list.html'
    context_object_name = 'taxcodes'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        return TaxCode.objects.filter(is_active=True)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Tax Codes'
        context['form'] = TaxCodeForm()
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        return context
    
    def post(self, request, *args, **kwargs):
        if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'create')):
            messages.error(request, 'Permission denied.')
            return redirect('finance:taxcode_list')
        
        form = TaxCodeForm(request.POST)
        if form.is_valid():
            taxcode = form.save()
            messages.success(request, f'Tax Code {taxcode.code} created.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
        return redirect('finance:taxcode_list')


# ============ FINANCIAL REPORTS ============

@login_required
def trial_balance(request):
    """
    Standard Trial Balance Report (As at Date) - UAE Accounting Standard.
    
    IFRS & UAE Audit Compliant - Shows NET BALANCE as of a specific date.
    Grouped by Account Type and Category for professional presentation.
    
    Grouping Hierarchy:
    1. ASSETS (Current Assets, Non-Current Assets)
    2. LIABILITIES (Current Liabilities)
    3. EQUITY / CAPITAL
    4. INCOME
    5. EXPENSES (Cost of Sales, Operating Expenses)
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    # Get as-of date (default: today)
    today = date.today()
    as_of_date_str = request.GET.get('as_of_date', today.isoformat())
    export_format = request.GET.get('format', '')
    show_zero_balances = request.GET.get('show_zero', '') == '1'
    
    try:
        as_of_date = datetime.strptime(as_of_date_str, '%Y-%m-%d').date()
    except ValueError:
        as_of_date = today
    
    accounts = Account.objects.filter(is_active=True).order_by('account_type', 'account_category', 'code')
    
    # Group definitions for UAE standard Trial Balance
    GROUP_STRUCTURE = {
        'asset': {
            'name': 'ASSETS',
            'subgroups': {
                'current_assets': {
                    'name': 'Current Assets',
                    'categories': ['cash_bank', 'trade_receivables', 'tax_receivables', 'inventory', 'prepaid', 'other_current_assets'],
                },
                'non_current_assets': {
                    'name': 'Non-Current Assets (Fixed Assets)',
                    'categories': ['fixed_furniture', 'fixed_it', 'fixed_vehicles', 'fixed_other', 'intangible', 'accum_depreciation'],
                },
            }
        },
        'liability': {
            'name': 'LIABILITIES',
            'subgroups': {
                'current_liabilities': {
                    'name': 'Current Liabilities',
                    'categories': ['trade_payables', 'tax_payables', 'accrued_liabilities', 'other_current_liabilities'],
                },
                'non_current_liabilities': {
                    'name': 'Non-Current Liabilities',
                    'categories': ['long_term_liabilities'],
                },
            }
        },
        'equity': {
            'name': 'EQUITY / CAPITAL',
            'subgroups': {
                'capital': {
                    'name': 'Capital & Reserves',
                    'categories': ['capital', 'reserves', 'retained_earnings'],
                },
            }
        },
        'income': {
            'name': 'INCOME',
            'subgroups': {
                'revenue': {
                    'name': 'Revenue',
                    'categories': ['operating_revenue', 'other_income'],
                },
            }
        },
        'expense': {
            'name': 'EXPENSES',
            'subgroups': {
                'cost_of_sales': {
                    'name': 'Cost of Sales',
                    'categories': ['cost_of_sales'],
                },
                'operating_expenses': {
                    'name': 'Operating Expenses',
                    'categories': ['rent_expense', 'salary_expense', 'banking_expense', 'bad_debts', 
                                   'depreciation_expense', 'utilities', 'project_costs', 'marketing', 
                                   'admin_expense', 'other_expense'],
                },
            }
        },
    }
    
    # Process accounts and calculate balances
    grouped_data = {}
    total_debit = Decimal('0.00')
    total_credit = Decimal('0.00')
    flat_data = []  # For Excel export
    
    for account in accounts:
        # Calculate net balance
        # CRITICAL: Include Account.opening_balance in addition to journal entries
        # This ensures opening balances are reflected even if opening journal doesn't exist
        totals = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__status='posted',
            journal_entry__date__lte=as_of_date
        ).aggregate(
            total_debit=Coalesce(Sum('debit'), Decimal('0.00')),
            total_credit=Coalesce(Sum('credit'), Decimal('0.00'))
        )
        
        # Start with account's opening balance
        # Opening balance is stored as a positive value for debit-normal accounts
        # and negative for credit-normal accounts (or use account's debit_increases property)
        account_opening = account.opening_balance or Decimal('0.00')
        
        # Add journal movements to opening balance
        net_balance = account_opening + (totals['total_debit'] - totals['total_credit'])
        
        # Skip zero balance accounts unless requested
        if net_balance == 0 and not show_zero_balances:
            continue
        
        # Check account properties
        is_cash_or_bank = account.is_cash_account or 'cash' in account.name.lower() or 'bank' in account.name.lower()
        is_contra = getattr(account, 'is_contra_account', False) or 'accumulated' in account.name.lower()
        overdraft_allowed = getattr(account, 'overdraft_allowed', False) or 'overdraft' in account.name.lower()
        
        # Determine debit or credit column based on ACCOUNT NATURE
        # For debit-normal accounts (Assets, Expenses): show positive = Debit, negative = Credit
        # For credit-normal accounts (Liabilities, Equity, Income): show positive = Credit, negative = Debit
        
        if account.debit_increases:
            # Asset or Expense account - normally shows Debit balance
            if net_balance >= 0:
                debit_amount = net_balance
                credit_amount = Decimal('0.00')
            else:
                # Negative balance for debit-normal account
                # For Cash/Bank - this is ABNORMAL unless overdraft allowed
                if is_cash_or_bank and not overdraft_allowed:
                    # Show as DEBIT with warning flag (negative cash/overdraft)
                    # This prevents cash from appearing in Credit column
                    debit_amount = net_balance  # Will be negative
                    credit_amount = Decimal('0.00')
                else:
                    # Other assets with credit balance or overdraft accounts
                    debit_amount = Decimal('0.00')
                    credit_amount = abs(net_balance)
        else:
            # Liability, Equity, or Income account - normally shows Credit balance
            if net_balance <= 0:
                debit_amount = Decimal('0.00')
                credit_amount = abs(net_balance)
            else:
                # Positive balance for credit-normal account = abnormal
                debit_amount = net_balance
                credit_amount = Decimal('0.00')
        
        # Check for abnormal balance
        abnormal = False
        negative_cash_warning = False
        
        if is_contra:
            # Contra accounts have opposite normal balance
            if debit_amount > 0:
                abnormal = True
        elif account.debit_increases:
            # Assets/Expenses should have debit balance
            if credit_amount > 0:
                abnormal = True
            # Special check for negative cash (debit_amount < 0)
            if is_cash_or_bank and net_balance < 0:
                negative_cash_warning = True
                if not overdraft_allowed:
                    abnormal = True  # Negative cash without overdraft is abnormal
        else:
            # Liabilities/Equity/Income should have credit balance
            if debit_amount > 0:
                abnormal = True
        
        account_data = {
                'account': account,
            'code': account.code,
            'name': account.name,
            'account_type': account.account_type,
            'account_type_display': account.get_account_type_display(),
            'category': getattr(account, 'account_category', None) or 'other',
            'debit': debit_amount,
            'credit': credit_amount,
            'abnormal': abnormal,
            'is_contra': is_contra,
            'is_cash_or_bank': is_cash_or_bank,
            'negative_cash_warning': negative_cash_warning,
            'overdraft_allowed': overdraft_allowed,
            'net_balance': net_balance,  # Keep raw balance for reference
        }
        
        flat_data.append(account_data)
        total_debit += debit_amount
        total_credit += credit_amount
        
        # Group by account type and category
        acc_type = account.account_type
        acc_category = getattr(account, 'account_category', None)
        
        if acc_type not in grouped_data:
            grouped_data[acc_type] = {
                'name': GROUP_STRUCTURE.get(acc_type, {}).get('name', acc_type.title()),
                'subgroups': {},
                'total_debit': Decimal('0.00'),
                'total_credit': Decimal('0.00'),
            }
        
        # Find the appropriate subgroup
        subgroup_key = None
        if acc_type in GROUP_STRUCTURE:
            for sg_key, sg_data in GROUP_STRUCTURE[acc_type]['subgroups'].items():
                if acc_category in sg_data['categories']:
                    subgroup_key = sg_key
                    break
        
        if not subgroup_key:
            subgroup_key = 'other'
        
        if subgroup_key not in grouped_data[acc_type]['subgroups']:
            sg_name = GROUP_STRUCTURE.get(acc_type, {}).get('subgroups', {}).get(subgroup_key, {}).get('name', 'Other')
            grouped_data[acc_type]['subgroups'][subgroup_key] = {
                'name': sg_name,
                'accounts': [],
                'total_debit': Decimal('0.00'),
                'total_credit': Decimal('0.00'),
            }
        
        grouped_data[acc_type]['subgroups'][subgroup_key]['accounts'].append(account_data)
        grouped_data[acc_type]['subgroups'][subgroup_key]['total_debit'] += debit_amount
        grouped_data[acc_type]['subgroups'][subgroup_key]['total_credit'] += credit_amount
        grouped_data[acc_type]['total_debit'] += debit_amount
        grouped_data[acc_type]['total_credit'] += credit_amount
    
    # Sort groups in proper order
    group_order = ['asset', 'liability', 'equity', 'income', 'expense']
    sorted_groups = []
    for group_type in group_order:
        if group_type in grouped_data:
            sorted_groups.append({
                'type': group_type,
                **grouped_data[group_type]
            })
    
    # Validation: Total Debit MUST equal Total Credit
    is_balanced = total_debit == total_credit
    difference = total_debit - total_credit
    
    if export_format == 'excel':
        from .excel_exports import export_trial_balance
        from apps.settings_app.models import CompanySettings
        company = CompanySettings.get_settings()
        return export_trial_balance(flat_data, as_of_date_str, company.company_name if company else '')
    
    return render(request, 'finance/trial_balance.html', {
        'title': f'Trial Balance (As at {as_of_date_str})',
        'grouped_data': sorted_groups,
        'trial_data': flat_data,  # For backward compatibility
        'total_debit': total_debit,
        'total_credit': total_credit,
        'is_balanced': is_balanced,
        'difference': difference,
        'as_of_date': as_of_date_str,
        'show_zero_balances': show_zero_balances,
    })


@login_required
def trial_balance_with_movements(request):
    """
    Trial Balance with Movements Report.
    
    Shows Opening Balance, Period Movement, and Closing Balance.
    This is a separate report from the standard Trial Balance.
    
    Opening Balance = All posted journals BEFORE start_date
    Period Movement = All posted journals BETWEEN start_date and end_date
    Closing Balance = Opening + Period Movement
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    # Get date range (default: current fiscal year)
    today = date.today()
    default_start = date(today.year, 1, 1).isoformat()
    default_end = today.isoformat()
    
    start_date_str = request.GET.get('start_date', default_start)
    end_date_str = request.GET.get('end_date', default_end)
    export_format = request.GET.get('format', '')
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        start_date = date(today.year, 1, 1)
        end_date = today
    
    accounts = Account.objects.filter(is_active=True).order_by('account_type', 'code')
    
    trial_data = []
    
    # Totals
    total_opening_debit = Decimal('0.00')
    total_opening_credit = Decimal('0.00')
    total_period_debit = Decimal('0.00')
    total_period_credit = Decimal('0.00')
    total_closing_debit = Decimal('0.00')
    total_closing_credit = Decimal('0.00')
    
    for account in accounts:
        # Account's static opening balance (from Account model)
        account_opening = account.opening_balance or Decimal('0.00')
        
        # Opening Balance: Account opening + Sum of all posted journal lines BEFORE start_date
        opening_lines = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__status='posted',
            journal_entry__date__lt=start_date
        ).aggregate(
            debit=Coalesce(Sum('debit'), Decimal('0.00')),
            credit=Coalesce(Sum('credit'), Decimal('0.00'))
        )
        opening_debit = opening_lines['debit']
        opening_credit = opening_lines['credit']
        # CRITICAL: Include account opening balance in opening calculation
        opening_balance = account_opening + (opening_debit - opening_credit)
        
        # Period Movement: Sum of all posted journal lines BETWEEN start_date and end_date
        period_lines = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__status='posted',
            journal_entry__date__gte=start_date,
            journal_entry__date__lte=end_date
        ).aggregate(
            debit=Coalesce(Sum('debit'), Decimal('0.00')),
            credit=Coalesce(Sum('credit'), Decimal('0.00'))
        )
        period_debit = period_lines['debit']
        period_credit = period_lines['credit']
        
        # Closing Balance = Opening + Period Movement
        closing_balance = opening_balance + (period_debit - period_credit)
        
        # Convert balances to debit/credit columns based on account nature
        # Opening columns
        if opening_balance > 0:
            open_dr = opening_balance
            open_cr = Decimal('0.00')
        elif opening_balance < 0:
            open_dr = Decimal('0.00')
            open_cr = abs(opening_balance)
        else:
            open_dr = Decimal('0.00')
            open_cr = Decimal('0.00')
        
        # Closing columns
        if closing_balance > 0:
            close_dr = closing_balance
            close_cr = Decimal('0.00')
        elif closing_balance < 0:
            close_dr = Decimal('0.00')
            close_cr = abs(closing_balance)
        else:
            close_dr = Decimal('0.00')
            close_cr = Decimal('0.00')
        
        # Only include accounts with activity
        has_activity = (open_dr != 0 or open_cr != 0 or 
                       period_debit != 0 or period_credit != 0 or
                       close_dr != 0 or close_cr != 0)
        
        if has_activity:
            # Check for abnormal balance
            abnormal = False
            is_contra_asset = 'accumulated' in account.name.lower() or 'contra' in account.name.lower()
            
            if is_contra_asset:
                if close_dr > 0:
                    abnormal = True
            elif account.debit_increases and close_cr > 0:
                abnormal = True
            elif not account.debit_increases and close_dr > 0:
                abnormal = True
            
            trial_data.append({
                'account': account,
                'code': account.code,
                'name': account.name,
                'account_type': account.get_account_type_display(),
                'opening_debit': open_dr,
                'opening_credit': open_cr,
                'period_debit': period_debit,
                'period_credit': period_credit,
                'closing_debit': close_dr,
                'closing_credit': close_cr,
                'abnormal': abnormal,
            })
            
            total_opening_debit += open_dr
            total_opening_credit += open_cr
            total_period_debit += period_debit
            total_period_credit += period_credit
            total_closing_debit += close_dr
            total_closing_credit += close_cr
    
    # Validation checks
    opening_balanced = total_opening_debit == total_opening_credit
    closing_balanced = total_closing_debit == total_closing_credit
    is_balanced = opening_balanced and closing_balanced
    
    # Excel Export
    if export_format == 'excel':
        from .excel_exports import export_trial_balance_with_movements
        from apps.settings_app.models import CompanySettings
        company = CompanySettings.get_settings()
        return export_trial_balance_with_movements(trial_data, start_date_str, end_date_str, {
            'total_opening_debit': total_opening_debit,
            'total_opening_credit': total_opening_credit,
            'total_period_debit': total_period_debit,
            'total_period_credit': total_period_credit,
            'total_closing_debit': total_closing_debit,
            'total_closing_credit': total_closing_credit,
        }, company.company_name if company else '')
    
    return render(request, 'finance/trial_balance_with_movements.html', {
        'title': 'Trial Balance with Movements',
        'trial_data': trial_data,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'total_opening_debit': total_opening_debit,
        'total_opening_credit': total_opening_credit,
        'total_period_debit': total_period_debit,
        'total_period_credit': total_period_credit,
        'total_closing_debit': total_closing_debit,
        'total_closing_credit': total_closing_credit,
        'opening_balanced': opening_balanced,
        'closing_balanced': closing_balanced,
        'is_balanced': is_balanced,
    })


@login_required
def profit_loss(request):
    """
    Profit & Loss Statement (Income Statement).
    SINGLE SOURCE OF TRUTH: Reads only from JournalEntryLine for Income and Expense accounts.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    # Get date range
    end_date = request.GET.get('end_date', date.today().isoformat())
    start_date = request.GET.get('start_date', date(date.today().year, 1, 1).isoformat())
    
    # Income accounts - calculate balance from journal lines
    income_accounts = Account.objects.filter(
        is_active=True, 
        account_type=AccountType.INCOME
    ).order_by('code')
    
    income_data = []
    total_income = Decimal('0.00')
    for acc in income_accounts:
        # Income accounts: Credits increase, Debits decrease (Credit - Debit = Balance)
        lines = JournalEntryLine.objects.filter(
            account=acc,
            journal_entry__status='posted',
            journal_entry__date__gte=start_date,
            journal_entry__date__lte=end_date,
        ).aggregate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit')
        )
        debit = lines['total_debit'] or Decimal('0.00')
        credit = lines['total_credit'] or Decimal('0.00')
        balance = credit - debit  # Income is increased by credits
        
        if balance != 0:
            income_data.append({'account': acc, 'amount': balance})
            total_income += balance
    
    # Expense accounts - calculate balance from journal lines
    expense_accounts = Account.objects.filter(
        is_active=True, 
        account_type=AccountType.EXPENSE
    ).order_by('code')
    
    expense_data = []
    total_expenses = Decimal('0.00')
    for acc in expense_accounts:
        # Expense accounts: Debits increase, Credits decrease (Debit - Credit = Balance)
        lines = JournalEntryLine.objects.filter(
            account=acc,
            journal_entry__status='posted',
            journal_entry__date__gte=start_date,
            journal_entry__date__lte=end_date,
        ).aggregate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit')
        )
        debit = lines['total_debit'] or Decimal('0.00')
        credit = lines['total_credit'] or Decimal('0.00')
        balance = debit - credit  # Expense is increased by debits
        
        if balance != 0:
            expense_data.append({'account': acc, 'amount': balance})
            total_expenses += balance
    
    # Calculate profit
    net_profit_before_tax = total_income - total_expenses
    
    # Corporate tax calculation (9% on profit > AED 375,000)
    tax_threshold = Decimal('375000.00')
    tax_rate = Decimal('0.09')
    
    if net_profit_before_tax > tax_threshold:
        corporate_tax = (net_profit_before_tax - tax_threshold) * tax_rate
    else:
        corporate_tax = Decimal('0.00')
    
    net_profit_after_tax = net_profit_before_tax - corporate_tax
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel':
        from .excel_exports import export_profit_loss
        # Prepare data for export
        revenue_export = [{'code': d['account'].code, 'name': d['account'].name, 'balance': d['amount']} for d in income_data]
        expense_export = [{'code': d['account'].code, 'name': d['account'].name, 'balance': d['amount']} for d in expense_data]
        return export_profit_loss(revenue_export, expense_export, start_date, end_date)
    
    return render(request, 'finance/profit_loss.html', {
        'title': 'Profit & Loss Statement',
        'income_data': income_data,
        'expense_data': expense_data,
        'total_income': total_income,
        'total_expenses': total_expenses,
        'net_profit_before_tax': net_profit_before_tax,
        'corporate_tax': corporate_tax,
        'net_profit_after_tax': net_profit_after_tax,
        'start_date': start_date,
        'end_date': end_date,
    })


@login_required
def balance_sheet(request):
    """
    Balance Sheet - Assets = Liabilities + Equity.
    SINGLE SOURCE OF TRUTH: Reads only from JournalEntryLine for all account types.
    
    IFRS/GAAP Compliant:
    - Period-wise filtering (from-to dates)
    - Accumulated depreciation respects reporting period
    - Contra-assets properly reduce related asset values
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    # Period-wise filtering
    start_date = request.GET.get('start_date', date(date.today().year, 1, 1).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    
    def get_account_balance(account, up_to_date):
        """
        Calculate account balance from journal lines up to a specific date.
        Respects reporting period for accumulated depreciation.
        """
        lines = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__status='posted',
            journal_entry__date__lte=up_to_date,
        ).aggregate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit')
        )
        debit = lines['total_debit'] or Decimal('0.00')
        credit = lines['total_credit'] or Decimal('0.00')
        
        # Account type determines balance calculation
        if account.debit_increases and not account.is_contra_account:
            # Normal Assets and Expenses: Debit - Credit
            return account.opening_balance + (debit - credit)
        elif account.is_contra_account:
            # Contra accounts (like Accumulated Depreciation): Credit - Debit (shows as negative/reducing value)
            return account.opening_balance + (credit - debit)
        else:
            # Liabilities, Equity, Income: Credit - Debit
            return account.opening_balance + (credit - debit)
    
    # Get all asset accounts
    asset_accounts = Account.objects.filter(
        is_active=True, 
        account_type=AccountType.ASSET
    ).order_by('code')
    
    # Group assets: Fixed assets with their accumulated depreciation
    # Fixed asset categories that have depreciation
    fixed_asset_categories = ['fixed_furniture', 'fixed_it', 'fixed_vehicles', 'fixed_other']
    
    # Collect fixed assets and accumulated depreciation separately
    fixed_assets_data = {}  # {category: {'assets': [], 'depreciation': [], 'total_cost': 0, 'total_dep': 0, 'net_value': 0}}
    current_assets_data = []
    other_assets_data = []
    
    total_assets = Decimal('0.00')
    total_fixed_assets_cost = Decimal('0.00')
    total_accumulated_depreciation = Decimal('0.00')
    total_net_fixed_assets = Decimal('0.00')
    total_current_assets = Decimal('0.00')
    
    for acc in asset_accounts:
        balance = get_account_balance(acc, end_date)
        if balance == 0:
            continue
            
        # Check if it's accumulated depreciation
        if acc.is_contra_account or acc.account_category == 'accum_depreciation' or 'accumulated depreciation' in acc.name.lower():
            # Find related fixed asset category
            asset_name_lower = acc.name.lower()
            related_category = 'fixed_other'  # default
            
            if 'furniture' in asset_name_lower:
                related_category = 'fixed_furniture'
            elif 'it' in asset_name_lower or 'computer' in asset_name_lower or 'equipment' in asset_name_lower:
                related_category = 'fixed_it'
            elif 'vehicle' in asset_name_lower:
                related_category = 'fixed_vehicles'
            
            if related_category not in fixed_assets_data:
                fixed_assets_data[related_category] = {
                    'category_name': dict(AccountCategory.choices).get(related_category, 'Fixed Assets'),
                    'assets': [],
                    'depreciation': [],
                    'total_cost': Decimal('0.00'),
                    'total_dep': Decimal('0.00'),
                    'net_value': Decimal('0.00')
                }
            
            # Store as positive value for display, but it reduces assets
            fixed_assets_data[related_category]['depreciation'].append({
                'account': acc,
                'amount': abs(balance)
            })
            fixed_assets_data[related_category]['total_dep'] += abs(balance)
            total_accumulated_depreciation += abs(balance)
            
        # Check if it's a fixed asset
        elif acc.account_category in fixed_asset_categories:
            if acc.account_category not in fixed_assets_data:
                fixed_assets_data[acc.account_category] = {
                    'category_name': dict(AccountCategory.choices).get(acc.account_category, 'Fixed Assets'),
                    'assets': [],
                    'depreciation': [],
                    'total_cost': Decimal('0.00'),
                    'total_dep': Decimal('0.00'),
                    'net_value': Decimal('0.00')
                }
            
            fixed_assets_data[acc.account_category]['assets'].append({
                'account': acc,
                'amount': balance
            })
            fixed_assets_data[acc.account_category]['total_cost'] += balance
            total_fixed_assets_cost += balance
            
        # Current assets (Cash, Bank, Receivables, etc.)
        # CRITICAL: Check BOTH account_category AND is_cash_account flag
        # This ensures consistency with Cash Flow Statement
        elif (
            acc.account_category in ['cash_bank', 'trade_receivables', 'tax_receivables', 'inventory', 'prepaid', 'other_current_assets'] or
            acc.is_cash_account or  # Include ALL accounts flagged as cash
            (acc.account_category is None and ('cash' in acc.name.lower() or 'bank' in acc.name.lower()) and 
             'receivable' not in acc.name.lower() and 'pdc' not in acc.name.lower())  # Fallback for uncategorized
        ):
            current_assets_data.append({'account': acc, 'amount': balance})
            total_current_assets += balance
        else:
            # Other assets
            other_assets_data.append({'account': acc, 'amount': balance})
            total_assets += balance
    
    # Calculate net book values for fixed assets
    for category in fixed_assets_data:
        fixed_assets_data[category]['net_value'] = fixed_assets_data[category]['total_cost'] - fixed_assets_data[category]['total_dep']
    
    total_net_fixed_assets = total_fixed_assets_cost - total_accumulated_depreciation
    total_assets = total_current_assets + total_net_fixed_assets + sum(item['amount'] for item in other_assets_data)
    
    # Prepare flat asset_data for backward compatibility with Excel export
    asset_data = []
    for item in current_assets_data:
        asset_data.append(item)
    for category, data in sorted(fixed_assets_data.items()):
        for item in data['assets']:
            asset_data.append(item)
        for dep in data['depreciation']:
            asset_data.append({'account': dep['account'], 'amount': -dep['amount']})  # Show as negative
    for item in other_assets_data:
        asset_data.append(item)
    
    # Liabilities
    liability_accounts = Account.objects.filter(
        is_active=True, 
        account_type=AccountType.LIABILITY
    ).order_by('code')
    
    liability_data = []
    total_liabilities = Decimal('0.00')
    for acc in liability_accounts:
        balance = get_account_balance(acc, end_date)
        if balance != 0:
            liability_data.append({'account': acc, 'amount': balance})
            total_liabilities += balance
    
    # Equity
    equity_accounts = Account.objects.filter(
        is_active=True, 
        account_type=AccountType.EQUITY
    ).order_by('code')
    
    equity_data = []
    total_equity = Decimal('0.00')
    for acc in equity_accounts:
        balance = get_account_balance(acc, end_date)
        if balance != 0:
            equity_data.append({'account': acc, 'amount': balance})
            total_equity += balance
    
    # Calculate retained earnings (current year P&L from journal lines)
    income_accounts = Account.objects.filter(
        is_active=True, account_type=AccountType.INCOME
    )
    expense_accounts = Account.objects.filter(
        is_active=True, account_type=AccountType.EXPENSE
    )
    
    income_lines = JournalEntryLine.objects.filter(
        account__in=income_accounts,
        journal_entry__status='posted',
        journal_entry__date__lte=end_date,
    ).aggregate(
        total_debit=Sum('debit'),
        total_credit=Sum('credit')
    )
    income_total = (income_lines['total_credit'] or Decimal('0.00')) - (income_lines['total_debit'] or Decimal('0.00'))
    
    expense_lines = JournalEntryLine.objects.filter(
        account__in=expense_accounts,
        journal_entry__status='posted',
        journal_entry__date__lte=end_date,
    ).aggregate(
        total_debit=Sum('debit'),
        total_credit=Sum('credit')
    )
    expense_total = (expense_lines['total_debit'] or Decimal('0.00')) - (expense_lines['total_credit'] or Decimal('0.00'))
    
    current_year_profit = income_total - expense_total
    total_equity += current_year_profit
    
    # Balance check
    total_liabilities_equity = total_liabilities + total_equity
    is_balanced = abs(total_assets - total_liabilities_equity) < Decimal('0.01')  # Allow small rounding difference
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel':
        from .excel_exports import export_balance_sheet
        assets_export = [{'code': d['account'].code, 'name': d['account'].name, 'balance': d['amount']} for d in asset_data]
        liabilities_export = [{'code': d['account'].code, 'name': d['account'].name, 'balance': d['amount']} for d in liability_data]
        equity_export = [{'code': d['account'].code, 'name': d['account'].name, 'balance': d['amount']} for d in equity_data]
        if current_year_profit != 0:
            equity_export.append({'code': '', 'name': 'Current Year Profit/Loss', 'balance': current_year_profit})
        return export_balance_sheet(assets_export, liabilities_export, equity_export, end_date, start_date)
    
    return render(request, 'finance/balance_sheet.html', {
        'title': 'Balance Sheet',
        'asset_data': asset_data,
        'liability_data': liability_data,
        'equity_data': equity_data,
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'total_equity': total_equity,
        'current_year_profit': current_year_profit,
        'total_liabilities_equity': total_liabilities_equity,
        'is_balanced': is_balanced,
        'start_date': start_date,
        'end_date': end_date,
        # New structured data for proper presentation
        'current_assets_data': current_assets_data,
        'total_current_assets': total_current_assets,
        'fixed_assets_data': fixed_assets_data,
        'total_fixed_assets_cost': total_fixed_assets_cost,
        'total_accumulated_depreciation': total_accumulated_depreciation,
        'total_net_fixed_assets': total_net_fixed_assets,
        'other_assets_data': other_assets_data,
    })


@login_required
def general_ledger(request):
    """General Ledger - All transactions for an account."""
    from django.http import JsonResponse
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    account_id = request.GET.get('account')
    start_date = request.GET.get('start_date', date(date.today().year, 1, 1).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    export_format = request.GET.get('format', '')
    
    accounts = Account.objects.filter(is_active=True).order_by('code')
    selected_account = None
    transactions = []
    running_balance = Decimal('0.00')
    
    if account_id:
        selected_account = get_object_or_404(Account, pk=account_id)
        
        # Calculate opening balance as of start_date
        # = Account opening balance + all posted journals before start_date
        base_opening = selected_account.opening_balance or Decimal('0.00')
        
        # Get all movements before the start date
        pre_period = JournalEntryLine.objects.filter(
            account=selected_account,
            journal_entry__status='posted',
            journal_entry__date__lt=start_date
        ).aggregate(
            total_debit=Coalesce(Sum('debit'), Decimal('0.00')),
            total_credit=Coalesce(Sum('credit'), Decimal('0.00'))
        )
        
        # Calculate opening balance at start_date
        if selected_account.debit_increases:
            running_balance = base_opening + (pre_period['total_debit'] - pre_period['total_credit'])
        else:
            running_balance = base_opening + (pre_period['total_credit'] - pre_period['total_debit'])
        
        opening_balance = running_balance  # Store for display
        
        lines = JournalEntryLine.objects.filter(
            account=selected_account,
            journal_entry__status='posted',
            journal_entry__date__gte=start_date,
            journal_entry__date__lte=end_date,
        ).select_related('journal_entry').order_by('journal_entry__date', 'id')
        
        for line in lines:
            if selected_account.debit_increases:
                running_balance += line.debit - line.credit
            else:
                running_balance += line.credit - line.debit
            
            transactions.append({
                'date': line.journal_entry.date,
                'journal_pk': line.journal_entry.pk,
                'journal_id': line.journal_entry.pk,
                'entry_number': line.journal_entry.entry_number,
                'reference': line.journal_entry.reference,
                'source_module': line.journal_entry.source_module,
                'description': line.description or line.journal_entry.description,
                'debit': line.debit,
                'credit': line.credit,
                'balance': running_balance,
            })
    
    # JSON Export (for drill-down)
    if export_format == 'json' and selected_account:
        return JsonResponse({
            'account': {
                'code': selected_account.code,
                'name': selected_account.name,
            },
            'start_date': start_date,
            'end_date': end_date,
            'entries': [
                {
                    'date': str(t['date']),
                    'journal_id': t['journal_pk'],
                    'entry_number': t['entry_number'],
                    'source_module': t.get('source_module', 'Manual'),
                    'description': t['description'],
                    'debit': str(t['debit']) if t['debit'] else None,
                    'credit': str(t['credit']) if t['credit'] else None,
                    'balance': str(t['balance']),
                }
                for t in transactions
            ]
        })
    
    # Excel Export
    if export_format == 'excel' and selected_account:
        from .excel_exports import export_general_ledger
        return export_general_ledger(transactions, selected_account.name, start_date, end_date)
    
    # Opening balance for context (calculated at start_date, not just account opening)
    context_opening_balance = opening_balance if account_id and 'opening_balance' in locals() else Decimal('0.00')
    
    return render(request, 'finance/general_ledger.html', {
        'title': 'General Ledger',
        'accounts': accounts,
        'selected_account': selected_account,
        'transactions': transactions,
        'opening_balance': context_opening_balance,
        'closing_balance': running_balance,
        'start_date': start_date,
        'end_date': end_date,
    })


@login_required
def vat_report(request):
    """
    UAE VAT Return Report (FTA format).
    
    DATA SOURCE LOGIC (UAE FTA Compliant):
    1. If a VAT Return exists for the period AND status is 'posted', 'submitted', or 'accepted':
       → Use values from the VATReturn record (SOURCE OF TRUTH)
       → This ensures consistency with VAT Return History
    2. If no submitted VAT Return exists:
       → Calculate from transactions (draft/preview mode)
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    # Get date range
    start_date_str = request.GET.get('start_date', date(date.today().year, (date.today().month - 1) // 3 * 3 + 1, 1).isoformat())
    end_date_str = request.GET.get('end_date', date.today().isoformat())
    
    try:
        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)
    except ValueError:
        start_date = date(date.today().year, 1, 1)
        end_date = date.today()
    
    # Get all VAT returns for history display
    vat_returns = VATReturn.objects.filter(is_active=True).order_by('-period_start')
    
    # ========================================
    # CHECK FOR SUBMITTED VAT RETURN FOR THIS PERIOD
    # ========================================
    # Look for VAT Return that covers or overlaps the selected period
    submitted_vat_return = VATReturn.objects.filter(
        is_active=True,
        status__in=['posted', 'submitted', 'accepted'],  # Final statuses
        period_start__lte=end_date,
        period_end__gte=start_date,
    ).first()
    
    # Flag to indicate if we're showing submitted data or calculated data
    is_submitted_data = submitted_vat_return is not None
    is_period_locked = is_submitted_data and submitted_vat_return.status in ['submitted', 'accepted']
    
    if is_submitted_data:
        # ========================================
        # USE SUBMITTED VAT RETURN AS SOURCE OF TRUTH
        # ========================================
        # Box 1: Standard Rated Supplies
        standard_rated_supplies = submitted_vat_return.standard_rated_supplies
        standard_rated_vat = submitted_vat_return.standard_rated_vat
        
        # Box 2: Zero Rated Supplies
        zero_rated_supplies = submitted_vat_return.zero_rated_supplies
        
        # Box 3: Exempt Supplies
        exempt_supplies = submitted_vat_return.exempt_supplies
        
        # Box 9: Standard Rated Expenses
        standard_rated_expenses = submitted_vat_return.standard_rated_expenses
        
        # Box 10: Input VAT (Recoverable)
        current_input_vat = submitted_vat_return.input_vat
        
        # Totals
        current_sales = submitted_vat_return.total_sales
        current_output_vat = submitted_vat_return.output_vat
        current_purchases = submitted_vat_return.total_purchases
        
        # Net VAT (payable/refundable)
        current_net_vat = submitted_vat_return.net_vat
        
        # Adjustments
        adjustments = submitted_vat_return.adjustments
        adjustment_reason = submitted_vat_return.adjustment_reason
        
    else:
        # ========================================
        # CALCULATE FROM TRANSACTIONS (DRAFT/PREVIEW MODE)
        # ========================================
        
        # Get ALL VAT Payable accounts (Output VAT) - not just first!
        # CRITICAL: Look for accounts with 'vat' and 'output' or 'payable' in name
        vat_payable_accounts = Account.objects.filter(
            account_type=AccountType.LIABILITY, 
            is_active=True,
            name__icontains='vat'
        ).filter(
            Q(name__icontains='output') | Q(name__icontains='payable')
        ).exclude(
            # Exclude settlement accounts (used for VAT return posting)
            name__icontains='settlement'
        )
        
        # Fallback: Look by code pattern (21xx, 22xx for VAT Payable)
        if not vat_payable_accounts.exists():
            vat_payable_accounts = Account.objects.filter(
                Q(code__startswith='21') | Q(code__startswith='22') | Q(code__startswith='2000'), 
                account_type=AccountType.LIABILITY, 
                is_active=True,
                name__icontains='vat'
            )
        
        # Get ALL VAT Recoverable accounts (Input VAT) - not just first!
        # CRITICAL: Look for accounts with 'vat' and 'input' or 'recoverable' in name
        vat_recoverable_accounts = Account.objects.filter(
            account_type=AccountType.ASSET, 
            is_active=True,
            name__icontains='vat'
        ).filter(
            Q(name__icontains='input') | Q(name__icontains='recoverable')
        )
        
        # Fallback: Look by code pattern (12xx, 13xx for Input VAT)
        if not vat_recoverable_accounts.exists():
            vat_recoverable_accounts = Account.objects.filter(
                Q(code__startswith='12') | Q(code__startswith='13') | Q(code__startswith='1000'), 
                account_type=AccountType.ASSET, 
                is_active=True,
                name__icontains='vat'
            )
    
    # Get Sales accounts (Income - typically 4xxx)
    sales_accounts = Account.objects.filter(
        account_type=AccountType.INCOME, is_active=True
    )
    
    # Get Expense accounts (Expense - typically 5xxx)
    expense_accounts = Account.objects.filter(
        account_type=AccountType.EXPENSE, is_active=True
    )
    
        # Calculate Output VAT from ALL VAT Payable accounts
        # Output VAT = Credit entries to VAT Payable (when sales are made)
    output_vat_lines = JournalEntryLine.objects.filter(
            account__in=vat_payable_accounts,
        journal_entry__status='posted',
        journal_entry__date__gte=start_date,
        journal_entry__date__lte=end_date,
        ).exclude(
            # Exclude VAT Return settlement entries (already counted in submitted returns)
            journal_entry__source_module='vat'
        ).exclude(
            journal_entry__source_module='vat_return'
        ) if vat_payable_accounts.exists() else JournalEntryLine.objects.none()
    
    current_output_vat = output_vat_lines.aggregate(
        total=Sum('credit')
    )['total'] or Decimal('0.00')
    
        # Calculate Input VAT from ALL VAT Recoverable accounts
        # Input VAT = Debit entries to VAT Recoverable (when purchases are made)
    input_vat_lines = JournalEntryLine.objects.filter(
            account__in=vat_recoverable_accounts,
        journal_entry__status='posted',
        journal_entry__date__gte=start_date,
        journal_entry__date__lte=end_date,
        ).exclude(
            journal_entry__source_module='vat'
        ).exclude(
            journal_entry__source_module='vat_return'
        ) if vat_recoverable_accounts.exists() else JournalEntryLine.objects.none()
    
    current_input_vat = input_vat_lines.aggregate(
        total=Sum('debit')
    )['total'] or Decimal('0.00')
    
    # Calculate Sales from Income account journal lines (Credits = Sales)
    sales_lines = JournalEntryLine.objects.filter(
        account__in=sales_accounts,
        journal_entry__status='posted',
        journal_entry__date__gte=start_date,
        journal_entry__date__lte=end_date,
    )
    
    current_sales = sales_lines.aggregate(
        total=Sum('credit')
    )['total'] or Decimal('0.00')
    
    # Calculate Purchases from Expense account journal lines (Debits = Expenses)
    expense_lines = JournalEntryLine.objects.filter(
        account__in=expense_accounts,
        journal_entry__status='posted',
        journal_entry__date__gte=start_date,
        journal_entry__date__lte=end_date,
    )
    
    current_purchases = expense_lines.aggregate(
        total=Sum('debit')
    )['total'] or Decimal('0.00')
    
    # Standard Rated = Amounts where VAT was charged (5%)
    # Simplified: assumes all sales/purchases are standard rated
    standard_rated_supplies = current_sales
    standard_rated_vat = current_output_vat
    standard_rated_expenses = current_purchases
    
    # Zero rated and Exempt (not yet calculated from transactions)
    zero_rated_supplies = Decimal('0.00')
    exempt_supplies = Decimal('0.00')
    
    # Net VAT
    current_net_vat = current_output_vat - current_input_vat
    
    # No adjustments in draft mode
    adjustments = Decimal('0.00')
    adjustment_reason = ''
    
    # ========================================
    # EXCEL EXPORT
    # ========================================
    export_format = request.GET.get('format', '')
    if export_format == 'excel':
        from .excel_exports import export_vat_report
        vat_data = {
            'standard_sales': standard_rated_supplies,
            'output_vat': current_output_vat,
            'zero_rated_sales': zero_rated_supplies,
            'exempt_sales': exempt_supplies,
            'standard_purchases': standard_rated_expenses,
            'input_vat': current_input_vat,
            'net_vat': current_net_vat,
            'adjustments': adjustments if is_submitted_data else Decimal('0.00'),
            'is_submitted': is_submitted_data,
            'vat_return_number': submitted_vat_return.return_number if is_submitted_data else None,
            'vat_return_status': submitted_vat_return.status if is_submitted_data else 'draft',
        }
        return export_vat_report(vat_data, start_date_str, end_date_str)
    
    return render(request, 'finance/vat_report.html', {
        'title': 'VAT Report (FTA)',
        'vat_returns': vat_returns,
        
        # Data source indicator
        'is_submitted_data': is_submitted_data,
        'is_period_locked': is_period_locked,
        'submitted_vat_return': submitted_vat_return,
        
        # Box 1: Standard Rated Supplies (5%)
        'standard_rated_supplies': standard_rated_supplies,
        'standard_rated_vat': standard_rated_vat,
        
        # Box 2: Zero Rated Supplies (0%)
        'zero_rated_supplies': zero_rated_supplies,
        
        # Box 3: Exempt Supplies
        'exempt_supplies': exempt_supplies,
        
        # Box 4: Out of scope (not applicable for UAE)
        'out_of_scope': Decimal('0.00'),
        
        # Box 9: Standard Rated Expenses
        'standard_rated_expenses': standard_rated_expenses,
        
        # Box 10: Input VAT (Recoverable)
        'input_vat': current_input_vat,
        
        # Totals
        'total_sales': current_sales,
        'total_output_vat': current_output_vat,
        'total_purchases': current_purchases,
        'total_input_vat': current_input_vat,
        
        # Net VAT (Box 11)
        'net_vat': current_net_vat,
        'is_refund': current_net_vat < 0,
        
        # Adjustments
        'adjustments': adjustments if is_submitted_data else Decimal('0.00'),
        'adjustment_reason': adjustment_reason if is_submitted_data else '',
        
        # Date range
        'start_date': start_date_str,
        'end_date': end_date_str,
    })


@login_required
def corporate_tax_report(request):
    """
    UAE Corporate Tax Computation Report.
    SINGLE SOURCE OF TRUTH: Reads only from JournalEntryLine for P&L accounts.
    
    UAE Corporate Tax Law (Federal Decree-Law No. 47 of 2022):
    - Tax Rate: 9% on profit exceeding AED 375,000
    - Accounting Profit → Adjustments → Taxable Income → Tax
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    # Get fiscal year
    fiscal_year_id = request.GET.get('fiscal_year')
    fiscal_years = FiscalYear.objects.filter(is_active=True).order_by('-start_date')
    
    # Default to current year
    if fiscal_year_id:
        selected_fiscal_year = FiscalYear.objects.filter(pk=fiscal_year_id).first()
    else:
        selected_fiscal_year = fiscal_years.first()
    
    # Get date range from fiscal year
    if selected_fiscal_year:
        start_date = selected_fiscal_year.start_date.isoformat()
        end_date = selected_fiscal_year.end_date.isoformat()
    else:
        start_date = date(date.today().year, 1, 1).isoformat()
        end_date = date.today().isoformat()
    
    # Get existing computations
    tax_computations = CorporateTaxComputation.objects.filter(is_active=True).select_related(
        'fiscal_year', 'journal_entry', 'payment_journal_entry'
    ).order_by('-fiscal_year__start_date')
    
    # Get account classifications once
    income_accounts = Account.objects.filter(is_active=True, account_type=AccountType.INCOME)
    expense_accounts = Account.objects.filter(is_active=True, account_type=AccountType.EXPENSE)
    
    # ========================================
    # AUTO-UPDATE DRAFT COMPUTATIONS FROM GL
    # This ensures History always matches Computation
    # Only draft/final computations are updated (not filed/paid - those are locked snapshots)
    # ========================================
    for comp in tax_computations:
        if comp.status not in ['filed', 'paid']:  # Can update draft/final
            fy_start = comp.fiscal_year.start_date
            fy_end = comp.fiscal_year.end_date
            
            # Calculate from GL
            fy_income = JournalEntryLine.objects.filter(
                account__in=income_accounts,
                journal_entry__status='posted',
                journal_entry__date__gte=fy_start,
                journal_entry__date__lte=fy_end,
            ).aggregate(
                total_debit=Sum('debit'),
                total_credit=Sum('credit')
            )
            fy_revenue = (fy_income['total_credit'] or Decimal('0.00')) - (fy_income['total_debit'] or Decimal('0.00'))
            
            fy_expense = JournalEntryLine.objects.filter(
                account__in=expense_accounts,
                journal_entry__status='posted',
                journal_entry__date__gte=fy_start,
                journal_entry__date__lte=fy_end,
            ).aggregate(
                total_debit=Sum('debit'),
                total_credit=Sum('credit')
            )
            fy_expenses = (fy_expense['total_debit'] or Decimal('0.00')) - (fy_expense['total_credit'] or Decimal('0.00'))
            
            # Update if values changed
            if comp.revenue != fy_revenue or comp.expenses != fy_expenses:
                comp.revenue = fy_revenue
                comp.expenses = fy_expenses
                comp.save(update_fields=['revenue', 'expenses'])
                comp.calculate()  # Recalculate derived values
    
    # Refresh the queryset to get updated values
    tax_computations = CorporateTaxComputation.objects.filter(is_active=True).select_related(
        'fiscal_year', 'journal_entry', 'payment_journal_entry'
    ).order_by('-fiscal_year__start_date')
    
    # Calculate current year tax from Journal Lines (SINGLE SOURCE OF TRUTH)
    
    # Revenue from journal lines (Credits to Income accounts)
    income_lines = JournalEntryLine.objects.filter(
        account__in=income_accounts,
        journal_entry__status='posted',
        journal_entry__date__gte=start_date,
        journal_entry__date__lte=end_date,
    ).aggregate(
        total_debit=Sum('debit'),
        total_credit=Sum('credit')
    )
    current_revenue = (income_lines['total_credit'] or Decimal('0.00')) - (income_lines['total_debit'] or Decimal('0.00'))
    
    # Expenses from journal lines (Debits to Expense accounts)
    expense_lines = JournalEntryLine.objects.filter(
        account__in=expense_accounts,
        journal_entry__status='posted',
        journal_entry__date__gte=start_date,
        journal_entry__date__lte=end_date,
    ).aggregate(
        total_debit=Sum('debit'),
        total_credit=Sum('credit')
    )
    current_expenses = (expense_lines['total_debit'] or Decimal('0.00')) - (expense_lines['total_credit'] or Decimal('0.00'))
    
    # Accounting profit (before adjustments)
    accounting_profit = current_revenue - current_expenses
    
    # UAE Corporate Tax (9% on profit > AED 375,000)
    tax_threshold = Decimal('375000.00')
    tax_rate = Decimal('9.00')
    
    # For quick calculation (without adjustments) - display only
    if accounting_profit > tax_threshold:
        taxable_amount = accounting_profit - tax_threshold
        tax_payable_estimate = (taxable_amount * tax_rate / 100).quantize(Decimal('0.01'))
    else:
        taxable_amount = Decimal('0.00')
        tax_payable_estimate = Decimal('0.00')
    
    # Get existing computation for selected fiscal year
    existing_computation = None
    # Computed values for display (always use current GL values)
    computation_revenue = current_revenue
    computation_expenses = current_expenses
    computation_accounting_profit = accounting_profit
    computation_non_deductible = Decimal('0.00')
    computation_exempt_income = Decimal('0.00')
    computation_other_adjustments = Decimal('0.00')
    computation_taxable_income = Decimal('0.00')
    computation_taxable_above_threshold = Decimal('0.00')
    computation_tax_payable = Decimal('0.00')
    
    if selected_fiscal_year:
        existing_computation = CorporateTaxComputation.objects.filter(
            fiscal_year=selected_fiscal_year, is_active=True
        ).first()
        
        if existing_computation:
            # ========================================
            # RECALCULATE USING CURRENT GL VALUES + STORED ADJUSTMENTS
            # This ensures consistency between GL Summary and Tax Computation
            # ========================================
            computation_non_deductible = existing_computation.non_deductible_expenses
            computation_exempt_income = existing_computation.exempt_income
            computation_other_adjustments = existing_computation.other_adjustments
            
            # Taxable Income = GL Accounting Profit + Adjustments
            total_adjustments = computation_non_deductible - computation_exempt_income + computation_other_adjustments
            computation_taxable_income = accounting_profit + total_adjustments
            
            # Apply threshold
            if computation_taxable_income > tax_threshold:
                computation_taxable_above_threshold = computation_taxable_income - tax_threshold
                computation_tax_payable = (computation_taxable_above_threshold * tax_rate / 100).quantize(Decimal('0.01'))
            else:
                computation_taxable_above_threshold = Decimal('0.00')
                computation_tax_payable = Decimal('0.00')
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel' and selected_fiscal_year:
        from .excel_exports import export_corporate_tax
        tax_data = {
            'fiscal_year': selected_fiscal_year.name,
            'start_date': start_date,
            'end_date': end_date,
            'revenue': current_revenue,
            'expenses': current_expenses,
            'accounting_profit': accounting_profit,
            'tax_threshold': tax_threshold,
            'tax_rate': tax_rate,
            'taxable_amount': taxable_amount,
            'tax_payable': tax_payable_estimate,
            'computation': existing_computation,
        }
        return export_corporate_tax(tax_data)
    
    return render(request, 'finance/corporate_tax_report.html', {
        'title': 'Corporate Tax Report',
        'fiscal_years': fiscal_years,
        'selected_fiscal_year': selected_fiscal_year,
        'tax_computations': tax_computations,
        'existing_computation': existing_computation,
        # GL-derived values (SINGLE SOURCE OF TRUTH)
        'current_revenue': current_revenue,
        'current_expenses': current_expenses,
        'accounting_profit': accounting_profit,
        'tax_threshold': tax_threshold,
        'tax_rate': tax_rate,
        'taxable_amount': taxable_amount,
        'tax_payable_estimate': tax_payable_estimate,
        # Computed values for Tax Computation section (GL + Adjustments)
        'comp_revenue': computation_revenue,
        'comp_expenses': computation_expenses,
        'comp_accounting_profit': computation_accounting_profit,
        'comp_non_deductible': computation_non_deductible,
        'comp_exempt_income': computation_exempt_income,
        'comp_other_adjustments': computation_other_adjustments,
        'comp_taxable_income': computation_taxable_income,
        'comp_taxable_above_threshold': computation_taxable_above_threshold,
        'comp_tax_payable': computation_tax_payable,
        # Date range
        'start_date': start_date,
        'end_date': end_date,
        'can_create': request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'create'),
        'can_post': request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit'),
    })


@login_required
def corporate_tax_create(request):
    """
    Create Corporate Tax Computation with adjustments.
    
    IMPORTANT: Revenue/Expenses/Accounting Profit are ALWAYS derived from GL.
    User can only enter adjustments (non-deductible, exempt income, etc.)
    This ensures consistency between GL and Tax Computation.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'create')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:corporate_tax_report')
    
    if request.method == 'POST':
        fiscal_year_id = request.POST.get('fiscal_year')
        fiscal_year = get_object_or_404(FiscalYear, pk=fiscal_year_id)
        
        # Check if already exists
        existing = CorporateTaxComputation.objects.filter(fiscal_year=fiscal_year, is_active=True).first()
        if existing:
            messages.error(request, f'Tax computation already exists for {fiscal_year.name}.')
            return redirect('finance:corporate_tax_report')
        
        # ========================================
        # CALCULATE REVENUE/EXPENSES FROM GL (SINGLE SOURCE OF TRUTH)
        # ========================================
        start_date = fiscal_year.start_date
        end_date = fiscal_year.end_date
        
        income_accounts = Account.objects.filter(is_active=True, account_type=AccountType.INCOME)
        expense_accounts = Account.objects.filter(is_active=True, account_type=AccountType.EXPENSE)
        
        # Revenue from journal lines (Credits to Income accounts)
        income_lines = JournalEntryLine.objects.filter(
            account__in=income_accounts,
            journal_entry__status='posted',
            journal_entry__date__gte=start_date,
            journal_entry__date__lte=end_date,
        ).aggregate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit')
        )
        gl_revenue = (income_lines['total_credit'] or Decimal('0.00')) - (income_lines['total_debit'] or Decimal('0.00'))
        
        # Expenses from journal lines (Debits to Expense accounts)
        expense_lines = JournalEntryLine.objects.filter(
            account__in=expense_accounts,
            journal_entry__status='posted',
            journal_entry__date__gte=start_date,
            journal_entry__date__lte=end_date,
        ).aggregate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit')
        )
        gl_expenses = (expense_lines['total_debit'] or Decimal('0.00')) - (expense_lines['total_credit'] or Decimal('0.00'))
        
        # ========================================
        # GET ADJUSTMENTS FROM FORM (User Input)
        # ========================================
        try:
            non_deductible = Decimal(request.POST.get('non_deductible_expenses', '0') or '0')
            exempt_income = Decimal(request.POST.get('exempt_income', '0') or '0')
            other_adjustments = Decimal(request.POST.get('other_adjustments', '0') or '0')
            notes = request.POST.get('notes', '')
        except:
            messages.error(request, 'Invalid adjustment amounts.')
            return redirect('finance:corporate_tax_report')
        
        # Create computation with GL-derived values
        computation = CorporateTaxComputation.objects.create(
            fiscal_year=fiscal_year,
            revenue=gl_revenue,  # From GL - NOT user input
            expenses=gl_expenses,  # From GL - NOT user input
            non_deductible_expenses=non_deductible,
            exempt_income=exempt_income,
            other_adjustments=other_adjustments,
            notes=notes,
            created_by=request.user,
        )
        computation.calculate()
        
        messages.success(request, f'Corporate Tax computation created for {fiscal_year.name}. Tax Payable: AED {computation.tax_payable:,.2f}')
        return redirect('finance:corporate_tax_report')
    
    # GET - redirect to report
    return redirect('finance:corporate_tax_report')


@login_required
def corporate_tax_post_provision(request, pk):
    """
    Post Corporate Tax Provision Journal.
    Dr Corporate Tax Expense (P&L)
    Cr Corporate Tax Payable (Liability)
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:corporate_tax_report')
    
    computation = get_object_or_404(CorporateTaxComputation, pk=pk)
    
    try:
        journal = computation.post_provision(user=request.user)
        messages.success(request, f'Tax provision posted. Journal: {journal.entry_number}')
    except ValidationError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f'Error posting provision: {e}')
    
    return redirect('finance:corporate_tax_report')


@login_required
def corporate_tax_recalculate(request, pk):
    """
    Recalculate Corporate Tax Computation from current GL values.
    
    This updates the stored Revenue/Expenses/Accounting Profit to match
    the current General Ledger, ensuring consistency between:
    - Tax Summary
    - Tax Computation
    - Tax History
    
    Adjustments (non-deductible, exempt income, other) are preserved.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:corporate_tax_report')
    
    computation = get_object_or_404(CorporateTaxComputation, pk=pk)
    
    # Cannot recalculate filed/paid computations
    if computation.status in ['filed', 'paid']:
        messages.error(request, f'Cannot recalculate - computation is already {computation.status}.')
        return redirect('finance:corporate_tax_report')
    
    try:
        fiscal_year = computation.fiscal_year
        start_date = fiscal_year.start_date
        end_date = fiscal_year.end_date
        
        # ========================================
        # RECALCULATE FROM CURRENT GL
        # ========================================
        income_accounts = Account.objects.filter(is_active=True, account_type=AccountType.INCOME)
        expense_accounts = Account.objects.filter(is_active=True, account_type=AccountType.EXPENSE)
        
        # Revenue from journal lines
        income_lines = JournalEntryLine.objects.filter(
            account__in=income_accounts,
            journal_entry__status='posted',
            journal_entry__date__gte=start_date,
            journal_entry__date__lte=end_date,
        ).aggregate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit')
        )
        gl_revenue = (income_lines['total_credit'] or Decimal('0.00')) - (income_lines['total_debit'] or Decimal('0.00'))
        
        # Expenses from journal lines
        expense_lines = JournalEntryLine.objects.filter(
            account__in=expense_accounts,
            journal_entry__status='posted',
            journal_entry__date__gte=start_date,
            journal_entry__date__lte=end_date,
        ).aggregate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit')
        )
        gl_expenses = (expense_lines['total_debit'] or Decimal('0.00')) - (expense_lines['total_credit'] or Decimal('0.00'))
        
        # Update stored values
        old_revenue = computation.revenue
        old_expenses = computation.expenses
        
        computation.revenue = gl_revenue
        computation.expenses = gl_expenses
        computation.save(update_fields=['revenue', 'expenses'])
        
        # Recalculate derived values
        computation.calculate()
        
        messages.success(
            request, 
            f'Tax computation for {fiscal_year.name} recalculated from current GL. '
            f'Revenue: {old_revenue:,.2f} → {gl_revenue:,.2f}, '
            f'Expenses: {old_expenses:,.2f} → {gl_expenses:,.2f}'
        )
        
    except Exception as e:
        messages.error(request, f'Error recalculating: {e}')
    
    return redirect('finance:corporate_tax_report')


@login_required
def corporate_tax_pay(request, pk):
    """
    Pay Corporate Tax.
    Dr Corporate Tax Payable (clears liability)
    Cr Bank
    """
    computation = get_object_or_404(CorporateTaxComputation, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:corporate_tax_report')
    
    if request.method == 'POST':
        bank_account_id = request.POST.get('bank_account')
        payment_date = request.POST.get('payment_date')
        reference = request.POST.get('reference', '')
        
        bank_account = BankAccount.objects.filter(pk=bank_account_id, is_active=True).first()
        if not bank_account:
            messages.error(request, 'Invalid bank account.')
            return redirect('finance:corporate_tax_report')
        
        from datetime import datetime
        try:
            if payment_date:
                payment_date = datetime.strptime(payment_date, '%Y-%m-%d').date()
            else:
                payment_date = date.today()
        except ValueError:
            payment_date = date.today()
        
        try:
            journal = computation.post_payment(
                bank_account=bank_account,
                payment_date=payment_date,
                reference=reference,
                user=request.user
            )
            messages.success(request, f'Tax payment of AED {computation.paid_amount:,.2f} recorded. Journal: {journal.entry_number}')
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f'Error processing payment: {e}')
        
        return redirect('finance:corporate_tax_report')
    
    # GET - Show payment form
    bank_accounts = BankAccount.objects.filter(is_active=True)
    context = {
        'title': f'Pay Corporate Tax - {computation.fiscal_year.name}',
        'computation': computation,
        'bank_accounts': bank_accounts,
        'today': date.today().strftime('%Y-%m-%d'),
    }
    return render(request, 'finance/corporate_tax_pay.html', context)


@login_required
def journal_register(request):
    """
    Journal Register - Comprehensive read-only control & audit report.
    Lists ALL journal entries from all sources with full filtering capabilities.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    # Date filters
    start_date = request.GET.get('start_date', date(date.today().year, 1, 1).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    
    # Base queryset with related data for performance
    entries = JournalEntry.objects.filter(
        is_active=True,
        date__gte=start_date,
        date__lte=end_date,
    ).select_related(
        'fiscal_year', 'period', 'posted_by', 'created_by', 'reversal_of'
    ).prefetch_related('lines', 'lines__account')
    
    # Filter by fiscal year
    fiscal_year_id = request.GET.get('fiscal_year')
    if fiscal_year_id:
        entries = entries.filter(fiscal_year_id=fiscal_year_id)
    
    # Filter by period
    period_id = request.GET.get('period')
    if period_id:
        entries = entries.filter(period_id=period_id)
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        entries = entries.filter(status=status)
    
    # Filter by entry type (source module)
    entry_type = request.GET.get('entry_type')
    if entry_type:
        entries = entries.filter(entry_type=entry_type)
    
    # Filter by source module (using actual source_module field)
    source_module = request.GET.get('source_module')
    if source_module:
        # Direct filter on source_module field
        entries = entries.filter(source_module=source_module)
    
    # Filter by account (journals affecting a specific account)
    account_id = request.GET.get('account')
    if account_id:
        entries = entries.filter(lines__account_id=account_id).distinct()
    
    # Filter by created_by
    created_by = request.GET.get('created_by')
    if created_by:
        entries = entries.filter(created_by_id=created_by)
    
    # Search filter (reference, description, entry number)
    search = request.GET.get('search')
    if search:
        entries = entries.filter(
            Q(entry_number__icontains=search) |
            Q(reference__icontains=search) |
            Q(description__icontains=search)
        )
    
    # Amount range filters
    min_amount = request.GET.get('min_amount')
    if min_amount:
        try:
            entries = entries.filter(total_debit__gte=Decimal(min_amount))
        except:
            pass
    
    max_amount = request.GET.get('max_amount')
    if max_amount:
        try:
            entries = entries.filter(total_debit__lte=Decimal(max_amount))
        except:
            pass
    
    # Reversal only filter
    reversal_only = request.GET.get('reversal_only')
    if reversal_only == '1':
        entries = entries.filter(Q(entry_type='reversal') | Q(reversal_of__isnull=False))
    
    # Sorting
    sort_by = request.GET.get('sort', '-date')
    valid_sorts = ['date', '-date', 'entry_number', '-entry_number', 'total_debit', '-total_debit']
    if sort_by in valid_sorts:
        entries = entries.order_by(sort_by, '-created_at')
    else:
        entries = entries.order_by('-date', '-created_at')
    
    # Pagination
    paginator = Paginator(entries, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Calculate summary statistics
    all_entries = JournalEntry.objects.filter(
        is_active=True,
        date__gte=start_date,
        date__lte=end_date,
    )
    
    summary = {
        'total_entries': entries.count(),
        'total_debit': entries.aggregate(Sum('total_debit'))['total_debit__sum'] or Decimal('0.00'),
        'total_credit': entries.aggregate(Sum('total_credit'))['total_credit__sum'] or Decimal('0.00'),
        'posted_count': entries.filter(status='posted').count(),
        'draft_count': entries.filter(status='draft').count(),
        'reversed_count': entries.filter(status='reversed').count(),
    }
    
    # Prepare filter options
    fiscal_years = FiscalYear.objects.filter(is_active=True).order_by('-start_date')
    periods = AccountingPeriod.objects.filter(is_active=True).select_related('fiscal_year').order_by('-start_date')
    accounts = Account.objects.filter(is_active=True).order_by('code')
    users = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    
    status_choices = JournalEntry.STATUS_CHOICES
    entry_type_choices = JournalEntry.ENTRY_TYPE_CHOICES
    
    # Source module choices - match JournalEntry.SOURCE_MODULE_CHOICES
    source_module_choices = [
        ('manual', 'Manual Entry'),
        ('sales', 'Sales Invoice'),
        ('purchase', 'Vendor Bill'),
        ('payment', 'Payment'),
        ('bank_transfer', 'Bank Transfer'),
        ('expense_claim', 'Expense Claim'),
        ('payroll', 'Payroll'),
        ('inventory', 'Inventory'),
        ('fixed_asset', 'Fixed Asset'),
        ('project', 'Project'),
        ('pdc', 'PDC Cheque'),
        ('property', 'Property/Rent'),
        ('vat', 'VAT Adjustment'),
        ('corporate_tax', 'Corporate Tax'),
        ('petty_cash', 'Petty Cash'),
        ('opening_balance', 'Opening Balance'),
        ('year_end', 'Year-End Closing'),
    ]
    
    # Determine source module for each entry (for display) - using actual source_module field
    def get_source_info(entry):
        """Get source module info from entry's source_module field."""
        source_classes = {
            'manual': ('Manual Entry', 'bg-secondary'),
            'sales': ('Sales Invoice', 'bg-success'),
            'purchase': ('Vendor Bill', 'bg-primary'),
            'payment': ('Payment', 'bg-info'),
            'bank_transfer': ('Bank Transfer', 'bg-secondary'),
            'expense_claim': ('Expense Claim', 'bg-warning'),
            'payroll': ('Payroll', 'bg-purple'),
            'inventory': ('Inventory', 'bg-teal'),
            'fixed_asset': ('Fixed Asset', 'bg-dark'),
            'project': ('Project', 'bg-orange'),
            'pdc': ('PDC Cheque', 'bg-pink'),
            'property': ('Property/Rent', 'bg-cyan'),
            'vat': ('VAT Adjustment', 'bg-danger'),
            'corporate_tax': ('Corporate Tax', 'bg-dark'),
            'petty_cash': ('Petty Cash', 'bg-warning'),
            'opening_balance': ('Opening Balance', 'bg-info'),
            'year_end': ('Year-End Closing', 'bg-dark'),
        }
        
        source_module = entry.source_module or 'manual'
        label, css_class = source_classes.get(source_module, ('Unknown', 'bg-light'))
        
        # Override for special entry types
        if entry.entry_type == 'reversal' or entry.reversal_of:
            return (source_module, f'{label} (Reversal)', 'bg-warning')
        if entry.entry_type == 'opening':
            return ('opening_balance', 'Opening Balance', 'bg-info')
        
        return (source_module, label, css_class)
    
    # Add source info to entries
    entries_with_source = []
    for entry in page_obj:
        source_key, source_label, source_class = get_source_info(entry)
        entries_with_source.append({
            'entry': entry,
            'source_key': source_key,
            'source_label': source_label,
            'source_class': source_class,
        })
    
    context = {
        'title': 'Journal Register',
        'entries': entries_with_source,
        'page_obj': page_obj,
        'summary': summary,
        'start_date': start_date,
        'end_date': end_date,
        'fiscal_years': fiscal_years,
        'periods': periods,
        'accounts': accounts,
        'users': users,
        'status_choices': status_choices,
        'entry_type_choices': entry_type_choices,
        'source_module_choices': source_module_choices,
    }
    
    # Export functionality
    export_format = request.GET.get('export')
    if export_format:
        return journal_register_export(request, entries, export_format, start_date, end_date)
    
    return render(request, 'finance/journal_register.html', context)


def journal_register_export(request, entries, export_format, start_date, end_date):
    """Export journal register to various formats."""
    import csv
    from django.http import HttpResponse
    
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="journal_register_{start_date}_to_{end_date}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Entry Number', 'Date', 'Fiscal Year', 'Period', 'Reference', 
            'Description', 'Total Debit', 'Total Credit', 'Status', 
            'Entry Type', 'Created By', 'Created At', 'Posted By', 'Posted At'
        ])
        
        for entry in entries:
            writer.writerow([
                entry.entry_number,
                entry.date,
                entry.fiscal_year.name if entry.fiscal_year else '',
                entry.period.name if entry.period else '',
                entry.reference,
                entry.description,
                entry.total_debit,
                entry.total_credit,
                entry.get_status_display(),
                entry.get_entry_type_display(),
                entry.created_by.get_full_name() if entry.created_by else '',
                entry.created_at,
                entry.posted_by.get_full_name() if entry.posted_by else '',
                entry.posted_date,
            ])
        
        return response
    
    elif export_format == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Journal Register'
            
            # Title
            ws.merge_cells('A1:N1')
            ws['A1'] = f'Journal Register: {start_date} to {end_date}'
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Headers
            headers = [
                'Entry Number', 'Date', 'Fiscal Year', 'Period', 'Reference', 
                'Description', 'Total Debit', 'Total Credit', 'Status', 
                'Entry Type', 'Created By', 'Created At', 'Posted By', 'Posted At'
            ]
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for row_num, entry in enumerate(entries, 4):
                ws.cell(row=row_num, column=1, value=entry.entry_number)
                ws.cell(row=row_num, column=2, value=entry.date)
                ws.cell(row=row_num, column=3, value=entry.fiscal_year.name if entry.fiscal_year else '')
                ws.cell(row=row_num, column=4, value=entry.period.name if entry.period else '')
                ws.cell(row=row_num, column=5, value=entry.reference)
                ws.cell(row=row_num, column=6, value=entry.description)
                ws.cell(row=row_num, column=7, value=float(entry.total_debit))
                ws.cell(row=row_num, column=8, value=float(entry.total_credit))
                ws.cell(row=row_num, column=9, value=entry.get_status_display())
                ws.cell(row=row_num, column=10, value=entry.get_entry_type_display())
                ws.cell(row=row_num, column=11, value=entry.created_by.get_full_name() if entry.created_by else '')
                ws.cell(row=row_num, column=12, value=entry.created_at.strftime('%Y-%m-%d %H:%M') if entry.created_at else '')
                ws.cell(row=row_num, column=13, value=entry.posted_by.get_full_name() if entry.posted_by else '')
                ws.cell(row=row_num, column=14, value=entry.posted_date.strftime('%Y-%m-%d %H:%M') if entry.posted_date else '')
            
            # Auto-width columns
            for col in range(1, 15):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="journal_register_{start_date}_to_{end_date}.xlsx"'
            return response
            
        except ImportError:
            messages.error(request, 'Excel export requires openpyxl library.')
            return redirect('finance:journal_register')
    
    return redirect('finance:journal_register')


@login_required
def journal_register_detail(request, pk):
    """
    Journal Register Detail - Read-only drill-down view of a journal entry.
    Shows full journal details with audit trail.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:journal_register')
    
    entry = get_object_or_404(
        JournalEntry.objects.select_related(
            'fiscal_year', 'period', 'posted_by', 'created_by', 'reversal_of'
        ).prefetch_related('lines', 'lines__account', 'reversed_by'),
        pk=pk
    )
    
    # Determine source module
    ref = entry.reference.upper() if entry.reference else ''
    
    if entry.entry_type == 'opening':
        source_info = {'module': 'Opening Balance', 'class': 'bg-info'}
    elif entry.entry_type == 'reversal' or entry.reversal_of:
        source_info = {'module': 'Reversal', 'class': 'bg-warning'}
    elif ref.startswith('INV') or 'INVOICE' in ref:
        source_info = {'module': 'Sales Invoice', 'class': 'bg-success'}
    elif ref.startswith('BILL') or 'BILL' in ref:
        source_info = {'module': 'Purchase Bill', 'class': 'bg-primary'}
    elif ref.startswith('PR') or ref.startswith('PM') or 'PAYMENT' in ref:
        source_info = {'module': 'Payment', 'class': 'bg-purple'}
    elif ref.startswith('TRANSFER') or 'TRANSFER' in ref:
        source_info = {'module': 'Bank Transfer', 'class': 'bg-secondary'}
    elif ref.startswith('EXPENSE') or 'EXPENSE' in ref:
        source_info = {'module': 'Expense Claim', 'class': 'bg-orange'}
    elif ref.startswith('VAT') or 'VAT' in ref:
        source_info = {'module': 'VAT Adjustment', 'class': 'bg-danger'}
    elif ref.startswith('ADJ') or entry.entry_type == 'adjustment':
        source_info = {'module': 'Adjustment', 'class': 'bg-dark'}
    else:
        source_info = {'module': 'Manual Journal', 'class': 'bg-light text-dark'}
    
    # Get linked records
    linked_records = []
    
    # Check for linked invoice
    if hasattr(entry, 'invoice') and entry.invoice.exists():
        for inv in entry.invoice.all():
            linked_records.append({
                'type': 'Sales Invoice',
                'number': inv.invoice_number,
                'url': f'/sales/invoices/{inv.pk}/',
            })
    
    # Check for linked payments
    if hasattr(entry, 'payments') and entry.payments.exists():
        for payment in entry.payments.all():
            linked_records.append({
                'type': 'Payment',
                'number': payment.payment_number,
                'url': f'/finance/payments/{payment.pk}/edit/',
            })
    
    # Check for linked bank transfers
    if hasattr(entry, 'bank_transfers') and entry.bank_transfers.exists():
        for transfer in entry.bank_transfers.all():
            linked_records.append({
                'type': 'Bank Transfer',
                'number': transfer.transfer_number,
                'url': f'/finance/bank-transfers/',
            })
    
    # Check for linked expense claims
    if hasattr(entry, 'expense_claims') and entry.expense_claims.exists():
        for claim in entry.expense_claims.all():
            linked_records.append({
                'type': 'Expense Claim',
                'number': claim.claim_number,
                'url': f'/finance/expense-claims/{claim.pk}/',
            })
    
    # Check for opening balance entry
    if hasattr(entry, 'opening_balance_entry') and entry.opening_balance_entry.exists():
        for ob in entry.opening_balance_entry.all():
            linked_records.append({
                'type': 'Opening Balance',
                'number': ob.entry_number,
                'url': f'/finance/opening-balances/{ob.pk}/',
            })
    
    # Period lock status
    is_period_locked = entry.period.is_locked if entry.period else False
    
    context = {
        'title': f'Journal Entry: {entry.entry_number}',
        'entry': entry,
        'source_info': source_info,
        'linked_records': linked_records,
        'is_period_locked': is_period_locked,
        'can_edit': (
            entry.status == 'draft' and 
            not is_period_locked and
            (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit'))
        ),
    }
    
    return render(request, 'finance/journal_register_detail.html', context)


# ============ FISCAL YEAR & PERIOD VIEWS ============

class FiscalYearListView(PermissionRequiredMixin, ListView):
    model = FiscalYear
    template_name = 'finance/fiscalyear_list.html'
    context_object_name = 'fiscal_years'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        return FiscalYear.objects.filter(is_active=True)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Fiscal Years'
        context['form'] = FiscalYearForm()
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        context['today'] = date.today().isoformat()
        return context
    
    def post(self, request, *args, **kwargs):
        if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'create')):
            messages.error(request, 'Permission denied.')
            return redirect('finance:fiscalyear_list')
        
        form = FiscalYearForm(request.POST)
        if form.is_valid():
            fy = form.save()
            messages.success(request, f'Fiscal Year {fy.name} created.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
        return redirect('finance:fiscalyear_list')


@login_required
def fiscalyear_close(request, pk):
    """Close a fiscal year."""
    fy = get_object_or_404(FiscalYear, pk=pk)
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:fiscalyear_list')
    
    if fy.is_closed:
        messages.error(request, 'Fiscal year is already closed.')
        return redirect('finance:fiscalyear_list')
    
    fy.close(request.user)
    messages.success(request, f'Fiscal Year {fy.name} closed.')
    return redirect('finance:fiscalyear_list')


class AccountingPeriodListView(PermissionRequiredMixin, ListView):
    model = AccountingPeriod
    template_name = 'finance/period_list.html'
    context_object_name = 'periods'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        return AccountingPeriod.objects.filter(is_active=True).select_related('fiscal_year')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Accounting Periods'
        context['form'] = AccountingPeriodForm()
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        context['today'] = date.today().isoformat()
        return context
    
    def post(self, request, *args, **kwargs):
        if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'create')):
            messages.error(request, 'Permission denied.')
            return redirect('finance:period_list')
        
        form = AccountingPeriodForm(request.POST)
        if form.is_valid():
            period = form.save()
            messages.success(request, f'Period {period.name} created.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
        return redirect('finance:period_list')


@login_required
def period_lock(request, pk):
    """Lock/unlock an accounting period."""
    from django.utils import timezone
    period = get_object_or_404(AccountingPeriod, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:period_list')
    
    period.is_locked = not period.is_locked
    if period.is_locked:
        period.locked_date = timezone.now()
        period.locked_by = request.user
    else:
        period.locked_date = None
        period.locked_by = None
    period.save()
    
    status = 'locked' if period.is_locked else 'unlocked'
    messages.success(request, f'Period {period.name} {status}.')
    return redirect('finance:period_list')


# ============ BANK ACCOUNT VIEWS ============

class BankAccountListView(PermissionRequiredMixin, ListView):
    model = BankAccount
    template_name = 'finance/bankaccount_list.html'
    context_object_name = 'bank_accounts'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        return BankAccount.objects.filter(is_active=True).select_related('gl_account')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Bank Accounts'
        context['form'] = BankAccountForm()
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        
        # Update balances from GL
        for ba in context['bank_accounts']:
            ba.update_balance()
        
        return context
    
    def post(self, request, *args, **kwargs):
        if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'create')):
            messages.error(request, 'Permission denied.')
            return redirect('finance:bankaccount_list')
        
        form = BankAccountForm(request.POST)
        if form.is_valid():
            ba = form.save()
            messages.success(request, f'Bank Account {ba.name} created.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
        return redirect('finance:bankaccount_list')


class BankAccountUpdateView(UpdatePermissionMixin, UpdateView):
    model = BankAccount
    form_class = BankAccountForm
    template_name = 'finance/bankaccount_form.html'
    success_url = reverse_lazy('finance:bankaccount_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Bank Account: {self.object.name}'
        return context
    
    def form_valid(self, form):
        messages.success(self.request, f'Bank Account {form.instance.name} updated.')
        return super().form_valid(form)


# ============ BANK TRANSFER VIEWS ============

class BankTransferListView(PermissionRequiredMixin, ListView):
    model = BankTransfer
    template_name = 'finance/banktransfer_list.html'
    context_object_name = 'transfers'
    module_name = 'finance'
    permission_type = 'view'
    paginate_by = 25
    
    def get_queryset(self):
        return BankTransfer.objects.filter(is_active=True).select_related('from_bank', 'to_bank')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Bank Transfers'
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        context['today'] = date.today().isoformat()
        return context


class BankTransferCreateView(CreatePermissionMixin, CreateView):
    model = BankTransfer
    form_class = BankTransferForm
    template_name = 'finance/banktransfer_form.html'
    success_url = reverse_lazy('finance:banktransfer_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Bank Transfer'
        context['today'] = date.today().isoformat()
        return context
    
    def form_valid(self, form):
        messages.success(self.request, 'Bank Transfer created.')
        return super().form_valid(form)


@login_required
def banktransfer_confirm(request, pk):
    """Confirm a bank transfer and create journal entry."""
    transfer = get_object_or_404(BankTransfer, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:banktransfer_list')
    
    if transfer.status != 'draft':
        messages.error(request, 'Only draft transfers can be confirmed.')
        return redirect('finance:banktransfer_list')
    
    try:
        transfer.confirm(request.user)
        messages.success(request, f'Bank Transfer {transfer.transfer_number} confirmed.')
    except Exception as e:
        messages.error(request, str(e))
    
    return redirect('finance:banktransfer_list')


# ============ EXPENSE CLAIM VIEWS ============

class ExpenseClaimListView(PermissionRequiredMixin, ListView):
    model = ExpenseClaim
    template_name = 'finance/expenseclaim_list.html'
    context_object_name = 'claims'
    module_name = 'finance'
    permission_type = 'view'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = ExpenseClaim.objects.filter(is_active=True).select_related('employee')
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Expense Claims'
        context['status_choices'] = ExpenseClaim.STATUS_CHOICES
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        context['can_approve'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'approve')
        context['today'] = date.today().isoformat()
        return context


class ExpenseClaimCreateView(CreatePermissionMixin, CreateView):
    model = ExpenseClaim
    form_class = ExpenseClaimForm
    template_name = 'finance/expenseclaim_form.html'
    success_url = reverse_lazy('finance:expenseclaim_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Expense Claim'
        context['today'] = date.today().isoformat()
        if self.request.POST:
            context['items_formset'] = ExpenseItemFormSet(self.request.POST, self.request.FILES)
        else:
            context['items_formset'] = ExpenseItemFormSet()
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        items_formset = context['items_formset']
        
        if items_formset.is_valid():
            form.instance.employee = self.request.user
            self.object = form.save()
            items_formset.instance = self.object
            items_formset.save()
            self.object.calculate_totals()
            messages.success(self.request, f'Expense Claim {self.object.claim_number} created.')
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))


class ExpenseClaimDetailView(PermissionRequiredMixin, DetailView):
    model = ExpenseClaim
    template_name = 'finance/expenseclaim_detail.html'
    context_object_name = 'claim'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Expense Claim: {self.object.claim_number}'
        context['can_approve'] = (
            self.request.user.is_superuser or 
            PermissionChecker.has_permission(self.request.user, 'finance', 'approve')
        ) and self.object.status == 'submitted'
        return context


@login_required
def expenseclaim_submit(request, pk):
    """Submit expense claim for approval."""
    claim = get_object_or_404(ExpenseClaim, pk=pk)
    
    if claim.status != 'draft':
        messages.error(request, 'Only draft claims can be submitted.')
        return redirect('finance:expenseclaim_list')
    
    claim.status = 'submitted'
    claim.save()
    messages.success(request, f'Expense Claim {claim.claim_number} submitted for approval.')
    return redirect('finance:expenseclaim_list')


@login_required
def expenseclaim_approve(request, pk):
    """Approve an expense claim."""
    from django.utils import timezone
    claim = get_object_or_404(ExpenseClaim, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'approve')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:expenseclaim_list')
    
    if claim.status != 'submitted':
        messages.error(request, 'Only submitted claims can be approved.')
        return redirect('finance:expenseclaim_list')
    
    claim.status = 'approved'
    claim.approved_by = request.user
    claim.approved_date = timezone.now()
    claim.save()
    
    messages.success(request, f'Expense Claim {claim.claim_number} approved.')
    return redirect('finance:expenseclaim_list')


# ============ BUDGET VIEWS ============

class BudgetListView(PermissionRequiredMixin, ListView):
    model = Budget
    template_name = 'finance/budget_list.html'
    context_object_name = 'budgets'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        return Budget.objects.filter(is_active=True).select_related('fiscal_year')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Budgets'
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        return context


class BudgetCreateView(CreatePermissionMixin, CreateView):
    model = Budget
    form_class = BudgetForm
    template_name = 'finance/budget_form.html'
    success_url = reverse_lazy('finance:budget_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Budget'
        if self.request.POST:
            context['lines_formset'] = BudgetLineFormSet(self.request.POST)
        else:
            context['lines_formset'] = BudgetLineFormSet()
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        lines_formset = context['lines_formset']
        
        if lines_formset.is_valid():
            self.object = form.save()
            lines_formset.instance = self.object
            lines_formset.save()
            messages.success(self.request, f'Budget {self.object.name} created.')
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))


class BudgetDetailView(PermissionRequiredMixin, DetailView):
    model = Budget
    template_name = 'finance/budget_detail.html'
    context_object_name = 'budget'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Budget: {self.object.name}'
        context['can_edit'] = (
            self.request.user.is_superuser or 
            PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        ) and self.object.status == 'draft'
        return context


# ============ VAT RETURN VIEWS ============

class VATReturnListView(PermissionRequiredMixin, ListView):
    model = VATReturn
    template_name = 'finance/vatreturn_list.html'
    context_object_name = 'vat_returns'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        return VATReturn.objects.filter(is_active=True)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'VAT Returns'
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        return context


class VATReturnCreateView(CreatePermissionMixin, CreateView):
    model = VATReturn
    form_class = VATReturnForm
    template_name = 'finance/vatreturn_form.html'
    success_url = reverse_lazy('finance:vatreturn_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create VAT Return'
        context['today'] = date.today().isoformat()
        return context
    
    def form_valid(self, form):
        self.object = form.save()
        self.object.calculate()
        messages.success(self.request, f'VAT Return {self.object.return_number} created.')
        return redirect(self.success_url)


class VATReturnUpdateView(UpdatePermissionMixin, UpdateView):
    """Edit VAT Return - only for draft returns"""
    model = VATReturn
    form_class = VATReturnForm
    template_name = 'finance/vatreturn_form.html'
    context_object_name = 'vat_return'
    module_name = 'finance'
    
    def get_queryset(self):
        # Only allow editing draft returns
        return VATReturn.objects.filter(status='draft')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit VAT Return: {self.object.return_number}'
        context['is_edit'] = True
        return context
    
    def form_valid(self, form):
        self.object = form.save()
        self.object.calculate()
        messages.success(self.request, f'VAT Return {self.object.return_number} updated.')
        return redirect('finance:vatreturn_detail', pk=self.object.pk)


class VATReturnDetailView(PermissionRequiredMixin, DetailView):
    model = VATReturn
    template_name = 'finance/vatreturn_detail.html'
    context_object_name = 'vat_return'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'VAT Return: {self.object.return_number}'
        context['can_edit'] = (
            self.request.user.is_superuser or 
            PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        ) and self.object.status == 'draft'
        context['can_post'] = context['can_edit'] and self.object.can_post
        context['can_reverse'] = (
            self.request.user.is_superuser or 
            PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        ) and self.object.can_reverse
        return context


@login_required
def vatreturn_post(request, pk):
    """
    Post VAT Return - Creates journal entry to clear VAT control accounts.
    
    JOURNAL ENTRY (UAE FTA Compliant):
    Dr Output VAT Control        = Output VAT Amount
    Cr Input VAT Control         = Input VAT Amount
    Cr VAT Payable to FTA        = Net VAT (difference)
    
    This does NOT affect P&L or Corporate Tax calculations.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:vatreturn_list')
    
    vat_return = get_object_or_404(VATReturn, pk=pk)
    
    if not vat_return.can_post:
        messages.error(request, f'VAT Return {vat_return.return_number} cannot be posted. Status must be "draft".')
        return redirect('finance:vatreturn_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            journal = vat_return.post(user=request.user)
            
            # Log audit entry
            from django.contrib.admin.models import LogEntry, CHANGE
            from django.contrib.contenttypes.models import ContentType
            LogEntry.objects.create(
                user_id=request.user.pk,
                content_type_id=ContentType.objects.get_for_model(vat_return).pk,
                object_id=vat_return.pk,
                object_repr=str(vat_return),
                action_flag=CHANGE,
                change_message=f'Posted VAT Return. Journal: {journal.entry_number}. Period: {vat_return.period_start} to {vat_return.period_end}. Net VAT: {vat_return.net_vat}'
            )
            
            messages.success(
                request, 
                f'VAT Return {vat_return.return_number} posted successfully. '
                f'Journal Entry: {journal.entry_number}. '
                f'Net VAT {"Payable" if vat_return.net_vat > 0 else "Refund"}: AED {abs(vat_return.net_vat):,.2f}'
            )
        except Exception as e:
            messages.error(request, f'Error posting VAT Return: {str(e)}')
    
    return redirect('finance:vatreturn_detail', pk=pk)


@login_required
def vatreturn_reverse(request, pk):
    """
    Reverse VAT Return - Creates reversal journal entry.
    
    This:
    1. Creates exact reversal of the posting journal
    2. Unlocks the VAT period
    3. Sets status back to 'draft'
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:vatreturn_list')
    
    vat_return = get_object_or_404(VATReturn, pk=pk)
    
    if not vat_return.can_reverse:
        messages.error(request, f'VAT Return {vat_return.return_number} cannot be reversed. Status must be "posted".')
        return redirect('finance:vatreturn_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            original_journal = vat_return.journal_entry
            reversal = vat_return.reverse(user=request.user)
            
            # Log audit entry
            from django.contrib.admin.models import LogEntry, CHANGE
            from django.contrib.contenttypes.models import ContentType
            LogEntry.objects.create(
                user_id=request.user.pk,
                content_type_id=ContentType.objects.get_for_model(vat_return).pk,
                object_id=vat_return.pk,
                object_repr=str(vat_return),
                action_flag=CHANGE,
                change_message=f'Reversed VAT Return. Original Journal: {original_journal.entry_number}. Reversal Journal: {reversal.entry_number}. Period unlocked.'
            )
            
            messages.success(
                request, 
                f'VAT Return {vat_return.return_number} reversed successfully. '
                f'Reversal Journal: {reversal.entry_number}. '
                f'VAT period {vat_return.period_start} to {vat_return.period_end} is now unlocked.'
            )
        except Exception as e:
            messages.error(request, f'Error reversing VAT Return: {str(e)}')
    
    return redirect('finance:vatreturn_detail', pk=pk)


@login_required
def vatreturn_submit_to_fta(request, pk):
    """
    Submit VAT Return to FTA - Marks as submitted and prevents reversal.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:vatreturn_list')
    
    vat_return = get_object_or_404(VATReturn, pk=pk)
    
    if vat_return.status != 'posted':
        messages.error(request, f'VAT Return must be posted before submitting to FTA.')
        return redirect('finance:vatreturn_detail', pk=pk)
    
    if request.method == 'POST':
        from django.utils import timezone
        
        fta_reference = request.POST.get('fta_reference', '')
        
        vat_return.status = 'submitted'
        vat_return.filed_date = timezone.now()
        vat_return.filed_by = request.user
        vat_return.fta_reference = fta_reference
        vat_return.save(update_fields=['status', 'filed_date', 'filed_by', 'fta_reference'])
        
        # Log audit entry
        from django.contrib.admin.models import LogEntry, CHANGE
        from django.contrib.contenttypes.models import ContentType
        LogEntry.objects.create(
            user_id=request.user.pk,
            content_type_id=ContentType.objects.get_for_model(vat_return).pk,
            object_id=vat_return.pk,
            object_repr=str(vat_return),
            action_flag=CHANGE,
            change_message=f'Submitted VAT Return to FTA. Reference: {fta_reference}. Period: {vat_return.period_start} to {vat_return.period_end}'
        )
        
        messages.success(request, f'VAT Return {vat_return.return_number} submitted to FTA.')
    
    return redirect('finance:vatreturn_detail', pk=pk)


# ============ ADDITIONAL REPORTS ============

@login_required
def cash_flow(request):
    """
    Cash Flow Statement - DIRECT METHOD (IFRS/UAE Compliant)
    
    CRITICAL ACCOUNTING RULES:
    - Shows ONLY actual cash/bank movements
    - EXCLUDES: Ledger balances, AR/AP balances, Depreciation, Provisions, PDC (until cleared)
    - Source: Journal entries that HIT cash/bank accounts from actual transactions
    
    Valid source_modules for Cash Flow:
    - payment, sales, purchase, bank_transfer, payroll, expense_claim, pdc (only clearance)
    
    EXCLUDES source_modules:
    - opening_balance, reversal, adjustment, depreciation, provision, accrual
    
    Classification:
    A. OPERATING: Customer receipts, Supplier payments, Salaries, Taxes, Operating expenses
    B. INVESTING: Fixed asset purchases/sales, Investment activity
    C. FINANCING: Capital, Loans, Drawings, Dividends
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    start_date_str = request.GET.get('start_date', date(date.today().year, 1, 1).isoformat())
    end_date_str = request.GET.get('end_date', date.today().isoformat())
    bank_filter = request.GET.get('bank', '')
    
    try:
        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)
    except ValueError:
        start_date = date(date.today().year, 1, 1)
        end_date = date.today()
    
    # ========================================
    # IDENTIFY CASH & CASH EQUIVALENTS (IFRS IAS 7 Compliant)
    # 
    # IMPORTANT: For Cash Flow to reconcile with Balance Sheet,
    # we MUST include the same accounts in both.
    # 
    # Cash & Cash Equivalents include:
    # - Cash on hand
    # - Bank current accounts
    # - Bank savings accounts
    # - Fixed Deposits (treated as cash equivalents for consistency)
    # 
    # NOTE: Per strict IFRS, Fixed Deposits with maturity > 3 months
    # should be classified as investments. However, for Balance Sheet
    # reconciliation, we include them here.
    # ========================================
    cash_bank_accounts = Account.objects.filter(
        is_active=True,
        is_cash_account=True
        # REMOVED: is_fixed_deposit=False - Now including FDs for BS reconciliation
    ).order_by('code')
    
    # Fallback if no accounts marked as cash accounts
    if not cash_bank_accounts.exists():
        cash_bank_accounts = Account.objects.filter(
            is_active=True,
            account_type=AccountType.ASSET,
        ).filter(
            Q(name__icontains='bank') |
            Q(name__icontains='cash in hand') |
            Q(name__icontains='petty cash') |
            Q(name__icontains='fixed deposit') |  # INCLUDE FD for BS reconciliation
            Q(account_category='cash_bank')
        ).exclude(
            Q(name__icontains='receivable') |
            Q(name__icontains='pdc')  # Exclude PDC Receivable
        ).order_by('code')
    
    if bank_filter:
        cash_bank_accounts = cash_bank_accounts.filter(pk=bank_filter)
    
    cash_account_ids = list(cash_bank_accounts.values_list('id', flat=True))
    
    # ========================================
    # OPENING CASH BALANCE
    # IFRS FIX: Include ALL Cash & Cash Equivalents (including Fixed Deposits)
    # 
    # IMPORTANT: Opening balance includes:
    # 1. Account's opening_balance field
    # 2. All transactions BEFORE start_date
    # 3. Opening balance entries dated ON start_date (these are positions, not activities)
    # ========================================
    opening_cash = Decimal('0.00')
    opening_cash_detail = []  # For validation
    
    # Opening balance entry identifiers (source modules that represent opening positions)
    OPENING_BALANCE_SOURCES = ['opening_balance', 'system_opening', 'system']
    
    for acc in cash_bank_accounts:
        # IFRS FIX: Include Fixed Deposits for Balance Sheet reconciliation
        # Previously excluded FDs, but this caused mismatch with Balance Sheet
        # Now including all cash_bank_accounts (including FDs) for consistency
            
        acc_opening = acc.opening_balance or Decimal('0.00')
        
        # Add all transactions BEFORE start_date
        pre_period = JournalEntryLine.objects.filter(
            account=acc,
            journal_entry__status='posted',
            journal_entry__date__lt=start_date
        ).aggregate(
            total_debit=Coalesce(Sum('debit'), Decimal('0.00')),
            total_credit=Coalesce(Sum('credit'), Decimal('0.00'))
        )
        acc_opening += pre_period['total_debit'] - pre_period['total_credit']
        
        # CRITICAL FIX: Also include opening balance entries dated ON start_date
        # These are opening positions, NOT cash flow activities
        opening_entries_on_start = JournalEntryLine.objects.filter(
            account=acc,
            journal_entry__status='posted',
            journal_entry__date=start_date
        ).filter(
            Q(journal_entry__source_module__in=OPENING_BALANCE_SOURCES) |
            Q(journal_entry__entry_type='opening') |
            Q(journal_entry__entry_type='opening_balance') |  # Added: check for 'opening_balance' type
            Q(journal_entry__description__icontains='opening balance') |  # Added: check description
            Q(journal_entry__reference__icontains='OPENING BALANCE') |
            Q(journal_entry__reference__istartswith='OB-')
        ).aggregate(
            total_debit=Coalesce(Sum('debit'), Decimal('0.00')),
            total_credit=Coalesce(Sum('credit'), Decimal('0.00'))
        )
        acc_opening += opening_entries_on_start['total_debit'] - opening_entries_on_start['total_credit']
        
        opening_cash += acc_opening
        opening_cash_detail.append({'account': acc.name, 'balance': acc_opening})
    
    # ========================================
    # CLOSING CASH BALANCE
    # IFRS FIX: Include ALL Cash & Cash Equivalents (including Fixed Deposits)
    # This ensures Cash Flow closing matches Balance Sheet cash total
    # ========================================
    closing_cash = Decimal('0.00')
    closing_cash_detail = []  # For reconciliation validation
    for acc in cash_bank_accounts:
        # IFRS FIX: Include Fixed Deposits for Balance Sheet reconciliation
        # Previously excluded FDs, now including for consistency
            
        acc_balance = acc.opening_balance or Decimal('0.00')
        period_totals = JournalEntryLine.objects.filter(
            account=acc,
            journal_entry__status='posted',
            journal_entry__date__lte=end_date
        ).aggregate(
            total_debit=Coalesce(Sum('debit'), Decimal('0.00')),
            total_credit=Coalesce(Sum('credit'), Decimal('0.00'))
        )
        acc_balance += period_totals['total_debit'] - period_totals['total_credit']
        closing_cash += acc_balance
        closing_cash_detail.append({'account': acc.name, 'balance': acc_balance})
    
    # ========================================
    # VALID SOURCE MODULES FOR CASH FLOW
    # (Only actual cash movements)
    # ========================================
    VALID_CASH_SOURCES = [
        'payment', 'sales', 'purchase', 'bank_transfer', 'payroll', 
        'expense_claim', 'petty_cash', 'manual', 'pdc',
        'property', 'project', 'inventory', 'fixed_asset',
    ]
    
    # Excluded sources (non-cash entries)
    EXCLUDED_SOURCES = [
        'opening_balance', 'reversal', 'adjustment', 'depreciation',
        'provision', 'accrual', 'corporate_tax', 'vat', 'year_end',
        'system', 'system_opening',  # Add system-generated entries
    ]
    
    # ========================================
    # GET CASH MOVEMENTS (Direct Method)
    # CRITICAL: EXCLUDE ALL OPENING BALANCE ENTRIES
    # Opening balances are POSITIONS, not TRANSACTIONS
    # ========================================
    
    # Get journals that:
    # 1. Have a line hitting cash/bank account
    # 2. Are posted
    # 3. Are within date range
    # 4. Are NOT from excluded sources
    # 5. Are NOT opening balance entries (by reference, type, or description)
    
    cash_journal_lines = JournalEntryLine.objects.filter(
        account_id__in=cash_account_ids,
        journal_entry__status='posted',
        journal_entry__date__gte=start_date,
        journal_entry__date__lte=end_date
    ).exclude(
        journal_entry__source_module__in=EXCLUDED_SOURCES
    ).exclude(
        # Exclude journals that are only accrual/depreciation related
        journal_entry__entry_type__in=['adjusting', 'closing', 'opening', 'opening_balance']  # Include 'opening_balance'
    ).exclude(
        # CRITICAL: Exclude opening balance by reference patterns
        journal_entry__reference__icontains='OPENING BALANCE'
    ).exclude(
        journal_entry__reference__istartswith='OB-'
    ).exclude(
        # Exclude opening balance by description
        journal_entry__description__icontains='Opening Balance'
    ).exclude(
        # Exclude system-generated opening entries
        journal_entry__is_system_generated=True,
        journal_entry__entry_type='opening'
    ).select_related('journal_entry', 'account').order_by('journal_entry__date')
    
    # Process cash movements
    operating_items = []
    investing_items = []
    financing_items = []
    
    operating_total = Decimal('0.00')
    investing_total = Decimal('0.00')
    financing_total = Decimal('0.00')
    
    processed_journals = set()
    
    for cash_line in cash_journal_lines:
        journal = cash_line.journal_entry
        
        # Skip if already processed
        if journal.pk in processed_journals:
            continue
        processed_journals.add(journal.pk)
        
        # Calculate net cash movement (Debit = Cash In, Credit = Cash Out for asset accounts)
        journal_cash_movement = Decimal('0.00')
        for line in journal.lines.filter(account_id__in=cash_account_ids):
            journal_cash_movement += line.debit - line.credit
        
        # Skip zero movements
        if journal_cash_movement == 0:
            continue
        
        # Get counter-party accounts (non-cash accounts in this journal)
        counter_lines = journal.lines.exclude(account_id__in=cash_account_ids)
        
        if not counter_lines.exists():
            # Check if this is a Fixed Deposit movement (cash-to-FD or FD-to-cash)
            # These should appear in Investing Activities per IAS 7
            fd_lines = journal.lines.filter(account__is_fixed_deposit=True)
            non_fd_cash_lines = journal.lines.filter(account_id__in=cash_account_ids).exclude(account__is_fixed_deposit=True)
            
            if fd_lines.exists() and non_fd_cash_lines.exists():
                # This is a Fixed Deposit transfer - show in Investing Activities
                fd_movement = Decimal('0.00')
                for fd_line in fd_lines:
                    # Debit to FD = investing outflow (buying FD)
                    # Credit to FD = investing inflow (FD matured/withdrawn)
                    fd_movement += fd_line.credit - fd_line.debit
                
                if fd_movement != 0:
                    fd_account = fd_lines.first().account
                    item = {
                        'date': journal.date,
                        'reference': journal.reference or journal.entry_number,
                        'description': journal.description or f"Fixed Deposit {'withdrawal' if fd_movement > 0 else 'placement'}",
                        'counter_account': f"{fd_account.code} - {fd_account.name}",
                        'counter_account_type': 'asset',
                        'amount': fd_movement,
                        'journal_id': journal.pk,
                        'source_module': journal.source_module,
                        'is_inflow': fd_movement > 0,
                        'category': 'Fixed deposit withdrawal' if fd_movement > 0 else 'Fixed deposit placement',
                    }
                    investing_items.append(item)
                    investing_total += fd_movement
            # Skip other cash-to-cash transfers
            continue
        
        # Determine classification based on counter account
        primary_counter = counter_lines.first()
        counter_account = primary_counter.account
        counter_type = counter_account.account_type
        counter_name = counter_account.name.lower()
        counter_category = getattr(counter_account, 'account_category', '') or ''
        
        # Skip non-cash entries (double-check)
        non_cash_keywords = ['depreciation', 'amortization', 'provision', 'accrual', 'allowance']
        if any(kw in counter_name for kw in non_cash_keywords):
            continue
        
        # NOTE: PDC Receivable clearance IS actual cash (when PDC cheque clears)
        # Dr Bank (cash in), Cr PDC Receivable (receivable reduces)
        # This should be classified as "Cash received from customers" - do NOT skip!
        
        # Build item
        # Positive = Cash In, Negative = Cash Out
        item = {
            'date': journal.date,
            'reference': journal.reference or journal.entry_number,
            'description': journal.description or primary_counter.description or counter_account.name,
            'counter_account': f"{counter_account.code} - {counter_account.name}",
            'counter_account_type': counter_type,
            'amount': journal_cash_movement,
            'journal_id': journal.pk,
            'source_module': journal.source_module,
            'is_inflow': journal_cash_movement > 0,
        }
        
        # ========================================
        # CLASSIFICATION LOGIC (Direct Method)
        # ========================================
        
        # A. OPERATING ACTIVITIES
        # - Customer receipts (AR clearing)
        # - Supplier payments (AP clearing)
        # - Salary payments
        # - Tax payments
        # - Operating expense payments
        
        if counter_type == AccountType.INCOME:
            # Direct cash sales
            item['category'] = 'Cash received from customers'
            operating_items.append(item)
            operating_total += journal_cash_movement
            
        elif counter_type == AccountType.EXPENSE:
            # Direct expense payments
            expense_category = 'Cash paid for operating expenses'
            if 'salary' in counter_name or 'wage' in counter_name or 'payroll' in counter_name:
                expense_category = 'Cash paid to employees'
            elif 'rent' in counter_name:
                expense_category = 'Cash paid for rent'
            elif 'utility' in counter_name or 'electric' in counter_name:
                expense_category = 'Cash paid for utilities'
            elif 'bank charge' in counter_name:
                expense_category = 'Bank charges paid'
            item['category'] = expense_category
            operating_items.append(item)
            operating_total += journal_cash_movement
            
        elif (
            'receivable' in counter_name or 
            'debtor' in counter_name or 
            counter_category == 'trade_receivables' or
            counter_name.startswith('ar ') or  # "AR - Customer A"
            counter_name.startswith('ar-') or  # "AR-Customer"
            counter_account.code.startswith('11')  # 11xx = Receivables (standard COA)
        ):
            # AR clearing = Customer payment received
            # PDC Receivable clearance = PDC cheque deposited and cleared
            # IFRS: Once PDC clears, it's normal customer cash receipt - merge into single category
            # Internal tracking via description, external presentation as single line
            item['category'] = 'Cash received from customers'
            operating_items.append(item)
            operating_total += journal_cash_movement
            
        elif (
            'payable' in counter_name or 
            'creditor' in counter_name or 
            counter_category == 'trade_payables' or
            counter_name.startswith('ap ') or  # "AP - Vendor A"
            counter_name.startswith('ap-') or  # "AP-Vendor"
            counter_account.code.startswith('21') or  # 21xx = Payables (standard COA)
            'accrued' in counter_name  # Accrued Expenses = Operating
        ):
            # AP clearing = Payment to supplier
            # Accrued expenses clearing = Operating expense payment
            if 'vat' in counter_name or 'tax' in counter_name:
                item['category'] = 'VAT/Tax paid'
            elif 'salary' in counter_name or 'employee' in counter_name or 'wage' in counter_name:
                item['category'] = 'Cash paid to employees'
            elif 'accrued' in counter_name:
                item['category'] = 'Cash paid for accrued expenses'
            else:
                item['category'] = 'Cash paid to suppliers'
            operating_items.append(item)
            operating_total += journal_cash_movement
            
        elif 'vat' in counter_name or counter_category in ['tax_payables', 'tax_receivables']:
            # VAT payments/refunds
            item['category'] = 'VAT paid / (received)'
            operating_items.append(item)
            operating_total += journal_cash_movement
        
        # B. INVESTING ACTIVITIES
        # - Fixed asset purchases
        # - Fixed asset sales
        # - Investment purchases/sales
        
        elif counter_category in ['fixed_furniture', 'fixed_it', 'fixed_vehicles', 'fixed_other', 'intangible']:
            # Fixed asset purchase
            item['category'] = 'Purchase of fixed assets'
            investing_items.append(item)
            investing_total += journal_cash_movement
            
        elif 'fixed asset' in counter_name or 'furniture' in counter_name or 'equipment' in counter_name or 'vehicle' in counter_name:
            item['category'] = 'Purchase of fixed assets' if journal_cash_movement < 0 else 'Sale of fixed assets'
            investing_items.append(item)
            investing_total += journal_cash_movement
            
        elif 'investment' in counter_name:
            item['category'] = 'Investment activity'
            investing_items.append(item)
            investing_total += journal_cash_movement
        
        # C. FINANCING ACTIVITIES
        # - Capital introduced / contributions
        # - Loans received/repaid
        # - Drawings / withdrawals
        # - Dividends paid
        # - Share capital / equity issued
        
        elif counter_type == AccountType.EQUITY or counter_category in ['capital', 'reserves', 'retained_earnings']:
            if 'drawing' in counter_name or 'withdrawal' in counter_name:
                item['category'] = 'Owner drawings / withdrawals'
            elif 'dividend' in counter_name:
                item['category'] = 'Dividends paid'
            elif 'capital' in counter_name or 'partner' in counter_name or 'owner' in counter_name:
                item['category'] = 'Capital contributed by owners' if journal_cash_movement > 0 else 'Capital withdrawn'
            elif 'share' in counter_name:
                item['category'] = 'Share capital issued' if journal_cash_movement > 0 else 'Share buyback'
            elif 'retained' in counter_name:
                item['category'] = 'Retained earnings adjustment'
            else:
                item['category'] = 'Other financing - Equity'
            financing_items.append(item)
            financing_total += journal_cash_movement
            
        elif 'loan' in counter_name or 'borrowing' in counter_name or 'mortgage' in counter_name:
            item['category'] = 'Proceeds from loans' if journal_cash_movement > 0 else 'Loan repayments'
            financing_items.append(item)
            financing_total += journal_cash_movement
            
        elif counter_type == AccountType.LIABILITY and ('long term' in counter_name or 'term loan' in counter_name):
            # Long-term liability typically = financing
            item['category'] = 'Financing - Long term liability'
            financing_items.append(item)
            financing_total += journal_cash_movement
        
        # Default to operating (but NOT opening balance entries)
        else:
            # Double-check this isn't an opening balance entry
            ref_lower = (journal.reference or '').lower()
            desc_lower = (journal.description or '').lower()
            if 'opening' in ref_lower or 'opening' in desc_lower or ref_lower.startswith('ob-'):
                continue  # Skip opening balance entries
            
            item['category'] = 'Other operating cash flow'
            operating_items.append(item)
            operating_total += journal_cash_movement
    
    # ========================================
    # CALCULATE EXCLUDED ADJUSTMENTS
    # (Reversals, corrections that hit cash but are NOT activities)
    # 
    # IMPORTANT: Exclude opening balance entries as they're included in opening_cash
    # ========================================
    excluded_adjustment_lines = JournalEntryLine.objects.filter(
        account_id__in=cash_account_ids,
        journal_entry__status='posted',
        journal_entry__date__gte=start_date,
        journal_entry__date__lte=end_date
    ).filter(
        Q(journal_entry__source_module__in=EXCLUDED_SOURCES) |
        Q(journal_entry__entry_type__in=['adjusting', 'closing'])
    ).exclude(
        # Don't count opening balance entries (they're part of opening balance, not period activity)
        journal_entry__entry_type='opening'
    ).exclude(
        journal_entry__source_module__in=OPENING_BALANCE_SOURCES
    ).exclude(
        journal_entry__reference__icontains='OPENING BALANCE'
    ).exclude(
        journal_entry__reference__istartswith='OB-'
    ).select_related('journal_entry', 'account')
    
    excluded_adjustments = Decimal('0.00')
    excluded_adjustment_details = []  # For transparency in reporting
    excluded_journals_processed = set()
    for line in excluded_adjustment_lines:
        j = line.journal_entry
        if j.pk not in excluded_journals_processed:
            excluded_journals_processed.add(j.pk)
            journal_adjustment = Decimal('0.00')
            for cash_line in j.lines.filter(account_id__in=cash_account_ids):
                journal_adjustment += cash_line.debit - cash_line.credit
            if journal_adjustment != Decimal('0.00'):
                excluded_adjustments += journal_adjustment
                excluded_adjustment_details.append({
                    'reference': j.reference,
                    'description': j.description,
                    'date': j.date,
                    'amount': journal_adjustment,
                    'source_module': j.source_module,
                })
    
    # ========================================
    # VALIDATION: Opening + Net Change + Adjustments = Closing
    # ========================================
    net_change = operating_total + investing_total + financing_total
    calculated_closing = opening_cash + net_change + excluded_adjustments
    validation_ok = abs(calculated_closing - closing_cash) < Decimal('0.01')
    validation_difference = closing_cash - calculated_closing
    
    # ========================================
    # GROUP BY CATEGORY FOR DISPLAY
    # ========================================
    def group_by_category(items):
        grouped = {}
        for item in items:
            key = item.get('category', 'Other')
            if key not in grouped:
                grouped[key] = {'category': key, 'items': [], 'total': Decimal('0.00')}
            grouped[key]['items'].append(item)
            grouped[key]['total'] += item['amount']
        return grouped
    
    operating_grouped = group_by_category(operating_items)
    investing_grouped = group_by_category(investing_items)
    financing_grouped = group_by_category(financing_items)
    
    # Prepare summary for export
    operating_summary = [
        {'description': data['category'], 'amount': data['total']} 
        for data in operating_grouped.values()
    ]
    investing_summary = [
        {'description': data['category'], 'amount': data['total']} 
        for data in investing_grouped.values()
    ]
    financing_summary = [
        {'description': data['category'], 'amount': data['total']} 
        for data in financing_grouped.values()
    ]
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel':
        from .excel_exports import export_cash_flow
        return export_cash_flow(
            operating=operating_summary, 
            investing=investing_summary, 
            financing=financing_summary, 
            start_date=start_date_str, 
            end_date=end_date_str,
            opening_balance=opening_cash,
            closing_balance=closing_cash,
            opening_detail=opening_cash_detail,
            closing_detail=closing_cash_detail,
            excluded_adjustments=excluded_adjustments
        )
    
    # Get all cash accounts for filter dropdown (include Fixed Deposits for consistency)
    all_cash_bank_accounts = Account.objects.filter(
        is_active=True,
    ).filter(
        Q(is_cash_account=True) |
        Q(name__icontains='bank') |
        Q(name__icontains='cash in hand') |
        Q(name__icontains='fixed deposit') |  # INCLUDE Fixed Deposits for BS reconciliation
        Q(account_category='cash_bank')
    ).exclude(
        Q(name__icontains='receivable') |
        Q(name__icontains='pdc')
    ).order_by('code')
    
    return render(request, 'finance/cash_flow.html', {
        'title': 'Cash Flow Statement - Direct Method (IFRS)',
        'opening_cash': opening_cash,
        'closing_cash': closing_cash,
        'calculated_closing': calculated_closing,
        'operating_total': operating_total,
        'investing_total': investing_total,
        'financing_total': financing_total,
        'net_change': net_change,
        'operating_grouped': operating_grouped,
        'investing_grouped': investing_grouped,
        'financing_grouped': financing_grouped,
        'operating_items': operating_items,
        'investing_items': investing_items,
        'financing_items': financing_items,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'bank_filter': bank_filter,
        'cash_bank_accounts': all_cash_bank_accounts,
        'validation_ok': validation_ok,
        'validation_difference': validation_difference,
        # Audit details for reconciliation
        'opening_cash_detail': opening_cash_detail,
        'closing_cash_detail': closing_cash_detail,
        'excluded_adjustments': excluded_adjustments,
        'excluded_adjustment_details': excluded_adjustment_details,
    })


@login_required
def ar_aging(request):
    """
    Accounts Receivable Aging Report.
    SINGLE SOURCE OF TRUTH: Reads only from JournalEntryLine (AR account).
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    # Support date filter
    as_of_date_str = request.GET.get('date', '')
    if as_of_date_str:
        try:
            today = date.fromisoformat(as_of_date_str)
        except ValueError:
            today = date.today()
    else:
        today = date.today()
    
    # Get AR account (typically 1200 or similar)
    ar_account = Account.objects.filter(
        code__startswith='12', account_type=AccountType.ASSET, is_active=True
    ).first()
    
    if not ar_account:
        messages.warning(request, 'Accounts Receivable account not found in Chart of Accounts.')
        return render(request, 'finance/ar_aging.html', {
            'title': 'Accounts Receivable Aging',
            'aging_data': {'current': [], '1_30': [], '31_60': [], '61_90': [], 'over_90': []},
            'totals': {'current': Decimal('0.00'), '1_30': Decimal('0.00'), '31_60': Decimal('0.00'), '61_90': Decimal('0.00'), 'over_90': Decimal('0.00')},
            'grand_total': Decimal('0.00'),
            'as_of_date': today,
        })
    
    # Get all open AR items (invoices) from journal lines
    # Open items = Journal lines where there's a debit to AR without matching credit (payment)
    ar_lines = JournalEntryLine.objects.filter(
        account=ar_account,
        journal_entry__status='posted',
        debit__gt=0,
    ).select_related('journal_entry').order_by('journal_entry__date')
    
    # Build aging data from AR journal lines
    aging_data = {
        'current': [],
        '1_30': [],
        '31_60': [],
        '61_90': [],
        'over_90': [],
    }
    
    totals = {
        'current': Decimal('0.00'),
        '1_30': Decimal('0.00'),
        '31_60': Decimal('0.00'),
        '61_90': Decimal('0.00'),
        'over_90': Decimal('0.00'),
    }
    
    # Group by reference (invoice number) and calculate net outstanding
    invoice_balances = {}
    for line in ar_lines:
        ref = line.journal_entry.reference
        if ref not in invoice_balances:
            invoice_balances[ref] = {
                'date': line.journal_entry.date,
                'reference': ref,
                'description': line.description or line.journal_entry.description,
                'debit': Decimal('0.00'),
                'credit': Decimal('0.00'),
            }
        invoice_balances[ref]['debit'] += line.debit
    
    # Get credits (payments) to AR
    ar_credits = JournalEntryLine.objects.filter(
        account=ar_account,
        journal_entry__status='posted',
        credit__gt=0,
    ).select_related('journal_entry')
    
    for line in ar_credits:
        ref = line.journal_entry.reference
        if ref in invoice_balances:
            invoice_balances[ref]['credit'] += line.credit
    
    # Calculate aging
    for ref, data in invoice_balances.items():
        outstanding = data['debit'] - data['credit']
        if outstanding <= 0:
            continue  # Fully paid
        
        days_old = (today - data['date']).days
        
        item_data = {
            'reference': ref,
            'description': data['description'],
            'date': data['date'],
            'total': data['debit'],
            'paid': data['credit'],
            'outstanding': outstanding,
            'days_old': days_old,
        }
        
        if days_old <= 30:
            aging_data['current'].append(item_data)
            totals['current'] += outstanding
        elif days_old <= 60:
            aging_data['1_30'].append(item_data)
            totals['1_30'] += outstanding
        elif days_old <= 90:
            aging_data['31_60'].append(item_data)
            totals['31_60'] += outstanding
        elif days_old <= 120:
            aging_data['61_90'].append(item_data)
            totals['61_90'] += outstanding
        else:
            aging_data['over_90'].append(item_data)
            totals['over_90'] += outstanding
    
    grand_total = sum(totals.values())
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel':
        from .excel_exports import export_ar_aging
        # Flatten data for export
        customer_data = []
        for bucket, items in aging_data.items():
            for item in items:
                customer_data.append({
                    'name': item['reference'],
                    'current': item['outstanding'] if bucket == 'current' else 0,
                    'days_1_30': item['outstanding'] if bucket == '1_30' else 0,
                    'days_31_60': item['outstanding'] if bucket == '31_60' else 0,
                    'days_61_90': item['outstanding'] if bucket == '61_90' else 0,
                    'days_over_90': item['outstanding'] if bucket == 'over_90' else 0,
                    'total': item['outstanding'],
                })
        return export_ar_aging(customer_data, today.isoformat())
    
    return render(request, 'finance/ar_aging.html', {
        'title': 'Accounts Receivable Aging',
        'aging_data': aging_data,
        'totals': totals,
        'grand_total': grand_total,
        'as_of_date': today,
        'ar_account': ar_account,
    })


@login_required
def ap_aging(request):
    """
    Accounts Payable Aging Report.
    SINGLE SOURCE OF TRUTH: Reads only from JournalEntryLine (AP account).
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    # Support date filter
    as_of_date_str = request.GET.get('date', '')
    if as_of_date_str:
        try:
            today = date.fromisoformat(as_of_date_str)
        except ValueError:
            today = date.today()
    else:
        today = date.today()
    
    # Get AP account (typically 2000 or similar)
    ap_account = Account.objects.filter(
        code__startswith='20', account_type=AccountType.LIABILITY, is_active=True
    ).first()
    
    if not ap_account:
        messages.warning(request, 'Accounts Payable account not found in Chart of Accounts.')
        return render(request, 'finance/ap_aging.html', {
            'title': 'Accounts Payable Aging',
            'aging_data': {'current': [], '1_30': [], '31_60': [], '61_90': [], 'over_90': []},
            'totals': {'current': Decimal('0.00'), '1_30': Decimal('0.00'), '31_60': Decimal('0.00'), '61_90': Decimal('0.00'), 'over_90': Decimal('0.00')},
            'grand_total': Decimal('0.00'),
            'as_of_date': today,
        })
    
    # Get all open AP items (bills) from journal lines
    # Open items = Journal lines where there's a credit to AP without matching debit (payment)
    ap_lines = JournalEntryLine.objects.filter(
        account=ap_account,
        journal_entry__status='posted',
        credit__gt=0,
    ).select_related('journal_entry').order_by('journal_entry__date')
    
    # Build aging data from AP journal lines
    aging_data = {
        'current': [],
        '1_30': [],
        '31_60': [],
        '61_90': [],
        'over_90': [],
    }
    
    totals = {
        'current': Decimal('0.00'),
        '1_30': Decimal('0.00'),
        '31_60': Decimal('0.00'),
        '61_90': Decimal('0.00'),
        'over_90': Decimal('0.00'),
    }
    
    # Group by reference (bill number) and calculate net outstanding
    bill_balances = {}
    for line in ap_lines:
        ref = line.journal_entry.reference
        if ref not in bill_balances:
            bill_balances[ref] = {
                'date': line.journal_entry.date,
                'reference': ref,
                'description': line.description or line.journal_entry.description,
                'debit': Decimal('0.00'),
                'credit': Decimal('0.00'),
            }
        bill_balances[ref]['credit'] += line.credit
    
    # Get debits (payments) to AP
    ap_debits = JournalEntryLine.objects.filter(
        account=ap_account,
        journal_entry__status='posted',
        debit__gt=0,
    ).select_related('journal_entry')
    
    for line in ap_debits:
        ref = line.journal_entry.reference
        if ref in bill_balances:
            bill_balances[ref]['debit'] += line.debit
    
    # Calculate aging
    for ref, data in bill_balances.items():
        outstanding = data['credit'] - data['debit']
        if outstanding <= 0:
            continue  # Fully paid
        
        days_old = (today - data['date']).days
        
        item_data = {
            'reference': ref,
            'description': data['description'],
            'date': data['date'],
            'total': data['credit'],
            'paid': data['debit'],
            'outstanding': outstanding,
            'days_old': days_old,
        }
        
        if days_old <= 30:
            aging_data['current'].append(item_data)
            totals['current'] += outstanding
        elif days_old <= 60:
            aging_data['1_30'].append(item_data)
            totals['1_30'] += outstanding
        elif days_old <= 90:
            aging_data['31_60'].append(item_data)
            totals['31_60'] += outstanding
        elif days_old <= 120:
            aging_data['61_90'].append(item_data)
            totals['61_90'] += outstanding
        else:
            aging_data['over_90'].append(item_data)
            totals['over_90'] += outstanding
    
    grand_total = sum(totals.values())
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel':
        from .excel_exports import export_ap_aging
        # Flatten data for export
        vendor_data = []
        for bucket, items in aging_data.items():
            for item in items:
                vendor_data.append({
                    'name': item['reference'],
                    'current': item['outstanding'] if bucket == 'current' else 0,
                    'days_1_30': item['outstanding'] if bucket == '1_30' else 0,
                    'days_31_60': item['outstanding'] if bucket == '31_60' else 0,
                    'days_61_90': item['outstanding'] if bucket == '61_90' else 0,
                    'days_over_90': item['outstanding'] if bucket == 'over_90' else 0,
                    'total': item['outstanding'],
                })
        return export_ap_aging(vendor_data, today.isoformat())
    
    return render(request, 'finance/ap_aging.html', {
        'title': 'Accounts Payable Aging',
        'aging_data': aging_data,
        'totals': totals,
        'grand_total': grand_total,
        'as_of_date': today,
        'ap_account': ap_account,
    })


@login_required
def bank_ledger(request):
    """Bank Ledger - Transactions for a specific bank account."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    bank_id = request.GET.get('bank')
    start_date = request.GET.get('start_date', date(date.today().year, 1, 1).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    
    bank_accounts = BankAccount.objects.filter(is_active=True)
    selected_bank = None
    transactions = []
    running_balance = Decimal('0.00')
    
    if bank_id:
        selected_bank = get_object_or_404(BankAccount, pk=bank_id)
        gl_account = selected_bank.gl_account
        running_balance = gl_account.opening_balance
        
        lines = JournalEntryLine.objects.filter(
            account=gl_account,
            journal_entry__status='posted',
            journal_entry__date__gte=start_date,
            journal_entry__date__lte=end_date,
        ).select_related('journal_entry').order_by('journal_entry__date', 'id')
        
        for line in lines:
            running_balance += line.debit - line.credit
            
            transactions.append({
                'date': line.journal_entry.date,
                'journal_pk': line.journal_entry.pk,
                'entry_number': line.journal_entry.entry_number,
                'reference': line.journal_entry.reference,
                'description': line.description or line.journal_entry.description,
                'debit': line.debit,
                'credit': line.credit,
                'balance': running_balance,
            })
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel' and selected_bank:
        from .excel_exports import export_bank_ledger
        return export_bank_ledger(transactions, selected_bank.name, start_date, end_date)
    
    return render(request, 'finance/bank_ledger.html', {
        'title': 'Bank Ledger',
        'bank_accounts': bank_accounts,
        'selected_bank': selected_bank,
        'transactions': transactions,
        'opening_balance': selected_bank.gl_account.opening_balance if selected_bank else Decimal('0.00'),
        'closing_balance': running_balance,
        'start_date': start_date,
        'end_date': end_date,
    })


@login_required
def budget_vs_actual(request):
    """Budget vs Actual Report."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    budget_id = request.GET.get('budget')
    budgets = Budget.objects.filter(is_active=True, status__in=['approved', 'locked'])
    
    selected_budget = None
    comparison_data = []
    total_budget = Decimal('0.00')
    total_actual = Decimal('0.00')
    
    if budget_id:
        selected_budget = get_object_or_404(Budget, pk=budget_id)
        
        for line in selected_budget.lines.all().select_related('account'):
            account = line.account
            actual = abs(account.balance) if account.balance else Decimal('0.00')
            budgeted = line.amount
            variance = budgeted - actual
            variance_pct = (variance / budgeted * 100) if budgeted else Decimal('0.00')
            
            comparison_data.append({
                'account': account,
                'budgeted': budgeted,
                'actual': actual,
                'variance': variance,
                'variance_pct': variance_pct,
                'is_over': actual > budgeted,
            })
            
            total_budget += budgeted
            total_actual += actual
    
    total_variance = total_budget - total_actual
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel' and selected_budget:
        from .excel_exports import export_budget_vs_actual
        export_data = [{
            'account': f"{d['account'].code} - {d['account'].name}",
            'budget': d['budgeted'],
            'actual': d['actual'],
            'variance': d['variance'],
            'variance_pct': float(d['variance_pct']),
        } for d in comparison_data]
        return export_budget_vs_actual(export_data, selected_budget.name, selected_budget.fiscal_year.name)
    
    return render(request, 'finance/budget_vs_actual.html', {
        'title': 'Budget vs Actual',
        'budgets': budgets,
        'selected_budget': selected_budget,
        'comparison_data': comparison_data,
        'total_budget': total_budget,
        'total_actual': total_actual,
        'total_variance': total_variance,
    })


@login_required
def payment_post(request, pk):
    """
    Post a payment and create journal entry.
    Uses Account Mapping (SAP/Oracle-style Account Determination) for account selection.
    
    Payment Received: Dr Bank, Cr AR (clearing entry)
    Payment Made: Dr AP, Cr Bank (clearing entry)
    """
    payment = get_object_or_404(Payment, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:payment_list')
    
    if payment.status != 'draft':
        messages.error(request, 'Only draft payments can be posted.')
        return redirect('finance:payment_list')
    
    if not payment.bank_account:
        messages.error(request, 'Bank account is required to post payment.')
        return redirect('finance:payment_list')
    
    # Create journal entry
    journal = JournalEntry.objects.create(
        date=payment.payment_date,
        reference=payment.payment_number,
        description=f"{payment.get_payment_type_display()}: {payment.party_name}",
        entry_type='standard',
        source_module='payment',
    )
    
    # Get AR/AP accounts using Account Mapping (SAP/Oracle standard)
    # Fallback to hardcoded codes for backward compatibility
    ar_account = AccountMapping.get_account_or_default('customer_receipt_ar_clear', '1200')
    if not ar_account:
        ar_account = Account.objects.filter(code__startswith='12', account_type='asset').first()
    
    ap_account = AccountMapping.get_account_or_default('vendor_payment_ap_clear', '2000')
    if not ap_account:
        ap_account = Account.objects.filter(code__startswith='20', account_type='liability').first()
    
    bank_account = payment.bank_account.gl_account
    
    if payment.payment_type == 'received':
        # Debit Bank, Credit AR (clears receivable)
        JournalEntryLine.objects.create(
            journal_entry=journal,
            account=bank_account,
            description=f"Payment from {payment.party_name}",
            debit=payment.amount,
        )
        if ar_account:
            JournalEntryLine.objects.create(
                journal_entry=journal,
                account=ar_account,
                description=f"Payment from {payment.party_name}",
                credit=payment.amount,
            )
        else:
            messages.warning(request, 'Accounts Receivable account not configured in Account Mapping.')
    else:
        # Debit AP (clears payable), Credit Bank
        if ap_account:
            JournalEntryLine.objects.create(
                journal_entry=journal,
                account=ap_account,
                description=f"Payment to {payment.party_name}",
                debit=payment.amount,
            )
        else:
            messages.warning(request, 'Accounts Payable account not configured in Account Mapping.')
        JournalEntryLine.objects.create(
            journal_entry=journal,
            account=bank_account,
            description=f"Payment to {payment.party_name}",
            credit=payment.amount,
        )
    
    journal.calculate_totals()
    
    try:
        journal.post(request.user)
        payment.journal_entry = journal
        payment.status = 'confirmed'
        payment.save()
        
        # Audit log with IP address
        from apps.core.audit import audit_payment_post
        audit_payment_post(payment, request.user, request=request)
        
        messages.success(request, f'Payment {payment.payment_number} posted successfully.')
    except Exception as e:
        journal.delete()
        messages.error(request, f'Failed to post payment: {e}')
    
    return redirect('finance:payment_list')


# ============ BANK STATEMENT VIEWS ============

class BankStatementListView(PermissionRequiredMixin, ListView):
    model = BankStatement
    template_name = 'finance/bankstatement_list.html'
    context_object_name = 'statements'
    module_name = 'finance'
    permission_type = 'view'
    paginate_by = 25
    
    def get_queryset(self):
        from .models import BankStatement
        queryset = BankStatement.objects.filter(is_active=True).select_related('bank_account')
        
        bank_id = self.request.GET.get('bank')
        if bank_id:
            queryset = queryset.filter(bank_account_id=bank_id)
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Bank Statements'
        context['bank_accounts'] = BankAccount.objects.filter(is_active=True)
        context['status_choices'] = BankStatement.STATUS_CHOICES
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        return context


class BankStatementCreateView(CreatePermissionMixin, CreateView):
    model = BankStatement
    template_name = 'finance/bankstatement_form.html'
    success_url = reverse_lazy('finance:bankstatement_list')
    module_name = 'finance'
    fields = ['bank_account', 'statement_start_date', 'statement_end_date', 
              'opening_balance', 'closing_balance', 'notes']
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields['bank_account'].widget.attrs['class'] = 'form-select'
        form.fields['bank_account'].queryset = BankAccount.objects.filter(is_active=True)
        for field_name in ['statement_start_date', 'statement_end_date']:
            form.fields[field_name].widget = forms.DateInput(attrs={'type': 'date', 'class': 'form-control'})
        for field_name in ['opening_balance', 'closing_balance']:
            form.fields[field_name].widget.attrs['class'] = 'form-control'
        form.fields['notes'].widget = forms.Textarea(attrs={'rows': 2, 'class': 'form-control'})
        return form
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Bank Statement'
        context['today'] = date.today().isoformat()
        return context
    
    def form_valid(self, form):
        messages.success(self.request, 'Bank Statement created successfully.')
        return super().form_valid(form)


@login_required
def bankstatement_template_download(request):
    """Download Excel template for bank statement import."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Statement Lines"
    
    # Headers
    headers = [
        'Transaction Date (YYYY-MM-DD) *',
        'Description *',
        'Reference',
        'Debit Amount (Money Out)',
        'Credit Amount (Money In)',
        'Balance',
        'Value Date (YYYY-MM-DD)'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = 22
    
    # Add sample data
    sample_data = [
        ['2026-01-05', 'Client Payment - INV-1001', 'TRF1001', '', '10500.00', '110500.00', '2026-01-05'],
        ['2026-01-07', 'Office Rent - January', 'CHQ789456', '6000.00', '', '104500.00', '2026-01-07'],
        ['2026-01-10', 'DEWA Utilities', 'DEWA012', '1200.00', '', '103300.00', '2026-01-10'],
        ['2026-01-12', 'Salary Transfer - January', 'SALJAN', '18000.00', '', '85300.00', '2026-01-12'],
        ['2026-01-15', 'Vendor Payment - Cloud Services', 'TRF2002', '3150.00', '', '82150.00', '2026-01-15'],
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            if col_num in [4, 5, 6]:  # Amount columns
                cell.alignment = Alignment(horizontal='right')
    
    # Add instructions sheet
    ws2 = wb.create_sheet(title="Instructions")
    instructions = [
        ["Bank Statement Import Instructions"],
        [""],
        ["Required Fields (marked with *):"],
        ["- Transaction Date: Date in YYYY-MM-DD format (e.g., 2026-01-15)"],
        ["- Description: Transaction description from bank statement"],
        [""],
        ["Optional Fields:"],
        ["- Reference: Bank reference or cheque number"],
        ["- Debit Amount: Money going out (leave blank if credit)"],
        ["- Credit Amount: Money coming in (leave blank if debit)"],
        ["- Balance: Running balance after transaction"],
        ["- Value Date: Value date if different from transaction date"],
        [""],
        ["Notes:"],
        ["- Either Debit or Credit must have a value, not both"],
        ["- Delete sample rows before importing your actual data"],
        ["- Maximum 1000 rows per import"],
    ]
    
    for row_num, row_data in enumerate(instructions, 1):
        cell = ws2.cell(row=row_num, column=1, value=row_data[0] if row_data else "")
        if row_num == 1:
            cell.font = Font(bold=True, size=14)
        if "Required" in str(row_data) or "Optional" in str(row_data) or "Notes" in str(row_data):
            cell.font = Font(bold=True)
    
    ws2.column_dimensions['A'].width = 70
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=bank_statement_template.xlsx'
    wb.save(response)
    return response


@login_required
def bankstatement_import(request, pk):
    """Import bank statement lines from Excel file."""
    from openpyxl import load_workbook
    from decimal import Decimal, InvalidOperation
    
    statement = get_object_or_404(BankStatement, pk=pk)
    
    if request.method != 'POST':
        return redirect('finance:bankstatement_detail', pk=pk)
    
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, 'Please select an Excel file to import.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    # Validate file extension
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls).')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    try:
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        imported_count = 0
        errors = []
        
        # Get existing max line number
        max_line = statement.lines.aggregate(max_line=models.Max('line_number'))['max_line'] or 0
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not row or not any(row):
                continue
            
            try:
                transaction_date = row[0]
                description = row[1]
                reference = row[2] or ''
                debit = row[3]
                credit = row[4]
                balance = row[5] if len(row) > 5 else None
                value_date = row[6] if len(row) > 6 else None
                
                # Validate required fields
                if not transaction_date:
                    errors.append(f"Row {row_num}: Transaction Date is required")
                    continue
                if not description:
                    errors.append(f"Row {row_num}: Description is required")
                    continue
                
                # Parse date
                if isinstance(transaction_date, str):
                    from datetime import datetime
                    transaction_date = datetime.strptime(transaction_date, '%Y-%m-%d').date()
                elif hasattr(transaction_date, 'date'):
                    transaction_date = transaction_date.date() if hasattr(transaction_date, 'date') else transaction_date
                
                # Parse value date
                if value_date:
                    if isinstance(value_date, str):
                        from datetime import datetime
                        value_date = datetime.strptime(value_date, '%Y-%m-%d').date()
                    elif hasattr(value_date, 'date'):
                        value_date = value_date.date() if hasattr(value_date, 'date') else value_date
                
                # Parse amounts
                debit_amount = Decimal(str(debit or 0).replace(',', ''))
                credit_amount = Decimal(str(credit or 0).replace(',', ''))
                balance_amount = Decimal(str(balance or 0).replace(',', '')) if balance else Decimal('0.00')
                
                if debit_amount == 0 and credit_amount == 0:
                    errors.append(f"Row {row_num}: Either Debit or Credit amount is required")
                    continue
                
                # Create statement line
                max_line += 1
                BankStatementLine.objects.create(
                    statement=statement,
                    line_number=max_line,
                    transaction_date=transaction_date,
                    value_date=value_date,
                    description=str(description)[:500],
                    reference=str(reference)[:200],
                    debit=debit_amount,
                    credit=credit_amount,
                    balance=balance_amount,
                    reconciliation_status='unmatched'
                )
                imported_count += 1
                
            except (ValueError, InvalidOperation) as e:
                errors.append(f"Row {row_num}: Invalid data format - {str(e)}")
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        # Update statement totals
        statement.total_debits = statement.lines.aggregate(total=models.Sum('debit'))['total'] or Decimal('0.00')
        statement.total_credits = statement.lines.aggregate(total=models.Sum('credit'))['total'] or Decimal('0.00')
        statement.save()
        
        if imported_count > 0:
            messages.success(request, f'Successfully imported {imported_count} transaction(s).')
        
        if errors:
            error_msg = f'{len(errors)} error(s) during import. First 5: ' + '; '.join(errors[:5])
            messages.warning(request, error_msg)
        
        if imported_count == 0 and not errors:
            messages.warning(request, 'No data found in the Excel file. Make sure to use the template format.')
            
    except Exception as e:
        messages.error(request, f'Error reading Excel file: {str(e)}')
    
    return redirect('finance:bankstatement_detail', pk=pk)


class BankStatementDetailView(PermissionRequiredMixin, DetailView):
    """Bank Statement detail - Main reconciliation interface."""
    model = BankStatement
    template_name = 'finance/bankstatement_detail.html'
    context_object_name = 'statement'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_context_data(self, **kwargs):
        from .forms import AdjustmentForm
        from apps.core.audit import get_entity_audit_history
        
        context = super().get_context_data(**kwargs)
        context['title'] = f'Bank Statement: {self.object.statement_number}'
        
        # Statement lines
        context['lines'] = self.object.lines.all().select_related(
            'matched_payment', 'matched_journal_line', 'adjustment_journal'
        )
        
        # Unmatched payments for this bank account
        context['unmatched_payments'] = Payment.objects.filter(
            bank_account=self.object.bank_account,
            status='confirmed',
        ).exclude(
            id__in=BankStatementLine.objects.filter(
                matched_payment__isnull=False
            ).values_list('matched_payment_id', flat=True)
        )
        
        # Unmatched journal entries for the bank GL account
        context['unmatched_journals'] = JournalEntryLine.objects.filter(
            account=self.object.bank_account.gl_account,
            journal_entry__status='posted',
        ).exclude(
            id__in=BankStatementLine.objects.filter(
                matched_journal_line__isnull=False
            ).values_list('matched_journal_line_id', flat=True)
        ).select_related('journal_entry')
        
        # Expense/Income accounts for adjustments
        context['expense_accounts'] = Account.objects.filter(
            is_active=True, account_type__in=['expense', 'income']
        ).order_by('account_type', 'code')
        
        # Permissions
        context['can_edit'] = (
            self.request.user.is_superuser or 
            PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        ) and self.object.status not in ['reconciled', 'locked']
        
        context['can_finalize'] = context['can_edit'] and self.object.unmatched_count == 0
        
        # Audit History
        context['audit_history'] = get_entity_audit_history('BankReconciliation', self.object.pk)
        
        return context


@login_required
def bankstatement_import(request, pk):
    """Import bank statement lines from CSV."""
    import csv
    from io import TextIOWrapper
    from .models import BankStatement, BankStatementLine
    
    statement = get_object_or_404(BankStatement, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if statement.status in ['reconciled', 'locked']:
        messages.error(request, 'Cannot import to reconciled/locked statements.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if request.method == 'POST':
        csv_file = request.FILES.get('csv_file')
        if not csv_file:
            messages.error(request, 'Please upload a CSV file.')
            return redirect('finance:bankstatement_detail', pk=pk)
        
        try:
            # Clear existing lines
            statement.lines.all().delete()
            
            # Parse CSV
            decoded_file = TextIOWrapper(csv_file.file, encoding='utf-8-sig')
            reader = csv.DictReader(decoded_file)
            
            line_number = 1
            for row in reader:
                # Expected columns: Date, Description, Reference, Debit, Credit, Balance
                BankStatementLine.objects.create(
                    statement=statement,
                    line_number=line_number,
                    transaction_date=row.get('Date', row.get('date', date.today())),
                    description=row.get('Description', row.get('description', '')),
                    reference=row.get('Reference', row.get('reference', '')),
                    debit=Decimal(row.get('Debit', row.get('debit', '0')) or '0'),
                    credit=Decimal(row.get('Credit', row.get('credit', '0')) or '0'),
                    balance=Decimal(row.get('Balance', row.get('balance', '0')) or '0'),
                )
                line_number += 1
            
            statement.calculate_totals()
            statement.status = 'in_progress'
            statement.save()
            
            messages.success(request, f'{line_number - 1} lines imported successfully.')
        except Exception as e:
            messages.error(request, f'Import failed: {e}')
    
    return redirect('finance:bankstatement_detail', pk=pk)


@login_required
def bankstatement_add_line(request, pk):
    """Manually add a line to bank statement."""
    from .models import BankStatement, BankStatementLine
    
    statement = get_object_or_404(BankStatement, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if statement.status in ['reconciled', 'locked']:
        messages.error(request, 'Cannot add lines to reconciled/locked statements.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            last_line = statement.lines.order_by('-line_number').first()
            next_line_number = (last_line.line_number + 1) if last_line else 1
            
            debit = Decimal(request.POST.get('debit', '0') or '0')
            credit = Decimal(request.POST.get('credit', '0') or '0')
            
            BankStatementLine.objects.create(
                statement=statement,
                line_number=next_line_number,
                transaction_date=request.POST.get('transaction_date'),
                description=request.POST.get('description', ''),
                reference=request.POST.get('reference', ''),
                debit=debit,
                credit=credit,
            )
            
            statement.calculate_totals()
            if statement.status == 'draft':
                statement.status = 'in_progress'
                statement.save()
            
            messages.success(request, 'Line added successfully.')
        except Exception as e:
            messages.error(request, f'Failed to add line: {e}')
    
    return redirect('finance:bankstatement_detail', pk=pk)


@login_required
def bankstatement_auto_match(request, pk):
    """Auto-match statement lines with accounting records."""
    from .models import BankStatement
    
    statement = get_object_or_404(BankStatement, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if statement.status in ['reconciled', 'locked']:
        messages.error(request, 'Cannot modify reconciled/locked statements.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    date_tolerance = int(request.GET.get('tolerance', 3))
    matched_count = statement.auto_match(date_tolerance=date_tolerance)
    
    if matched_count > 0:
        messages.success(request, f'{matched_count} lines matched automatically.')
    else:
        messages.info(request, 'No automatic matches found.')
    
    return redirect('finance:bankstatement_detail', pk=pk)


@login_required
def bankstatement_manual_match(request, pk, line_id):
    """Manually match a statement line with a payment or journal entry."""
    from .models import BankStatement, BankStatementLine
    
    statement = get_object_or_404(BankStatement, pk=pk)
    line = get_object_or_404(BankStatementLine, pk=line_id, statement=statement)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if statement.status in ['reconciled', 'locked']:
        messages.error(request, 'Cannot modify reconciled/locked statements.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if request.method == 'POST':
        match_type = request.POST.get('match_type')
        
        try:
            if match_type == 'payment':
                payment_id = request.POST.get('payment_id')
                payment = get_object_or_404(Payment, pk=payment_id)
                line.match_with_payment(payment, request.user)
                messages.success(request, f'Line matched with payment {payment.payment_number}.')
            
            elif match_type == 'journal':
                journal_line_id = request.POST.get('journal_line_id')
                journal_line = get_object_or_404(JournalEntryLine, pk=journal_line_id)
                line.match_with_journal(journal_line, request.user)
                messages.success(request, f'Line matched with journal {journal_line.journal_entry.entry_number}.')
            
        except Exception as e:
            messages.error(request, f'Matching failed: {e}')
    
    return redirect('finance:bankstatement_detail', pk=pk)


@login_required
def bankstatement_unmatch(request, pk, line_id):
    """Unmatch a statement line."""
    from .models import BankStatement, BankStatementLine
    
    statement = get_object_or_404(BankStatement, pk=pk)
    line = get_object_or_404(BankStatementLine, pk=line_id, statement=statement)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if statement.status in ['reconciled', 'locked']:
        messages.error(request, 'Cannot modify reconciled/locked statements.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    try:
        line.unmatch()
        messages.success(request, 'Line unmatched successfully.')
    except Exception as e:
        messages.error(request, f'Unmatch failed: {e}')
    
    return redirect('finance:bankstatement_detail', pk=pk)


@login_required
def bankstatement_adjustment(request, pk, line_id):
    """Create adjustment entry for unmatched bank item."""
    from .models import BankStatement, BankStatementLine
    
    statement = get_object_or_404(BankStatement, pk=pk)
    line = get_object_or_404(BankStatementLine, pk=line_id, statement=statement)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if statement.status in ['reconciled', 'locked']:
        messages.error(request, 'Cannot modify reconciled/locked statements.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    if request.method == 'POST':
        adjustment_type = request.POST.get('adjustment_type')
        expense_account_id = request.POST.get('expense_account_id')
        
        try:
            expense_account = get_object_or_404(Account, pk=expense_account_id)
            journal = line.create_adjustment(adjustment_type, expense_account, request.user)
            messages.success(request, f'Adjustment journal {journal.entry_number} created.')
        except Exception as e:
            messages.error(request, f'Adjustment failed: {e}')
    
    return redirect('finance:bankstatement_detail', pk=pk)


@login_required
def bankstatement_finalize(request, pk):
    """Finalize bank statement reconciliation."""
    from .models import BankStatement
    
    statement = get_object_or_404(BankStatement, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    try:
        statement.finalize(request.user)
        messages.success(request, f'Statement {statement.statement_number} reconciled successfully.')
    except ValidationError as e:
        for error in e.messages:
            messages.error(request, error)
    except Exception as e:
        messages.error(request, f'Finalization failed: {e}')
    
    return redirect('finance:bankstatement_detail', pk=pk)


@login_required
def bankstatement_lock(request, pk):
    """Lock a reconciled bank statement."""
    from .models import BankStatement
    
    statement = get_object_or_404(BankStatement, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'approve')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankstatement_detail', pk=pk)
    
    try:
        statement.lock(request.user)
        messages.success(request, f'Statement {statement.statement_number} locked.')
    except Exception as e:
        messages.error(request, f'Lock failed: {e}')
    
    return redirect('finance:bankstatement_detail', pk=pk)


# ============ BANK RECONCILIATION VIEWS ============

class BankReconciliationListView(PermissionRequiredMixin, ListView):
    model = BankReconciliation
    template_name = 'finance/bankreconciliation_list.html'
    context_object_name = 'reconciliations'
    module_name = 'finance'
    permission_type = 'view'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = BankReconciliation.objects.filter(is_active=True).select_related('bank_account', 'bank_statement')
        
        bank_id = self.request.GET.get('bank')
        if bank_id:
            queryset = queryset.filter(bank_account_id=bank_id)
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Bank Reconciliations'
        context['bank_accounts'] = BankAccount.objects.filter(is_active=True)
        context['status_choices'] = BankReconciliation.STATUS_CHOICES
        context['can_create'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'create')
        context['can_edit'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        context['can_approve'] = self.request.user.is_superuser or PermissionChecker.has_permission(self.request.user, 'finance', 'approve')
        return context


class BankReconciliationCreateView(CreatePermissionMixin, CreateView):
    model = BankReconciliation
    template_name = 'finance/bankreconciliation_form.html'
    success_url = reverse_lazy('finance:bankreconciliation_list')
    module_name = 'finance'
    fields = ['bank_account', 'bank_statement', 'reconciliation_date', 'period_start', 'period_end',
              'statement_opening_balance', 'statement_closing_balance', 'notes']
    
    def get_form(self, form_class=None):
        from .models import BankStatement
        form = super().get_form(form_class)
        for field_name in ['bank_account', 'bank_statement']:
            form.fields[field_name].widget.attrs['class'] = 'form-select'
        form.fields['bank_account'].queryset = BankAccount.objects.filter(is_active=True)
        form.fields['bank_statement'].queryset = BankStatement.objects.filter(
            is_active=True, status__in=['draft', 'in_progress', 'reconciled']
        )
        form.fields['bank_statement'].required = False
        for field_name in ['reconciliation_date', 'period_start', 'period_end']:
            form.fields[field_name].widget = forms.DateInput(attrs={'type': 'date', 'class': 'form-control'})
        for field_name in ['statement_opening_balance', 'statement_closing_balance']:
            form.fields[field_name].widget.attrs['class'] = 'form-control'
        form.fields['notes'].widget = forms.Textarea(attrs={'rows': 2, 'class': 'form-control'})
        return form
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Bank Reconciliation'
        context['today'] = date.today().isoformat()
        return context
    
    def form_valid(self, form):
        self.object = form.save()
        if self.object.bank_statement:
            self.object.calculate_from_statement()
        else:
            self.object.calculate()
        messages.success(self.request, 'Bank Reconciliation created successfully.')
        return redirect(self.success_url)


class BankReconciliationDetailView(PermissionRequiredMixin, DetailView):
    model = BankReconciliation
    template_name = 'finance/bankreconciliation_detail.html'
    context_object_name = 'reconciliation'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Reconciliation: {self.object.reconciliation_number}'
        context['items'] = self.object.items.all()
        
        # Permissions
        context['can_edit'] = (
            self.request.user.is_superuser or 
            PermissionChecker.has_permission(self.request.user, 'finance', 'edit')
        ) and self.object.status in ['draft', 'in_progress']
        
        context['can_complete'] = context['can_edit'] and self.object.is_reconciled
        context['can_approve'] = (
            self.request.user.is_superuser or 
            PermissionChecker.has_permission(self.request.user, 'finance', 'approve')
        ) and self.object.status == 'completed'
        
        return context


@login_required
def bankreconciliation_complete(request, pk):
    """Complete a bank reconciliation."""
    reconciliation = get_object_or_404(BankReconciliation, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankreconciliation_detail', pk=pk)
    
    try:
        reconciliation.complete(request.user)
        messages.success(request, f'Reconciliation {reconciliation.reconciliation_number} completed.')
    except ValidationError as e:
        for error in e.messages:
            messages.error(request, error)
    except Exception as e:
        messages.error(request, f'Failed to complete: {e}')
    
    return redirect('finance:bankreconciliation_detail', pk=pk)


@login_required
def bankreconciliation_approve(request, pk):
    """Approve a completed bank reconciliation."""
    reconciliation = get_object_or_404(BankReconciliation, pk=pk)
    
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'approve')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:bankreconciliation_detail', pk=pk)
    
    try:
        reconciliation.approve(request.user)
        messages.success(request, f'Reconciliation {reconciliation.reconciliation_number} approved.')
    except ValidationError as e:
        for error in e.messages:
            messages.error(request, error)
    except Exception as e:
        messages.error(request, f'Failed to approve: {e}')
    
    return redirect('finance:bankreconciliation_detail', pk=pk)


# ============ RECONCILIATION REPORTS ============

@login_required
def reconciliation_statement_report(request):
    """Bank Reconciliation Statement Report."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    bank_id = request.GET.get('bank')
    as_of_date = request.GET.get('date', date.today().isoformat())
    
    bank_accounts = BankAccount.objects.filter(is_active=True)
    selected_bank = None
    statement_data = None
    
    if bank_id:
        selected_bank = get_object_or_404(BankAccount, pk=bank_id)
        
        # Get GL balance
        gl_account = selected_bank.gl_account
        gl_lines = JournalEntryLine.objects.filter(
            account=gl_account,
            journal_entry__status='posted',
            journal_entry__date__lte=as_of_date,
        ).aggregate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit')
        )
        
        gl_balance = gl_account.opening_balance + (gl_lines['total_debit'] or Decimal('0.00')) - (gl_lines['total_credit'] or Decimal('0.00'))
        
        # Get latest bank statement
        from .models import BankStatement
        latest_statement = BankStatement.objects.filter(
            bank_account=selected_bank,
            statement_end_date__lte=as_of_date,
            status__in=['reconciled', 'locked']
        ).order_by('-statement_end_date').first()
        
        bank_balance = latest_statement.closing_balance if latest_statement else selected_bank.bank_statement_balance
        
        # Outstanding items (deposits in transit, outstanding checks)
        outstanding_deposits = Payment.objects.filter(
            bank_account=selected_bank,
            payment_type='received',
            status='confirmed',
            payment_date__lte=as_of_date,
        ).exclude(
            id__in=BankStatementLine.objects.filter(
                matched_payment__isnull=False,
                statement__status__in=['reconciled', 'locked']
            ).values_list('matched_payment_id', flat=True)
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        outstanding_checks = Payment.objects.filter(
            bank_account=selected_bank,
            payment_type='made',
            status='confirmed',
            payment_date__lte=as_of_date,
        ).exclude(
            id__in=BankStatementLine.objects.filter(
                matched_payment__isnull=False,
                statement__status__in=['reconciled', 'locked']
            ).values_list('matched_payment_id', flat=True)
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        adjusted_bank_balance = bank_balance + outstanding_deposits - outstanding_checks
        difference = gl_balance - adjusted_bank_balance
        
        statement_data = {
            'gl_balance': gl_balance,
            'bank_balance': bank_balance,
            'outstanding_deposits': outstanding_deposits,
            'outstanding_checks': outstanding_checks,
            'adjusted_bank_balance': adjusted_bank_balance,
            'difference': difference,
            'is_reconciled': abs(difference) < Decimal('0.01'),
        }
    
    return render(request, 'finance/reconciliation_statement_report.html', {
        'title': 'Bank Reconciliation Statement',
        'bank_accounts': bank_accounts,
        'selected_bank': selected_bank,
        'statement_data': statement_data,
        'as_of_date': as_of_date,
    })


@login_required
def unreconciled_transactions_report(request):
    """Unreconciled Transactions Report."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    bank_id = request.GET.get('bank')
    bank_accounts = BankAccount.objects.filter(is_active=True)
    selected_bank = None
    unreconciled_data = None
    
    if bank_id:
        selected_bank = get_object_or_404(BankAccount, pk=bank_id)
        
        # Unreconciled payments
        unreconciled_received = Payment.objects.filter(
            bank_account=selected_bank,
            payment_type='received',
            status='confirmed',
        ).exclude(
            id__in=BankStatementLine.objects.filter(
                matched_payment__isnull=False
            ).values_list('matched_payment_id', flat=True)
        )
        
        unreconciled_made = Payment.objects.filter(
            bank_account=selected_bank,
            payment_type='made',
            status='confirmed',
        ).exclude(
            id__in=BankStatementLine.objects.filter(
                matched_payment__isnull=False
            ).values_list('matched_payment_id', flat=True)
        )
        
        # Unreconciled bank statement lines
        from .models import BankStatement
        unreconciled_statement_lines = BankStatementLine.objects.filter(
            statement__bank_account=selected_bank,
            reconciliation_status='unmatched',
        ).select_related('statement')
        
        unreconciled_data = {
            'received': unreconciled_received,
            'made': unreconciled_made,
            'statement_lines': unreconciled_statement_lines,
            'total_received': unreconciled_received.aggregate(total=Sum('amount'))['total'] or Decimal('0.00'),
            'total_made': unreconciled_made.aggregate(total=Sum('amount'))['total'] or Decimal('0.00'),
        }
    
    return render(request, 'finance/unreconciled_transactions_report.html', {
        'title': 'Unreconciled Transactions',
        'bank_accounts': bank_accounts,
        'selected_bank': selected_bank,
        'unreconciled_data': unreconciled_data,
    })


@login_required
def reconciliation_adjustments_report(request):
    """Reconciliation Adjustments Report."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    start_date = request.GET.get('start_date', date(date.today().year, 1, 1).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    bank_id = request.GET.get('bank')
    
    bank_accounts = BankAccount.objects.filter(is_active=True)
    
    # Get adjustment journal entries (tagged via reference starting with ADJ-)
    adjustments = JournalEntry.objects.filter(
        reference__startswith='ADJ-',
        status='posted',
        date__gte=start_date,
        date__lte=end_date,
    ).order_by('-date')
    
    if bank_id:
        selected_bank = get_object_or_404(BankAccount, pk=bank_id)
        adjustments = adjustments.filter(
            lines__account=selected_bank.gl_account
        ).distinct()
    else:
        selected_bank = None
    
    return render(request, 'finance/reconciliation_adjustments_report.html', {
        'title': 'Reconciliation Adjustments',
        'bank_accounts': bank_accounts,
        'selected_bank': selected_bank,
        'adjustments': adjustments,
        'start_date': start_date,
        'end_date': end_date,
    })


@login_required
def cleared_vs_uncleared_report(request):
    """Cleared vs Uncleared Transactions Report."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    bank_id = request.GET.get('bank')
    bank_accounts = BankAccount.objects.filter(is_active=True)
    selected_bank = None
    report_data = None
    
    if bank_id:
        selected_bank = get_object_or_404(BankAccount, pk=bank_id)
        
        # Cleared (reconciled) payments
        cleared_payments = Payment.objects.filter(
            bank_account=selected_bank,
            status='reconciled',
        )
        
        # Uncleared (confirmed but not reconciled) payments
        uncleared_payments = Payment.objects.filter(
            bank_account=selected_bank,
            status='confirmed',
        )
        
        report_data = {
            'cleared': cleared_payments,
            'uncleared': uncleared_payments,
            'cleared_total': cleared_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00'),
            'uncleared_total': uncleared_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00'),
        }
    
    return render(request, 'finance/cleared_vs_uncleared_report.html', {
        'title': 'Cleared vs Uncleared Transactions',
        'bank_accounts': bank_accounts,
        'selected_bank': selected_bank,
        'report_data': report_data,
    })


@login_required
def bank_vs_gl_report(request):
    """Bank Ledger vs GL Ledger Difference Report."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    as_of_date = request.GET.get('date', date.today().isoformat())
    
    bank_accounts = BankAccount.objects.filter(is_active=True)
    comparison_data = []
    
    for bank in bank_accounts:
        # GL balance
        gl_account = bank.gl_account
        gl_lines = JournalEntryLine.objects.filter(
            account=gl_account,
            journal_entry__status='posted',
            journal_entry__date__lte=as_of_date,
        ).aggregate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit')
        )
        
        gl_balance = gl_account.opening_balance + (gl_lines['total_debit'] or Decimal('0.00')) - (gl_lines['total_credit'] or Decimal('0.00'))
        
        # Latest reconciled statement balance
        from .models import BankStatement
        latest_statement = BankStatement.objects.filter(
            bank_account=bank,
            status__in=['reconciled', 'locked']
        ).order_by('-statement_end_date').first()
        
        bank_balance = latest_statement.closing_balance if latest_statement else bank.bank_statement_balance
        
        difference = gl_balance - bank_balance
        
        comparison_data.append({
            'bank': bank,
            'gl_balance': gl_balance,
            'bank_balance': bank_balance,
            'difference': difference,
            'is_reconciled': abs(difference) < Decimal('0.01'),
            'last_reconciled': latest_statement.statement_end_date if latest_statement else None,
        })
    
    return render(request, 'finance/bank_vs_gl_report.html', {
        'title': 'Bank vs GL Ledger Comparison',
        'comparison_data': comparison_data,
        'as_of_date': as_of_date,
    })


# ============ OPENING BALANCE VIEWS ============

class OpeningBalanceListView(PermissionRequiredMixin, ListView):
    """List of all opening balance entries."""
    model = OpeningBalanceEntry
    template_name = 'finance/openingbalance_list.html'
    context_object_name = 'entries'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        queryset = OpeningBalanceEntry.objects.filter(is_active=True).select_related(
            'fiscal_year', 'journal_entry', 'posted_by'
        )
        
        entry_type = self.request.GET.get('type')
        if entry_type:
            queryset = queryset.filter(entry_type=entry_type)
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        return queryset.order_by('-entry_date', '-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Opening Balances'
        context['entry_types'] = OpeningBalanceEntry.ENTRY_TYPE_CHOICES
        
        # Get system-generated opening balance journal entry
        system_opening_journal = JournalEntry.objects.filter(
            entry_type='opening',
            is_system_generated=True,
            status='posted'
        ).prefetch_related('lines', 'lines__account').first()
        
        context['system_opening_journal'] = system_opening_journal
        
        # Determine if opening balances are locked (fiscal year closed)
        # Default: editable (not locked)
        is_fy_closed = False
        fiscal_year = None
        
        if system_opening_journal:
            fiscal_year = system_opening_journal.fiscal_year
            
            # If fiscal_year is not linked, find it based on journal date
            if not fiscal_year:
                fiscal_year = FiscalYear.objects.filter(
                    start_date__lte=system_opening_journal.date,
                    end_date__gte=system_opening_journal.date,
                    is_active=True
                ).first()
                
                # Link it to the journal for future reference
                if fiscal_year:
                    system_opening_journal.fiscal_year = fiscal_year
                    system_opening_journal.save(update_fields=['fiscal_year'])
            
            # Only locked if fiscal year is explicitly closed
            if fiscal_year:
                is_fy_closed = fiscal_year.is_closed
        
        context['is_fy_closed'] = is_fy_closed
        context['fiscal_year'] = fiscal_year
        
        if system_opening_journal:
            # Get opening balance lines grouped by account type
            lines = system_opening_journal.lines.all().select_related('account')
            
            # Aggregate by account type
            assets = []
            liabilities = []
            equity = []
            total_debit = Decimal('0.00')
            total_credit = Decimal('0.00')
            
            for line in lines:
                account = line.account
                line_data = {
                    'code': account.code,
                    'name': account.name,
                    'debit': line.debit,
                    'credit': line.credit,
                    'type': account.account_type,
                }
                
                total_debit += line.debit
                total_credit += line.credit
                
                if account.account_type == 'asset':
                    assets.append(line_data)
                elif account.account_type == 'liability':
                    liabilities.append(line_data)
                elif account.account_type == 'equity':
                    equity.append(line_data)
            
            context['opening_assets'] = assets
            context['opening_liabilities'] = liabilities
            context['opening_equity'] = equity
            context['opening_total_debit'] = total_debit
            context['opening_total_credit'] = total_credit
            context['is_balanced'] = total_debit == total_credit
        
        return context


class OpeningBalanceCreateView(CreatePermissionMixin, CreateView):
    """Create new opening balance entry."""
    model = OpeningBalanceEntry
    form_class = OpeningBalanceEntryForm
    template_name = 'finance/openingbalance_form.html'
    success_url = reverse_lazy('finance:openingbalance_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Opening Balance Entry'
        context['today'] = date.today().isoformat()
        if self.request.POST:
            context['formset'] = OpeningBalanceLineFormSet(self.request.POST)
        else:
            context['formset'] = OpeningBalanceLineFormSet()
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        
        if formset.is_valid():
            self.object = form.save(commit=False)
            self.object.created_by = self.request.user
            self.object.save()
            
            formset.instance = self.object
            formset.save()
            
            self.object.calculate_totals()
            
            messages.success(self.request, f'Opening Balance Entry {self.object.entry_number} created successfully.')
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))


class OpeningBalanceDetailView(PermissionRequiredMixin, DetailView):
    """View opening balance entry details."""
    model = OpeningBalanceEntry
    template_name = 'finance/openingbalance_detail.html'
    context_object_name = 'entry'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Opening Balance: {self.object.entry_number}'
        context['lines'] = self.object.lines.all().select_related('account', 'customer', 'vendor', 'bank_account')
        return context


class OpeningBalanceUpdateView(UpdatePermissionMixin, UpdateView):
    """Update opening balance entry."""
    model = OpeningBalanceEntry
    form_class = OpeningBalanceEntryForm
    template_name = 'finance/openingbalance_form.html'
    success_url = reverse_lazy('finance:openingbalance_list')
    module_name = 'finance'
    
    def get_queryset(self):
        return OpeningBalanceEntry.objects.filter(status='draft')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Opening Balance: {self.object.entry_number}'
        context['today'] = date.today().isoformat()
        if self.request.POST:
            context['formset'] = OpeningBalanceLineFormSet(self.request.POST, instance=self.object)
        else:
            context['formset'] = OpeningBalanceLineFormSet(instance=self.object)
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        
        if formset.is_valid():
            self.object = form.save(commit=False)
            self.object.updated_by = self.request.user
            self.object.save()
            
            formset.save()
            self.object.calculate_totals()
            
            messages.success(self.request, f'Opening Balance Entry {self.object.entry_number} updated successfully.')
            return redirect('finance:openingbalance_detail', pk=self.object.pk)
        else:
            return self.render_to_response(self.get_context_data(form=form))


@login_required
def openingbalance_post(request, pk):
    """Post an opening balance entry."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:openingbalance_list')
    
    entry = get_object_or_404(OpeningBalanceEntry, pk=pk)
    
    if request.method == 'POST':
        try:
            journal = entry.post(request.user)
            messages.success(request, f'Opening Balance Entry {entry.entry_number} posted successfully. Journal Entry: {journal.entry_number}')
        except ValidationError as e:
            messages.error(request, str(e))
    
    return redirect('finance:openingbalance_detail', pk=pk)


@login_required
def openingbalance_reverse(request, pk):
    """Reverse a posted opening balance entry."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:openingbalance_list')
    
    entry = get_object_or_404(OpeningBalanceEntry, pk=pk)
    
    if request.method == 'POST':
        try:
            entry.reverse(request.user)
            messages.success(request, f'Opening Balance Entry {entry.entry_number} reversed successfully.')
        except ValidationError as e:
            messages.error(request, str(e))
    
    return redirect('finance:openingbalance_detail', pk=pk)


# ============ SYSTEM OPENING BALANCE EDIT ============

@login_required
def system_opening_balance_edit(request):
    """
    Edit system-generated opening balances.
    
    ACCOUNTING RULES:
    - Editable BEFORE fiscal year is closed
    - Restricted AFTER fiscal year is closed
    - If account has other transactions, cannot change the account
    - All changes are logged in audit trail
    
    LOCK LOGIC (CORRECT):
    - ALLOW edit if fiscal_year.is_closed == False
    - DENY edit ONLY if fiscal_year.is_closed == True
    - If fiscal_year is None, determine by journal date
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:openingbalance_list')
    
    # Get the system opening balance journal entry
    journal = JournalEntry.objects.filter(
        entry_type='opening',
        is_system_generated=True
    ).prefetch_related('lines', 'lines__account').first()
    
    if not journal:
        messages.error(request, 'No system opening balance entry found.')
        return redirect('finance:openingbalance_list')
    
    # Determine fiscal year
    fiscal_year = journal.fiscal_year
    
    # If fiscal_year is not linked, find and link it based on journal date
    if not fiscal_year:
        # Find fiscal year that contains the journal date
        fiscal_year = FiscalYear.objects.filter(
            start_date__lte=journal.date,
            end_date__gte=journal.date,
            is_active=True
        ).first()
        
        # If found, link it to the journal
        if fiscal_year:
            journal.fiscal_year = fiscal_year
            journal.save(update_fields=['fiscal_year'])
    
    # Check if fiscal year is closed - ONLY deny if explicitly closed
    if fiscal_year and fiscal_year.is_closed:
        messages.error(request, f'Cannot edit opening balances - Fiscal Year {fiscal_year.name} is closed.')
        return redirect('finance:openingbalance_list')
    
    # If no fiscal year at all (shouldn't happen in production), allow editing
    # This is a fallback - opening balances should always be linked to a fiscal year
    
    # Get lines - ALL lines are editable when fiscal year is OPEN
    # Presence of subsequent transactions does NOT block opening balance edits
    # (This is standard accounting behavior - opening balances are the foundation)
    lines_data = []
    for line in journal.lines.all().select_related('account'):
        # Check if this account has other transactions (for informational purposes only)
        other_transactions = JournalEntryLine.objects.filter(
            account=line.account,
            journal_entry__status='posted'
        ).exclude(
            journal_entry=journal
        ).exists()
        
        lines_data.append({
            'id': line.id,
            'account': line.account,
            'description': line.description,
            'debit': line.debit,
            'credit': line.credit,
            'has_transactions': other_transactions,  # Informational only
            'editable': True,  # ALWAYS editable when FY is open
        })
    
    if request.method == 'POST':
        # Process the form
        try:
            from apps.core.audit import log_finance_audit
            
            changes_made = []
            total_debit = Decimal('0.00')
            total_credit = Decimal('0.00')
            
            for line_data in lines_data:
                line_id = line_data['id']
                line = JournalEntryLine.objects.get(pk=line_id)
                
                # Get new values from form
                new_debit = request.POST.get(f'debit_{line_id}', '0')
                new_credit = request.POST.get(f'credit_{line_id}', '0')
                
                try:
                    new_debit = Decimal(new_debit) if new_debit else Decimal('0.00')
                    new_credit = Decimal(new_credit) if new_credit else Decimal('0.00')
                except:
                    new_debit = Decimal('0.00')
                    new_credit = Decimal('0.00')
                
                # Track changes - opening balances are ALWAYS editable when FY is open
                # Presence of subsequent transactions does NOT block edits
                if line.debit != new_debit or line.credit != new_credit:
                    # Log the change
                    change_record = {
                        'account': line.account.code,
                        'account_name': line.account.name,
                        'old_debit': float(line.debit),
                        'old_credit': float(line.credit),
                        'new_debit': float(new_debit),
                        'new_credit': float(new_credit),
                        'has_subsequent_transactions': line_data['has_transactions'],
                    }
                    changes_made.append(change_record)
                    
                    # Update the line
                    line.debit = new_debit
                    line.credit = new_credit
                    line.save()
                
                total_debit += new_debit
                total_credit += new_credit
            
            # Validate balance
            if total_debit != total_credit:
                messages.error(request, f'Opening balance is not balanced. Debit: {total_debit}, Credit: {total_credit}')
                return redirect('finance:system_opening_balance_edit')
            
            # Update journal totals
            journal.total_debit = total_debit
            journal.total_credit = total_credit
            journal.save(update_fields=['total_debit', 'total_credit'])
            
            # RECALCULATE ACCOUNT BALANCES
            # Opening balance changes affect all subsequent balances
            if changes_made:
                recalculated_accounts = []
                for change in changes_made:
                    account_code = change['account']
                    try:
                        account = Account.objects.get(code=account_code)
                        # Recalculate the account's current balance
                        # Balance = Opening Balance (from this journal) + All subsequent transactions
                        account_lines = JournalEntryLine.objects.filter(
                            account=account,
                            journal_entry__status='posted'
                        ).aggregate(
                            total_debit=Coalesce(Sum('debit'), Decimal('0.00')),
                            total_credit=Coalesce(Sum('credit'), Decimal('0.00'))
                        )
                        
                        # Calculate new balance based on account type
                        if account.debit_increases:
                            new_balance = account_lines['total_debit'] - account_lines['total_credit']
                        else:
                            new_balance = account_lines['total_credit'] - account_lines['total_debit']
                        
                        # Update account balance
                        account.balance = new_balance
                        account.save(update_fields=['balance'])
                        recalculated_accounts.append(account_code)
                    except Account.DoesNotExist:
                        pass
                
                # Log audit with recalculation info
                log_finance_audit(
                    user=request.user,
                    action='update',
                    entity_type='OpeningBalance',
                    entity_id=journal.pk,
                    request=request,
                    details={
                        'journal_number': journal.entry_number,
                        'changes': changes_made,
                        'total_debit': float(total_debit),
                        'total_credit': float(total_credit),
                        'recalculated_accounts': recalculated_accounts,
                        'note': 'Downstream balances recalculated automatically',
                    }
                )
                messages.success(
                    request, 
                    f'Opening balances updated successfully. {len(changes_made)} line(s) changed. '
                    f'Account balances recalculated.'
                )
            else:
                messages.info(request, 'No changes were made.')
            
            return redirect('finance:openingbalance_list')
            
        except Exception as e:
            messages.error(request, f'Error updating opening balances: {str(e)}')
            return redirect('finance:system_opening_balance_edit')
    
    # Group lines by account type for display
    assets = [l for l in lines_data if l['account'].account_type == 'asset']
    liabilities = [l for l in lines_data if l['account'].account_type == 'liability']
    equity = [l for l in lines_data if l['account'].account_type == 'equity']
    
    # Calculate totals
    total_debit = sum(l['debit'] for l in lines_data)
    total_credit = sum(l['credit'] for l in lines_data)
    
    # Get all accounts for adding new lines
    available_accounts = Account.objects.filter(
        is_active=True,
        account_type__in=['asset', 'liability', 'equity']
    ).exclude(
        id__in=[l['account'].id for l in lines_data]
    ).order_by('account_type', 'code')
    
    context = {
        'title': 'Edit System Opening Balances',
        'journal': journal,
        'assets': assets,
        'liabilities': liabilities,
        'equity': equity,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'is_balanced': total_debit == total_credit,
        'available_accounts': available_accounts,
        'fiscal_year_closed': fiscal_year.is_closed if fiscal_year else False,
    }
    
    return render(request, 'finance/system_opening_balance_edit.html', context)


@login_required
def system_opening_balance_add_line(request):
    """Add a new line to system opening balance."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})
    
    journal = JournalEntry.objects.filter(
        entry_type='opening',
        is_system_generated=True
    ).first()
    
    if not journal:
        return JsonResponse({'success': False, 'error': 'No opening balance journal found'})
    
    # Check fiscal year
    if journal.fiscal_year and journal.fiscal_year.is_closed:
        return JsonResponse({'success': False, 'error': 'Fiscal year is closed'})
    
    account_id = request.POST.get('account_id')
    debit = request.POST.get('debit', '0')
    credit = request.POST.get('credit', '0')
    
    try:
        account = Account.objects.get(pk=account_id, is_active=True)
        debit = Decimal(debit) if debit else Decimal('0.00')
        credit = Decimal(credit) if credit else Decimal('0.00')
        
        # Check if account already exists in opening balance
        if journal.lines.filter(account=account).exists():
            return JsonResponse({'success': False, 'error': 'Account already exists in opening balances'})
        
        # Create new line
        line = JournalEntryLine.objects.create(
            journal_entry=journal,
            account=account,
            description=f'Opening Balance for {account.name}',
            debit=debit,
            credit=credit,
        )
        
        # Update journal totals
        journal.calculate_totals()
        
        # Log audit
        from apps.core.audit import log_finance_audit
        log_finance_audit(
            user=request.user,
            action='create',
            entity_type='OpeningBalanceLine',
            entity_id=line.pk,
            request=request,
            details={
                'journal_number': journal.entry_number,
                'account': account.code,
                'account_name': account.name,
                'debit': float(debit),
                'credit': float(credit),
            }
        )
        
        return JsonResponse({
            'success': True,
            'line_id': line.pk,
            'message': f'Opening balance for {account.code} added successfully'
        })
        
    except Account.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Account not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def system_opening_balance_delete_line(request, line_id):
    """Delete a line from system opening balance.
    
    ACCOUNTING RULES:
    - Allow delete when fiscal year is OPEN
    - Block delete ONLY when fiscal year is CLOSED
    - Recalculate account balances after deletion
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})
    
    try:
        line = JournalEntryLine.objects.select_related('journal_entry', 'account').get(pk=line_id)
        journal = line.journal_entry
        account = line.account
        
        # Check if it's the opening balance journal
        if journal.entry_type != 'opening' or not journal.is_system_generated:
            return JsonResponse({'success': False, 'error': 'Not an opening balance entry'})
        
        # Check fiscal year - ONLY block if closed
        if journal.fiscal_year and journal.fiscal_year.is_closed:
            return JsonResponse({'success': False, 'error': 'Fiscal year is closed'})
        
        # Check if account has other transactions (informational for audit)
        has_transactions = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__status='posted'
        ).exclude(journal_entry=journal).exists()
        
        # Log before deletion
        from apps.core.audit import log_finance_audit
        log_finance_audit(
            user=request.user,
            action='delete',
            entity_type='OpeningBalanceLine',
            entity_id=line.pk,
            request=request,
            details={
                'journal_number': journal.entry_number,
                'account': account.code,
                'account_name': account.name,
                'debit': float(line.debit),
                'credit': float(line.credit),
                'had_transactions': has_transactions,
                'note': 'Opening balance line deleted - balances recalculated' if has_transactions else 'Opening balance line deleted',
            }
        )
        
        # Delete the line
        line.delete()
        
        # Recalculate account balance
        account_lines = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__status='posted'
        ).aggregate(
            total_debit=Coalesce(Sum('debit'), Decimal('0.00')),
            total_credit=Coalesce(Sum('credit'), Decimal('0.00'))
        )
        
        if account.debit_increases:
            new_balance = account_lines['total_debit'] - account_lines['total_credit']
        else:
            new_balance = account_lines['total_credit'] - account_lines['total_debit']
        
        account.balance = new_balance
        account.save(update_fields=['balance'])
        
        # Update journal totals
        journal.calculate_totals()
        
        return JsonResponse({'success': True, 'message': 'Line deleted successfully'})
        
    except JournalEntryLine.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Line not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def seed_fy2025_opening_balance(request):
    """
    Web-accessible endpoint to seed FY 2025 opening balance.
    Superuser only. Runs the same logic as the management command.
    GET shows confirmation page, POST executes.
    """
    if not request.user.is_superuser:
        messages.error(request, 'Only superusers can seed opening balances.')
        return redirect('finance:openingbalance_list')
    
    if request.method == 'POST':
        from django.db import transaction as db_transaction
        
        try:
            with db_transaction.atomic():
                # Step 1: Get or create FY 2025
                fy, fy_created = FiscalYear.objects.get_or_create(
                    start_date=date(2025, 1, 1),
                    defaults={
                        'name': 'FY 2025',
                        'end_date': date(2025, 12, 31),
                        'is_active': True,
                        'is_closed': False,
                    }
                )
                if not fy_created:
                    # Also try by name
                    fy2 = FiscalYear.objects.filter(name__icontains='2025', is_active=True).first()
                    if fy2:
                        fy = fy2
                
                # Step 2: Check for existing entry
                existing = OpeningBalanceEntry.objects.filter(
                    fiscal_year=fy, entry_type='gl'
                ).first()
                if existing:
                    messages.warning(request, f'Opening balance entry already exists: {existing.entry_number}')
                    return redirect('finance:openingbalance_detail', pk=existing.pk)
                
                # Step 3: Account definitions
                account_defs = [
                    {'code': '1101', 'name': 'Cash on Hand', 'type': AccountType.ASSET,
                     'is_cash': True, 'is_fd': False},
                    {'code': '1102', 'name': 'Bank - ADCB Current Account', 'type': AccountType.ASSET,
                     'is_cash': True, 'is_fd': False},
                    {'code': '1103', 'name': 'Bank - ADCB Fixed Deposit', 'type': AccountType.ASSET,
                     'is_cash': False, 'is_fd': True},
                    {'code': '3201', 'name': 'Retained Earnings', 'type': AccountType.EQUITY,
                     'is_cash': False, 'is_fd': False},
                ]
                
                accounts = {}
                for ad in account_defs:
                    acc = Account.objects.filter(code=ad['code'], is_active=True).first()
                    if acc:
                        # Update name if different
                        if acc.name != ad['name']:
                            acc.name = ad['name']
                            acc.account_type = ad['type']
                            acc.is_cash_account = ad['is_cash']
                            acc.is_fixed_deposit = ad['is_fd']
                            acc.save()
                    else:
                        acc = Account.objects.create(
                            code=ad['code'], name=ad['name'],
                            account_type=ad['type'],
                            is_cash_account=ad['is_cash'],
                            is_fixed_deposit=ad['is_fd'],
                            is_system=True,
                        )
                    accounts[ad['code']] = acc
                
                # Step 4: Bank accounts
                bank_defs = [
                    {'name': 'ADCB Bank - Current Account', 'acct_num': 'ADCB-CURR-001',
                     'bank': 'Abu Dhabi Commercial Bank', 'gl': '1102'},
                    {'name': 'ADCB Bank - Fixed Deposit', 'acct_num': 'ADCB-FD-001',
                     'bank': 'Abu Dhabi Commercial Bank', 'gl': '1103'},
                ]
                bank_accounts = {}
                for bd in bank_defs:
                    ba = BankAccount.objects.filter(name=bd['name'], is_active=True).first()
                    if not ba:
                        ba = BankAccount.objects.create(
                            name=bd['name'], account_number=bd['acct_num'],
                            bank_name=bd['bank'], gl_account=accounts[bd['gl']],
                            currency='AED',
                        )
                    bank_accounts[bd['name']] = ba
                
                # Step 5: Create the opening balance entry
                entry = OpeningBalanceEntry(
                    entry_type='gl', fiscal_year=fy,
                    entry_date=date(2025, 1, 1),
                    description='Opening Balance Entry - Beginning of Fiscal Year 2025',
                    notes=(
                        'FY 2025 Opening Balances:\n'
                        '- Cash on Hand: 37,000 AED (Main Safe: 6,000 + Petty Cash: 4,000 + General Cash: 27,000)\n'
                        '- ADCB Current Account: 50,000 AED\n'
                        '- ADCB Fixed Deposit: 48,000 AED\n'
                        '- Retained Earnings: 135,000 AED (Credit - accumulated profits from FY 2024)\n'
                        '\nTotal Cash & Bank: 135,000 AED\n'
                        'Entry Date: 01/01/2025 | Reference Date: 31/12/2024'
                    ),
                )
                entry.save()
                
                # Step 6: Create lines
                lines_data = [
                    {'code': '1101', 'bank': None, 'ref': 'OB-2025-001',
                     'debit': Decimal('37000.00'), 'credit': Decimal('0.00'),
                     'desc': 'Opening Balance - Cash on Hand (Main Safe: 6,000 + Petty Cash: 4,000 + General Cash: 27,000)'},
                    {'code': '1102', 'bank': 'ADCB Bank - Current Account', 'ref': 'STMT-DEC-2024',
                     'debit': Decimal('50000.00'), 'credit': Decimal('0.00'),
                     'desc': 'Opening Balance - ADCB Current Account'},
                    {'code': '1103', 'bank': 'ADCB Bank - Fixed Deposit', 'ref': 'FD-CERT-2024',
                     'debit': Decimal('48000.00'), 'credit': Decimal('0.00'),
                     'desc': 'Opening Balance - ADCB Fixed Deposit'},
                    {'code': '3201', 'bank': None, 'ref': 'YE-2024',
                     'debit': Decimal('0.00'), 'credit': Decimal('135000.00'),
                     'desc': 'Opening Balance - Retained Earnings (Accumulated profits from FY 2024)'},
                ]
                
                for ld in lines_data:
                    OpeningBalanceLine.objects.create(
                        opening_balance_entry=entry,
                        account=accounts[ld['code']],
                        description=ld['desc'],
                        bank_account=bank_accounts.get(ld['bank']),
                        debit=ld['debit'], credit=ld['credit'],
                        reference_number=ld['ref'],
                        reference_date=date(2024, 12, 31),
                    )
                
                entry.calculate_totals()
                
                messages.success(
                    request,
                    f'✅ FY 2025 Opening Balance created successfully! '
                    f'Entry: {entry.entry_number} | '
                    f'Dr: AED {entry.total_debit:,.2f} = Cr: AED {entry.total_credit:,.2f}'
                )
                return redirect('finance:openingbalance_detail', pk=entry.pk)
        
        except Exception as e:
            messages.error(request, f'Error creating opening balance: {str(e)}')
            return redirect('finance:openingbalance_list')
    
    # GET - show confirmation
    existing = None
    fy = FiscalYear.objects.filter(
        Q(start_date__year=2025) | Q(name__icontains='2025'),
        is_active=True
    ).first()
    if fy:
        existing = OpeningBalanceEntry.objects.filter(
            fiscal_year=fy, entry_type='gl'
        ).first()
    
    context = {
        'title': 'Seed FY 2025 Opening Balance',
        'existing': existing,
        'lines': [
            {'account': '1101 - Cash on Hand', 'debit': '37,000.00', 'credit': '0.00', 'ref': 'OB-2025-001'},
            {'account': '1102 - Bank - ADCB Current Account', 'debit': '50,000.00', 'credit': '0.00', 'ref': 'STMT-DEC-2024'},
            {'account': '1103 - Bank - ADCB Fixed Deposit', 'debit': '48,000.00', 'credit': '0.00', 'ref': 'FD-CERT-2024'},
            {'account': '3201 - Retained Earnings', 'debit': '0.00', 'credit': '135,000.00', 'ref': 'YE-2024'},
        ],
    }
    return render(request, 'finance/seed_fy2025_opening_balance.html', context)


# ============ WRITE-OFF VIEWS ============

class WriteOffListView(PermissionRequiredMixin, ListView):
    """List of all write-off entries."""
    model = WriteOff
    template_name = 'finance/writeoff_list.html'
    context_object_name = 'writeoffs'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        queryset = WriteOff.objects.filter(is_active=True).select_related(
            'source_account', 'expense_account', 'customer', 'vendor', 'journal_entry'
        )
        
        writeoff_type = self.request.GET.get('type')
        if writeoff_type:
            queryset = queryset.filter(writeoff_type=writeoff_type)
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        return queryset.order_by('-writeoff_date', '-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Write-Offs & Adjustments'
        context['writeoff_types'] = WriteOff.WRITEOFF_TYPE_CHOICES
        return context


class WriteOffCreateView(CreatePermissionMixin, CreateView):
    """Create new write-off entry."""
    model = WriteOff
    form_class = WriteOffForm
    template_name = 'finance/writeoff_form.html'
    success_url = reverse_lazy('finance:writeoff_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create Write-Off / Adjustment'
        context['today'] = date.today().isoformat()
        return context
    
    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.created_by = self.request.user
        self.object.save()
        
        messages.success(self.request, f'Write-Off {self.object.writeoff_number} created successfully.')
        return redirect('finance:writeoff_detail', pk=self.object.pk)


class WriteOffDetailView(PermissionRequiredMixin, DetailView):
    """View write-off entry details."""
    model = WriteOff
    template_name = 'finance/writeoff_detail.html'
    context_object_name = 'writeoff'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Write-Off: {self.object.writeoff_number}'
        return context


class WriteOffUpdateView(UpdatePermissionMixin, UpdateView):
    """Update write-off entry."""
    model = WriteOff
    form_class = WriteOffForm
    template_name = 'finance/writeoff_form.html'
    module_name = 'finance'
    
    def get_queryset(self):
        return WriteOff.objects.filter(status='draft')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Write-Off: {self.object.writeoff_number}'
        context['today'] = date.today().isoformat()
        return context
    
    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.updated_by = self.request.user
        self.object.save()
        
        messages.success(self.request, f'Write-Off {self.object.writeoff_number} updated successfully.')
        return redirect('finance:writeoff_detail', pk=self.object.pk)


@login_required
def writeoff_approve(request, pk):
    """Approve a write-off entry."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'approve')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:writeoff_list')
    
    writeoff = get_object_or_404(WriteOff, pk=pk)
    
    if request.method == 'POST':
        try:
            writeoff.approve(request.user)
            messages.success(request, f'Write-Off {writeoff.writeoff_number} approved successfully.')
        except ValidationError as e:
            messages.error(request, str(e))
    
    return redirect('finance:writeoff_detail', pk=pk)


@login_required
def writeoff_post(request, pk):
    """Post a write-off entry."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:writeoff_list')
    
    writeoff = get_object_or_404(WriteOff, pk=pk)
    
    if request.method == 'POST':
        try:
            journal = writeoff.post(request.user)
            messages.success(request, f'Write-Off {writeoff.writeoff_number} posted successfully. Journal Entry: {journal.entry_number}')
        except ValidationError as e:
            messages.error(request, str(e))
    
    return redirect('finance:writeoff_detail', pk=pk)


@login_required
def writeoff_reverse(request, pk):
    """Reverse a posted write-off entry."""
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('finance:writeoff_list')
    
    writeoff = get_object_or_404(WriteOff, pk=pk)
    
    if request.method == 'POST':
        try:
            writeoff.reverse(request.user)
            messages.success(request, f'Write-Off {writeoff.writeoff_number} reversed successfully.')
        except ValidationError as e:
            messages.error(request, str(e))
    
    return redirect('finance:writeoff_detail', pk=pk)


# ============ EXCHANGE RATE VIEWS ============

class ExchangeRateListView(PermissionRequiredMixin, ListView):
    """List of all exchange rates with inline creation form."""
    model = ExchangeRate
    template_name = 'finance/exchangerate_list.html'
    context_object_name = 'rates'
    module_name = 'finance'
    permission_type = 'view'
    
    def get_queryset(self):
        queryset = ExchangeRate.objects.filter(is_active=True)
        
        currency = self.request.GET.get('currency')
        if currency:
            queryset = queryset.filter(currency_code=currency.upper())
        
        return queryset.order_by('-rate_date', 'currency_code')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Exchange Rates'
        # Get unique currencies
        context['currencies'] = ExchangeRate.objects.filter(is_active=True).values_list(
            'currency_code', flat=True
        ).distinct().order_by('currency_code')
        # Add form for inline creation
        context['form'] = ExchangeRateForm()
        context['today'] = date.today().isoformat()
        return context
    
    def post(self, request, *args, **kwargs):
        """Handle inline form submission."""
        if not PermissionChecker.has_permission(request.user, 'finance', 'create'):
            messages.error(request, 'Permission denied.')
            return redirect('finance:exchangerate_list')
        
        form = ExchangeRateForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            messages.success(request, f'Exchange Rate for {obj.currency_code} added successfully.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
        
        return redirect('finance:exchangerate_list')


class ExchangeRateCreateView(CreatePermissionMixin, CreateView):
    """Create new exchange rate - kept for backwards compatibility."""
    model = ExchangeRate
    form_class = ExchangeRateForm
    template_name = 'finance/exchangerate_form.html'
    success_url = reverse_lazy('finance:exchangerate_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add Exchange Rate'
        context['today'] = date.today().isoformat()
        return context
    
    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.created_by = self.request.user
        self.object.save()
        
        messages.success(self.request, f'Exchange Rate for {self.object.currency_code} added successfully.')
        return redirect(self.success_url)


class ExchangeRateUpdateView(UpdatePermissionMixin, UpdateView):
    """Update exchange rate."""
    model = ExchangeRate
    form_class = ExchangeRateForm
    template_name = 'finance/exchangerate_form.html'
    success_url = reverse_lazy('finance:exchangerate_list')
    module_name = 'finance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit Exchange Rate: {self.object.currency_code}'
        return context


# ============ VAT AUDIT REPORT ============

@login_required
def vat_audit_report(request):
    """
    VAT Audit Report - Line-level VAT details for FTA audit.
    Shows every transaction mapped to VAT box.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    # Date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date or not end_date:
        # Default to current month
        today = date.today()
        start_date = today.replace(day=1).isoformat()
        end_date = today.isoformat()
    
    # Get all VAT-related journal lines
    vat_tax_codes = TaxCode.objects.filter(is_active=True)
    vat_accounts = []
    for tc in vat_tax_codes:
        if tc.sales_account:
            vat_accounts.append(tc.sales_account.id)
        if tc.purchase_account:
            vat_accounts.append(tc.purchase_account.id)
    
    vat_accounts = list(set(vat_accounts))
    
    # Get all journal entries with VAT impact
    journal_lines = JournalEntryLine.objects.filter(
        journal_entry__status='posted',
        journal_entry__date__gte=start_date,
        journal_entry__date__lte=end_date,
    ).select_related('journal_entry', 'account').order_by('journal_entry__date', 'journal_entry__entry_number')
    
    # Group by VAT box
    vat_data = {
        'box1a': [],  # Standard rated supplies - Emirates
        'box1b': [],  # Standard rated supplies - GCC
        'box2': [],   # Tax refunds
        'box3': [],   # Zero-rated supplies
        'box4': [],   # Exempt supplies
        'box5': [],   # Total value of outputs
        'box6': [],   # Standard rated expenses
        'box7': [],   # Supplies subject to reverse charge
        'box8': [],   # Total value of inputs
        'box9': [],   # Output VAT due
        'box10': [],  # Input VAT recoverable
    }
    
    all_transactions = []
    
    for line in journal_lines:
        # Determine VAT box based on account
        vat_box = 'N/A'
        if line.account.account_type == 'income':
            # Output - Sales
            vat_box = 'Box 1a - Standard Supplies'
            vat_data['box1a'].append(line)
        elif line.account.account_type == 'expense':
            # Input - Expenses
            vat_box = 'Box 6 - Standard Expenses'
            vat_data['box6'].append(line)
        elif line.account.id in vat_accounts:
            # VAT accounts
            if line.credit > 0:
                vat_box = 'Box 9 - Output VAT'
                vat_data['box9'].append(line)
            else:
                vat_box = 'Box 10 - Input VAT'
                vat_data['box10'].append(line)
        
        all_transactions.append({
            'date': line.journal_entry.date,
            'entry_number': line.journal_entry.entry_number,
            'reference': line.journal_entry.reference,
            'description': line.description or line.journal_entry.description,
            'account': line.account,
            'debit': line.debit,
            'credit': line.credit,
            'vat_box': vat_box,
        })
    
    # Calculate totals by box
    box_totals = {}
    for box_name, lines in vat_data.items():
        total_debit = sum(l.debit for l in lines)
        total_credit = sum(l.credit for l in lines)
        box_totals[box_name] = {
            'count': len(lines),
            'debit': total_debit,
            'credit': total_credit,
            'net': total_debit - total_credit,
        }
    
    # Excel Export
    export_format = request.GET.get('format', '')
    if export_format == 'excel':
        from .excel_exports import export_vat_audit
        return export_vat_audit(start_date, end_date, all_transactions, box_totals)
    
    return render(request, 'finance/vat_audit_report.html', {
        'title': 'VAT Audit Report',
        'start_date': start_date,
        'end_date': end_date,
        'transactions': all_transactions,
        'box_totals': box_totals,
        'total_transactions': len(all_transactions),
    })


# ============ ACCOUNT MAPPING VIEWS ============
# SAP/Oracle-style Account Determination / Posting Profiles

@login_required
def account_mapping_list(request):
    """
    Account Mapping / Account Determination - Central configuration.
    One-time setup for all transaction types.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'view')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    can_edit = request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')
    
    # Group mappings by module
    modules = AccountMapping.MODULE_CHOICES
    mappings_by_module = {}
    
    for module_code, module_name in modules:
        mappings = AccountMapping.objects.filter(module=module_code).select_related('account')
        
        # Get all transaction types for this module
        module_types = [
            (code, label) for code, label in AccountMapping.TRANSACTION_TYPE_CHOICES
            if code.startswith(module_code) or 
               (module_code == 'general' and code in ['fx_gain', 'fx_loss', 'retained_earnings', 'opening_balance_equity', 'suspense', 'rounding']) or
               (module_code == 'banking' and code.startswith('bank_'))
        ]
        
        configured_types = {m.transaction_type: m for m in mappings}
        
        module_data = []
        for type_code, type_label in module_types:
            mapping = configured_types.get(type_code)
            module_data.append({
                'transaction_type': type_code,
                'label': type_label,
                'mapping': mapping,
                'account': mapping.account if mapping else None,
            })
        
        if module_data:
            mappings_by_module[module_code] = {
                'name': module_name,
                'items': module_data,
                'is_configured': AccountMapping.is_fully_configured(module_code),
            }
    
    # Get all active accounts for the dropdown
    accounts = Account.objects.filter(is_active=True).order_by('code')
    
    return render(request, 'finance/account_mapping_list.html', {
        'title': 'Account Mapping',
        'mappings_by_module': mappings_by_module,
        'accounts': accounts,
        'can_edit': can_edit,
    })


@login_required
def account_mapping_save(request):
    """
    Save account mapping via AJAX or form POST.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        return JsonResponse({'success': False, 'error': 'Permission denied.'})
    
    if request.method == 'POST':
        transaction_type = request.POST.get('transaction_type')
        account_id = request.POST.get('account_id')
        
        if not transaction_type:
            return JsonResponse({'success': False, 'error': 'Transaction type required.'})
        
        # Determine module from transaction type
        module = 'general'
        for mod_code, _ in AccountMapping.MODULE_CHOICES:
            if transaction_type.startswith(mod_code):
                module = mod_code
                break
        
        # Handle special cases
        if transaction_type.startswith('bank_'):
            module = 'banking'
        elif transaction_type in ['fx_gain', 'fx_loss', 'retained_earnings', 'opening_balance_equity', 'suspense', 'rounding']:
            module = 'general'
        
        if account_id:
            try:
                account = Account.objects.get(pk=account_id, is_active=True)
                mapping, created = AccountMapping.objects.update_or_create(
                    transaction_type=transaction_type,
                    defaults={
                        'module': module,
                        'account': account,
                    }
                )
                return JsonResponse({
                    'success': True, 
                    'message': f'Mapping saved: {account.code} - {account.name}',
                    'account_code': account.code,
                    'account_name': account.name,
                })
            except Account.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Account not found.'})
        else:
            # Remove mapping if account_id is empty
            AccountMapping.objects.filter(transaction_type=transaction_type).delete()
            return JsonResponse({'success': True, 'message': 'Mapping removed.'})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method.'})


@login_required
def accounting_settings(request):
    """
    View and edit global accounting settings.
    Controls auto-posting behavior per module.
    """
    if not (request.user.is_superuser or PermissionChecker.has_permission(request.user, 'finance', 'edit')):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard')
    
    settings_obj = AccountingSettings.get_settings()
    
    if request.method == 'POST':
        # Update settings
        settings_obj.auto_post_sales_invoice = request.POST.get('auto_post_sales_invoice') == 'on'
        settings_obj.auto_post_vendor_bill = request.POST.get('auto_post_vendor_bill') == 'on'
        settings_obj.auto_post_expense_claim = request.POST.get('auto_post_expense_claim') == 'on'
        settings_obj.auto_post_payroll = request.POST.get('auto_post_payroll') == 'on'
        settings_obj.auto_post_payment = request.POST.get('auto_post_payment') == 'on'
        settings_obj.auto_post_bank_transfer = request.POST.get('auto_post_bank_transfer') == 'on'
        settings_obj.require_approval_before_posting = request.POST.get('require_approval_before_posting') == 'on'
        settings_obj.allow_posting_to_closed_period = request.POST.get('allow_posting_to_closed_period') == 'on'
        settings_obj.round_to_fils = request.POST.get('round_to_fils') == 'on'
        
        # VAT rate
        try:
            vat_rate = Decimal(request.POST.get('default_vat_rate', '5.00'))
            settings_obj.default_vat_rate = vat_rate
        except:
            pass
        
        settings_obj.vat_registration_number = request.POST.get('vat_registration_number', '')
        
        settings_obj.save()
        messages.success(request, 'Accounting settings updated successfully.')
        return redirect('finance:accounting_settings')
    
    return render(request, 'finance/accounting_settings.html', {
        'title': 'Accounting Settings',
        'settings': settings_obj,
    })
