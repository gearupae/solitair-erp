import base64
import json
import mimetypes
import re
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django import forms
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.db.models import F, Q
from django.http import FileResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.views import View
from django.views.generic import CreateView, ListView, TemplateView

from apps.core.mixins import PermissionRequiredMixin
from apps.core.utils import PermissionChecker
from apps.settings_app.models import CompanySettings

from .models import (
    StockTakeLine,
    StockTakeScanLog,
    StockTakeSession,
    StockTakeUnknownScan,
)


def _can_inventory_view(user):
    return user.is_superuser or PermissionChecker.has_permission(user, 'inventory', 'view')


def _can_inventory_edit(user):
    return user.is_superuser or PermissionChecker.has_permission(user, 'inventory', 'edit')


def _normalize_barcode_scan_key(s: str) -> str:
    """
    Align scanner output with Excel numeric cells.

    Excel often stores GTIN/EAN as a number; openpyxl / str() yields '4104220122019.0'
    while a scan is '4104220122019'. Alphanumeric codes (e.g. ABC-abc-1234) are unchanged.
    """
    t = (s or '').strip()
    if not t:
        return ''
    if re.fullmatch(r'-?\d+\.0', t):
        return t[:-2]
    return t


def _parse_expected_workbook(upload_file):
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(upload_file.read()), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError('The spreadsheet is empty.')

    header = [str(c).strip().lower() if c is not None else '' for c in rows[0]]

    def find_col(*names):
        for i, h in enumerate(header):
            hn = h.replace(' ', '')
            for n in names:
                if hn == n.replace(' ', '').lower():
                    return i
        return None

    ic_sku = find_col('sku')
    ic_name = find_col('itemname', 'item name')
    ic_exp = find_col('expectedquantity', 'expected qty', 'expected quantity')
    ic_code = find_col(
        'barcode',
        'barcodeorqrcode',
        'barcodeorscancode',
        'barcode/scan code',
        'scancode',
        'scan code',
        'labelcode',
        'label code',
        'itembarcode',
        'item barcode',
        'qrcode',
        'qr code',
        'gtin',
        'upc',
        'ean',
    )
    if ic_sku is None or ic_exp is None:
        raise ValueError(
            'Required columns not found. First row must include: SKU, Expected Quantity '
            '(and optionally Item Name, Barcode or scan code).'
        )

    merged = {}
    for row in rows[1:]:
        if not row or ic_sku >= len(row) or row[ic_sku] is None:
            continue
        sku = _normalize_barcode_scan_key(str(row[ic_sku]).strip())
        if not sku:
            continue
        name = ''
        if ic_name is not None and ic_name < len(row) and row[ic_name] is not None:
            name = str(row[ic_name]).strip()
        scan_code = ''
        if ic_code is not None and ic_code < len(row) and row[ic_code] is not None:
            scan_code = _normalize_barcode_scan_key(str(row[ic_code]).strip())
        raw_exp = row[ic_exp] if ic_exp < len(row) else 0
        try:
            exp = Decimal(str(raw_exp))
        except (InvalidOperation, TypeError, ValueError):
            exp = Decimal('0')
        key = sku.upper()
        if key not in merged:
            merged[key] = {
                'sku': sku,
                'item_name': name or sku,
                'expected_qty': exp,
                'scan_code': scan_code,
            }
        else:
            merged[key]['expected_qty'] += exp
            if name:
                merged[key]['item_name'] = name
            if scan_code:
                merged[key]['scan_code'] = scan_code
    if not merged:
        raise ValueError('No data rows found below the header.')
    return list(merged.values())


class SessionForm(forms.ModelForm):
    class Meta:
        model = StockTakeSession
        fields = ['client_name', 'location', 'session_date', 'notes']
        widgets = {
            'client_name': forms.TextInput(attrs={'class': 'form-control'}),
            'location': forms.TextInput(attrs={'class': 'form-control'}),
            'session_date': forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
            'notes': forms.Textarea(attrs={'class': 'form-control', 'rows': 2}),
        }


class SessionListView(PermissionRequiredMixin, ListView):
    model = StockTakeSession
    template_name = 'stock_take/session_list.html'
    context_object_name = 'sessions'
    module_name = 'inventory'
    permission_type = 'view'

    def get_queryset(self):
        return StockTakeSession.objects.all()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Stock Take'
        ctx['can_edit'] = _can_inventory_edit(self.request.user)
        return ctx


class SessionCreateView(PermissionRequiredMixin, CreateView):
    model = StockTakeSession
    form_class = SessionForm
    template_name = 'stock_take/session_form.html'
    module_name = 'inventory'
    permission_type = 'edit'

    def get_initial(self):
        return {'session_date': date.today()}

    def form_valid(self, form):
        form.instance.status = StockTakeSession.STATUS_IN_PROGRESS
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('stock_take:session_scan', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'New Stock Take Session'
        return ctx


class ScanView(PermissionRequiredMixin, TemplateView):
    template_name = 'stock_take/scan.html'
    module_name = 'inventory'
    permission_type = 'view'

    def dispatch(self, request, *args, **kwargs):
        self.session_obj = get_object_or_404(StockTakeSession, pk=kwargs['pk'])
        if self.session_obj.status == StockTakeSession.STATUS_COMPLETED:
            return redirect('stock_take:session_report', pk=self.session_obj.pk)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        s = self.session_obj
        ctx['title'] = f'Stock Take — {s.client_name}'
        ctx['session'] = s
        lines = list(
            s.lines.all().values('sku', 'scan_code', 'item_name', 'expected_qty', 'actual_qty')
        )
        ctx['lines'] = lines
        ctx['lines_json'] = json.dumps(lines, default=str)
        ctx['can_scan'] = _can_inventory_edit(self.request.user)
        ctx['has_lines'] = s.lines.exists()
        ctx['unknown_count'] = s.unknown_scans.count()
        ctx['scan_url'] = reverse('stock_take:record_scan', kwargs={'pk': s.pk})
        ctx['upload_url'] = reverse('stock_take:session_upload', kwargs={'pk': s.pk})
        ctx['complete_url'] = reverse('stock_take:session_complete', kwargs={'pk': s.pk})
        ctx['expected_template_url'] = reverse('stock_take:expected_template', kwargs={'pk': s.pk})
        if s.lines.exists():
            ctx['public_scan_url'] = self.request.build_absolute_uri(
                reverse('stock_take:public_camera', kwargs={'token': s.public_scan_token})
            )
        else:
            ctx['public_scan_url'] = None
        return ctx


class PublicScanCameraView(TemplateView):
    """Mobile camera / wedge page for a session — public URL, no ERP login."""

    template_name = 'stock_take/public_camera.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        token = self.kwargs['token']
        session = StockTakeSession.objects.filter(public_scan_token=token).first()
        if not session or session.status != StockTakeSession.STATUS_IN_PROGRESS:
            ctx['invalid'] = True
            ctx['message'] = (
                'This link is not valid or the stock take session has already been completed.'
            )
            return ctx
        ctx['invalid'] = False
        ctx['session'] = session
        ctx['has_lines'] = session.lines.exists()
        ctx['scan_url'] = reverse(
            'stock_take:public_record_scan', kwargs={'token': session.public_scan_token}
        )
        return ctx


class ExpectedTemplateDownloadView(PermissionRequiredMixin, View):
    """Blank .xlsx matching the upload parser (SKU, Item Name, Barcode or scan code, Expected Qty)."""

    module_name = 'inventory'
    permission_type = 'view'

    def get(self, request, *args, **kwargs):
        from openpyxl import Workbook

        get_object_or_404(
            StockTakeSession, pk=kwargs['pk'], status=StockTakeSession.STATUS_IN_PROGRESS
        )
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        ws.append(['SKU', 'Item Name', 'Barcode or scan code', 'Expected Quantity'])
        ws.append(['DEMO-001', 'Sample widget', '5901234123457', 100])
        ws.append(['DEMO-002', 'Sample case pack', '20012345678903', 24])
        for col, w in (('A', 18), ('B', 28), ('C', 26), ('D', 22)):
            ws.column_dimensions[col].width = w
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return FileResponse(
            buf,
            as_attachment=True,
            filename='stock_take_expected_template.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )


class ReportView(PermissionRequiredMixin, TemplateView):
    template_name = 'stock_take/report.html'
    module_name = 'inventory'
    permission_type = 'view'

    def dispatch(self, request, *args, **kwargs):
        self.session_obj = get_object_or_404(StockTakeSession, pk=kwargs['pk'])
        if self.session_obj.status != StockTakeSession.STATUS_COMPLETED:
            return redirect('stock_take:session_scan', pk=self.session_obj.pk)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        s = self.session_obj
        company = CompanySettings.get_settings()
        ctx['company'] = company
        ctx['stock_take_logo_url'] = (
            self.request.build_absolute_uri(company.logo.url) if company.logo else ''
        )

        def _logo_data_url_for_pdf() -> str:
            """Inline data URL so PDF export works without fetch (same-origin / CDN issues)."""
            if not company.logo:
                return ''
            max_bytes = 900_000
            try:
                with company.logo.open('rb') as fh:
                    raw = fh.read()
            except OSError:
                return ''
            if not raw or len(raw) > max_bytes:
                return ''
            mime, _ = mimetypes.guess_type(company.logo.name)
            mime = mime or 'image/png'
            b64 = base64.b64encode(raw).decode('ascii')
            return f'data:{mime};base64,{b64}'

        ctx['stock_take_logo_data_url_json'] = json.dumps(_logo_data_url_for_pdf())

        def _fmt_ts(dt):
            if not dt:
                return ''
            t = timezone.localtime(dt) if timezone.is_aware(dt) else dt
            return t.strftime('%d/%m/%Y %H:%M')

        ctx['period_start_display'] = _fmt_ts(s.created_at)
        ctx['period_end_display'] = _fmt_ts(s.completed_at) if s.completed_at else '—'

        ctx['title'] = f'Stock Take Report — {s.client_name}'
        ctx['session'] = s
        lines = []
        for line in s.lines.all():
            var = line.actual_qty - line.expected_qty
            if var == 0:
                st = 'Match'
            elif var < 0:
                st = 'Short'
            else:
                st = 'Excess'
            lines.append(
                {
                    'sku': line.sku,
                    'scan_code': line.scan_code or '',
                    'item_name': line.item_name,
                    'expected_qty': line.expected_qty,
                    'actual_qty': line.actual_qty,
                    'variance': var,
                    'status': st,
                }
            )
        ctx['lines'] = lines
        ctx['unknown_scans'] = list(
            s.unknown_scans.all().values('barcode_raw', 'timestamp')[:500]
        )
        matched = sum(1 for x in lines if x['status'] == 'Match')
        short = sum(1 for x in lines if x['status'] == 'Short')
        excess = sum(1 for x in lines if x['status'] == 'Excess')
        unk = s.unknown_scans.count()
        ctx['summary'] = {
            'total_skus': len(lines),
            'matched': matched,
            'short': short,
            'excess': excess,
            'unidentified': unk,
        }
        ctx['lines_json'] = json.dumps(lines, default=str)
        ctx['summary_json'] = json.dumps(ctx['summary'], default=str)
        unk_rows = list(s.unknown_scans.all().values('barcode_raw', 'timestamp')[:500])
        ctx['unknown_json'] = json.dumps(
            [
                {'barcode_raw': u['barcode_raw'], 'timestamp': u['timestamp'].isoformat()}
                for u in unk_rows
            ],
            default=str,
        )
        return ctx


@login_required
@require_POST
def session_upload_expected(request, pk):
    """Parse .xlsx in the web worker for this HTTP request (does not run on the mobile UI thread)."""
    if not _can_inventory_edit(request.user):
        return JsonResponse({'ok': False, 'error': 'Permission denied.'}, status=403)
    session = get_object_or_404(
        StockTakeSession, pk=pk, status=StockTakeSession.STATUS_IN_PROGRESS
    )
    if session.scan_logs.exists():
        return JsonResponse(
            {'ok': False, 'error': 'Cannot replace the list after scans have been recorded.'},
            status=400,
        )
    f = request.FILES.get('file')
    if not f:
        return JsonResponse({'ok': False, 'error': 'No file uploaded.'}, status=400)
    if not f.name.lower().endswith(('.xlsx',)):
        return JsonResponse({'ok': False, 'error': 'Please upload an .xlsx file.'}, status=400)
    try:
        rows = _parse_expected_workbook(f)
    except ValueError as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'Could not read file: {e}'}, status=400)

    with transaction.atomic():
        session.lines.all().delete()
        StockTakeLine.objects.bulk_create(
            [
                StockTakeLine(
                    session=session,
                    sku=r['sku'],
                    scan_code=r.get('scan_code') or '',
                    item_name=r['item_name'],
                    expected_qty=r['expected_qty'],
                    actual_qty=Decimal('0'),
                )
                for r in rows
            ]
        )
    return JsonResponse({'ok': True, 'count': len(rows)})


def _line_for_scan_lookup(session, raw_norm):
    """Match shelf label: non-empty scan_code first (case-insensitive), then SKU."""
    raw_norm = _normalize_barcode_scan_key((raw_norm or '').strip())
    if not raw_norm:
        return None
    qs = StockTakeLine.objects.select_for_update().filter(session=session)
    # Legacy rows may still have Excel float strings (e.g. '....0') in the DB
    variants = [raw_norm]
    if raw_norm.isdigit():
        variants.append(f'{raw_norm}.0')
    seen = set()
    uniq = []
    for v in variants:
        k = v.casefold()
        if k not in seen:
            seen.add(k)
            uniq.append(v)
    for v in uniq:
        line = qs.filter(~Q(scan_code=''), scan_code__iexact=v).first()
        if line:
            return line
    for v in uniq:
        line = qs.filter(sku__iexact=v).first()
        if line:
            return line
    return None


def _record_scan_payload(session, data):
    """Barcode increment or manual set_actual. Returns (body_dict, http_status)."""
    if 'set_actual' in data:
        sku_key = _normalize_barcode_scan_key((data.get('sku') or '').strip())
        if not sku_key:
            return ({'ok': False, 'error': 'SKU is required.'}, 400)
        try:
            new_val = Decimal(str(data.get('set_actual')))
        except (InvalidOperation, TypeError, ValueError):
            return ({'ok': False, 'error': 'Invalid quantity.'}, 400)
        if new_val < 0:
            return ({'ok': False, 'error': 'Quantity cannot be negative.'}, 400)
        with transaction.atomic():
            line = _line_for_scan_lookup(session, sku_key)
            if not line:
                StockTakeUnknownScan.objects.create(
                    session=session,
                    barcode_raw=f'MANUAL:{sku_key}',
                )
                StockTakeScanLog.objects.create(
                    session=session,
                    sku='',
                    barcode_raw=f'MANUAL:{sku_key}',
                    actual_qty_after=Decimal('0'),
                    matched=False,
                )
                return ({'ok': True, 'matched': False, 'unknown': True}, 200)
            line.actual_qty = new_val
            line.save(update_fields=['actual_qty'])
            StockTakeScanLog.objects.create(
                session=session,
                sku=line.sku,
                barcode_raw=f'MANUAL_SET:{sku_key}={new_val}',
                actual_qty_after=line.actual_qty,
                matched=True,
            )
        return (
            {
                'ok': True,
                'matched': True,
                'sku': line.sku,
                'actual_qty': str(line.actual_qty),
                'expected_qty': str(line.expected_qty),
            },
            200,
        )

    raw = _normalize_barcode_scan_key((data.get('barcode') or '').strip())
    if not raw:
        return ({'ok': False, 'error': 'Empty barcode.'}, 400)

    with transaction.atomic():
        line = _line_for_scan_lookup(session, raw)
        auto_created = False
        if line:
            StockTakeLine.objects.filter(pk=line.pk).update(actual_qty=F('actual_qty') + 1)
            line.refresh_from_db()
        else:
            # Not in expected list: add a new line (unknown item, expected 0, count from 1).
            sku_key = raw[:120]
            sc_key = raw[:200]
            try:
                line = StockTakeLine.objects.create(
                    session=session,
                    sku=sku_key,
                    scan_code=sc_key,
                    item_name='Unknown',
                    expected_qty=Decimal('0'),
                    actual_qty=Decimal('1'),
                )
                auto_created = True
            except IntegrityError:
                line = _line_for_scan_lookup(session, raw)
                if line is None:
                    raise
                StockTakeLine.objects.filter(pk=line.pk).update(actual_qty=F('actual_qty') + 1)
                line.refresh_from_db()

        StockTakeScanLog.objects.create(
            session=session,
            sku=line.sku,
            barcode_raw=raw,
            actual_qty_after=line.actual_qty,
            matched=True,
        )

    return (
        {
            'ok': True,
            'matched': True,
            'sku': line.sku,
            'item_name': line.item_name,
            'expected_qty': str(line.expected_qty),
            'actual_qty': str(line.actual_qty),
            'auto_created': auto_created,
        },
        200,
    )


@login_required
@require_POST
def record_scan(request, pk):
    if not _can_inventory_edit(request.user):
        return JsonResponse({'ok': False, 'error': 'Permission denied.'}, status=403)
    session = get_object_or_404(
        StockTakeSession, pk=pk, status=StockTakeSession.STATUS_IN_PROGRESS
    )
    try:
        data = json.loads(request.body.decode('utf-8') or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': 'Invalid JSON.'}, status=400)

    payload, status = _record_scan_payload(session, data)
    return JsonResponse(payload, status=status)


@csrf_exempt
@require_POST
def public_record_scan(request, token):
    session = StockTakeSession.objects.filter(
        public_scan_token=token,
        status=StockTakeSession.STATUS_IN_PROGRESS,
    ).first()
    if not session:
        return JsonResponse(
            {'ok': False, 'error': 'Session not found or no longer active.'},
            status=404,
        )
    try:
        data = json.loads(request.body.decode('utf-8') or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': 'Invalid JSON.'}, status=400)

    payload, status = _record_scan_payload(session, data)
    return JsonResponse(payload, status=status)


@login_required
@require_POST
def session_complete(request, pk):
    if not _can_inventory_edit(request.user):
        return JsonResponse({'ok': False, 'error': 'Permission denied.'}, status=403)
    session = get_object_or_404(
        StockTakeSession, pk=pk, status=StockTakeSession.STATUS_IN_PROGRESS
    )
    if not session.lines.exists():
        return JsonResponse(
            {'ok': False, 'error': 'Upload an expected stock list before completing.'},
            status=400,
        )
    session.status = StockTakeSession.STATUS_COMPLETED
    session.completed_at = timezone.now()
    session.save(update_fields=['status', 'completed_at'])
    return JsonResponse(
        {'ok': True, 'redirect': reverse('stock_take:session_report', kwargs={'pk': session.pk})}
    )
