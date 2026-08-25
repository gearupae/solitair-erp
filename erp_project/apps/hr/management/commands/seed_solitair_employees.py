"""
Seed Solitair users, departments, and approval configuration from GearUp employee Excel.

Excel layout (Sheet1):
  Row 2: department names in paired columns (name at col, email at col+1 starts row 3)
  Row 3: Requester
  Row 4: First Manager
  Row 5: Second Manager
  Row 7: Initiator (global)
  Rows 8-10: Approval (value) approvers (global, multi-level by amount)

Run:
  python manage.py seed_solitair_employees
  python manage.py seed_solitair_employees --file "/path/to/file.xlsx"
"""
from __future__ import annotations

import re
from decimal import Decimal
from pathlib import Path

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand
from django.db import transaction

from apps.hr.models import Department, Designation, Employee
from apps.hr.user_provisioning import _unique_username
from apps.settings_app.models import ApprovalConfiguration, ApprovalConfigurationLevel, Role, UserProfile, UserRole

User = get_user_model()

DEPT_CODES = {
    'Marketing': 'MKT',
    'HR': 'HR',
    'Finance': 'FIN',
    'Sales': 'SAL',
    'Engineering': 'ENG',
    'IOCC': 'IOCC',
    'Quality & Safety': 'QAS',
    'Ground Operations': 'GROPS',
    'Flight Operations': 'FLOPS',
    'Training': 'TRN',
    'Admin': 'ADM',
    'IT': 'IT',
}

ROLE_PRIORITY = {
    'First Manager': 4,
    'Second Manager': 3,
    'Requester': 2,
    'Initiator': 2,
    'Approval (value)': 5,
}

ERP_ROLE_BY_EXCEL = {
    'Requester': 'purchase',
    'Initiator': 'purchase',
    'First Manager': 'manager',
    'Second Manager': 'manager',
    'Approval (value)': 'manager',
}

APPROVAL_LEVELS = (
    (Decimal('25000'), 'shegde@solitairholding.com'),
    (Decimal('100000'), 'thawada@solitairholding.com'),
    (Decimal('999999999'), 'shamadeh@solitairholding.com'),
)

APPROVAL_MODULES = (
    'purchase_request',
    'purchase_order',
    'service_request',
    'vendor_bill',
)


def _default_password() -> str:
    return getattr(settings, 'HR_EMPLOYEE_DEFAULT_PASSWORD', 'AlNajahEmployee123!')


def _dept_code(name: str) -> str:
    return DEPT_CODES.get(name, re.sub(r'[^A-Z0-9]', '', name.upper())[:10] or 'DEPT')


def _split_name(full_name: str) -> tuple[str, str]:
    parts = (full_name or '').strip().split(None, 1)
    if not parts:
        return 'User', ''
    return parts[0], parts[1] if len(parts) > 1 else ''


def _parse_excel(path: Path) -> tuple[list[str], dict, dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    departments: list[str] = []
    first_managers: dict[str, str] = {}
    col = 3
    while col <= ws.max_column:
        name = ws.cell(row=3, column=col).value
        if name and str(name).strip():
            dept_name = str(name).strip()
            departments.append(dept_name)
            mgr_email = ws.cell(row=5, column=col + 1).value
            if mgr_email and '@' in str(mgr_email):
                first_managers[dept_name] = str(mgr_email).strip().lower()
        col += 2

    people: dict[str, dict] = {}
    row_roles = {
        4: 'Requester',
        5: 'First Manager',
        6: 'Second Manager',
    }

    for idx, dept in enumerate(departments):
        col = 3 + idx * 2
        for row, excel_role in row_roles.items():
            name = ws.cell(row=row, column=col).value
            email = ws.cell(row=row, column=col + 1).value
            if not name or not email:
                continue
            email = str(email).strip().lower()
            if '@' not in email:
                continue
            entry = people.setdefault(
                email,
                {
                    'name': str(name).strip(),
                    'email': email,
                    'departments': set(),
                    'roles': set(),
                },
            )
            entry['departments'].add(dept)
            entry['roles'].add(excel_role)

    init_name = ws.cell(row=8, column=3).value
    init_email = ws.cell(row=8, column=4).value
    if init_name and init_email:
        email = str(init_email).strip().lower()
        people.setdefault(
            email,
            {'name': str(init_name).strip(), 'email': email, 'departments': set(), 'roles': set()},
        )['roles'].add('Initiator')

    for row in (9, 10, 11):
        name = ws.cell(row=row, column=3).value
        email = ws.cell(row=row, column=4).value
        if not name or not email:
            continue
        email = str(email).strip().lower()
        people.setdefault(
            email,
            {'name': str(name).strip(), 'email': email, 'departments': set(), 'roles': set()},
        )['roles'].add('Approval (value)')

    wb.close()
    return departments, people, first_managers


def _primary_role(roles: set[str]) -> str:
    return max(roles, key=lambda r: ROLE_PRIORITY.get(r, 0))


def _primary_department(roles: set[str], departments: set[str]) -> str | None:
    if not departments:
        return None
    if 'Requester' in roles:
        return sorted(departments)[0]
    if 'First Manager' in roles:
        return sorted(departments)[0]
    if 'Second Manager' in roles:
        return sorted(departments)[0]
    return sorted(departments)[0]


class Command(BaseCommand):
    help = 'Seed departments, users, employees, and approval config from GearUp employee Excel.'

    def add_arguments(self, parser):
        default_file = (
            Path(settings.BASE_DIR).parent
            / 'Excel Upload'
            / 'gearup_employee_list.xlsx'
        )
        parser.add_argument(
            '--file',
            default=str(default_file),
            help='Path to GearUp employee list Excel file.',
        )

    @transaction.atomic
    def handle(self, *args, **options):
        path = Path(options['file'])
        if not path.is_file():
            self.stderr.write(self.style.ERROR(f'File not found: {path}'))
            return

        departments, people, first_managers = _parse_excel(path)
        self.stdout.write(f'Parsed {len(departments)} departments, {len(people)} people from {path.name}')

        dept_objs: dict[str, Department] = {}
        for name in departments:
            code = _dept_code(name)
            dept, created = Department.objects.update_or_create(
                code=code,
                defaults={'name': name, 'is_active': True},
            )
            dept_objs[name] = dept
            self.stdout.write(f'  Department {"created" if created else "updated"}: {name} ({code})')

        role_cache = {r.code: r for r in Role.objects.filter(is_active=True)}
        user_by_email: dict[str, User] = {}
        password = _default_password()

        for email, info in sorted(people.items(), key=lambda x: x[1]['name']):
            first, last = _split_name(info['name'])
            primary_role = _primary_role(info['roles'])
            dept_name = _primary_department(info['roles'], info['departments'])
            dept = dept_objs.get(dept_name) if dept_name else None

            user = User.objects.filter(email__iexact=email).first()
            if not user:
                username = _unique_username(email, '')
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first[:150],
                    last_name=last[:150],
                    is_active=True,
                )
                created_user = True
            else:
                user.first_name = first[:150]
                user.last_name = last[:150]
                user.email = email
                user.is_active = True
                user.save(update_fields=['first_name', 'last_name', 'email', 'is_active'])
                created_user = False

            UserProfile.objects.get_or_create(user=user)

            erp_role_code = ERP_ROLE_BY_EXCEL.get(primary_role, 'employee')
            erp_role = role_cache.get(erp_role_code)
            if erp_role:
                UserRole.objects.update_or_create(
                    user=user,
                    role=erp_role,
                    defaults={'is_active': True},
                )

            desig_name = primary_role
            designation = None
            if dept:
                designation, _ = Designation.objects.get_or_create(
                    name=desig_name,
                    department=dept,
                    defaults={'is_active': True},
                )

            emp_code_base = re.sub(r'[^A-Z0-9]', '', email.split('@', 1)[0].upper())[:12] or 'EMP'
            emp_code = emp_code_base
            n = 1
            while (
                Employee.objects.filter(employee_code=emp_code)
                .exclude(email__iexact=email)
                .exists()
            ):
                n += 1
                emp_code = f'{emp_code_base}{n}'

            employee = Employee.objects.filter(email__iexact=email).first()
            emp_defaults = {
                'employee_code': emp_code,
                'user': user,
                'first_name': first,
                'last_name': last,
                'department': dept,
                'designation': designation,
                'status': 'active',
                'is_active': True,
            }
            if employee:
                for key, value in emp_defaults.items():
                    setattr(employee, key, value)
                employee.save()
                emp_created = False
            else:
                employee = Employee.objects.create(email=email, **emp_defaults)
                emp_created = True
            user_by_email[email] = user
            self.stdout.write(
                f'  User {"created" if created_user else "updated"}: {info["name"]} <{email}> '
                f'[{", ".join(sorted(info["roles"]))}]'
            )

        for dept_name, mgr_email in first_managers.items():
            manager = user_by_email.get(mgr_email)
            if manager and dept_name in dept_objs:
                dept = dept_objs[dept_name]
                dept.manager = manager
                dept.save(update_fields=['manager'])
                self.stdout.write(
                    f'  Manager for {dept_name}: {manager.get_full_name() or manager.username}'
                )

        # Multi-level value approvers
        for module in APPROVAL_MODULES:
            config, _ = ApprovalConfiguration.objects.update_or_create(
                module=module,
                defaults={
                    'approval_type': 'multi',
                    'is_active': True,
                },
            )
            config.levels.all().delete()
            for order, (amount, approver_email) in enumerate(APPROVAL_LEVELS):
                approver = user_by_email.get(approver_email)
                if not approver:
                    self.stderr.write(self.style.WARNING(f'Approver not found: {approver_email}'))
                    continue
                ApprovalConfigurationLevel.objects.create(
                    configuration=config,
                    amount_threshold=amount,
                    approver=approver,
                    order=order,
                    is_active=True,
                )
            self.stdout.write(
                f'  Approval config: {module} → multi-level '
                f'({", ".join(e for _, e in APPROVAL_LEVELS)})'
            )

        self.stdout.write(self.style.SUCCESS(
            f'\nDone. {len(people)} users seeded. Default password: {password}'
        ))
