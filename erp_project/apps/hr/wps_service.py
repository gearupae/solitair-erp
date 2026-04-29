"""UAE Central Bank WPS Salary Information File (SIF) — EDR + SCR pipe-delimited format."""
from __future__ import annotations

import re
from calendar import monthrange
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import Any

from django.utils import timezone

from apps.settings_app.models import Company


def strip_mol_personal_number(emirates_id: str) -> str:
    """MOL personal number: digits only (strip 784-YYYY-XXXXXXX-X dashes)."""
    if not emirates_id:
        return ''
    return re.sub(r'\D', '', str(emirates_id).strip())


def sanitize_pipe(value: str | None) -> str:
    if value is None:
        return ''
    return str(value).replace('|', '').replace('\n', '').replace('\r', '')


def aed_net_to_fils(net: Decimal) -> int:
    return int((net * 100).quantize(Decimal('1'), rounding=ROUND_HALF_UP))


def last_day_of_month(mf: date) -> date:
    return date(mf.year, mf.month, monthrange(mf.year, mf.month)[1])


def normalize_iban(raw: str | None) -> str:
    if not raw:
        return ''
    return re.sub(r'\s+', '', str(raw).strip())


def build_uae_central_bank_sif(
    company: Company,
    month_first: date,
    edr_rows: list[dict[str, Any]],
    *,
    upload_dt: datetime | None = None,
) -> str:
    """
    Build SIF file: EDR rows then one SCR row.
    edr_rows items: mol_personal, name, routing, iban, period_start, period_end,
    working_days (int), net_fils (int).
    """
    if upload_dt is None:
        upload_dt = timezone.localtime()
    elif timezone.is_naive(upload_dt):
        upload_dt = timezone.make_aware(upload_dt, timezone.get_current_timezone())

    upload_date = upload_dt.strftime('%Y-%m-%d')
    upload_time = upload_dt.strftime('%H%M')
    mm_yyyy = f'{month_first.month:02d}{month_first.year}'

    lines: list[str] = []
    total_fils = 0
    for r in edr_rows:
        try:
            nf = int(r.get('net_fils') or 0)
        except (TypeError, ValueError):
            nf = 0
        total_fils += nf
        try:
            wd = int(r.get('working_days') or 0)
        except (TypeError, ValueError):
            wd = 0
        ps = str(r.get('period_start') or month_first.isoformat())
        pe = str(r.get('period_end') or last_day_of_month(month_first).isoformat())
        lines.append(
            '|'.join(
                [
                    'EDR',
                    sanitize_pipe(str(r.get('mol_personal', ''))),
                    sanitize_pipe(str(r.get('name', ''))),
                    sanitize_pipe(str(r.get('routing', ''))),
                    sanitize_pipe(str(r.get('iban', ''))),
                    ps,
                    pe,
                    str(wd),
                    str(nf),
                    '0',
                    '0',
                ]
            )
        )

    mol = sanitize_pipe((company.mol_number or '').strip())
    co_iban = sanitize_pipe(normalize_iban(company.bank_iban or ''))
    co_routing = sanitize_pipe((company.bank_routing_code or '').strip())

    lines.append(
        '|'.join(
            [
                'SCR',
                mol,
                '',
                co_routing,
                upload_date,
                upload_time,
                mm_yyyy,
                str(len(edr_rows)),
                str(total_fils),
                'AED',
                co_iban,
            ]
        )
    )
    return '\n'.join(lines) + '\n'


def collect_wps_payload(company: Company, month_first: date) -> dict[str, Any]:
    """
    Paid payrolls for company + month. Returns edr_rows, preview_rows, warnings,
    missing_iban (names), employee_count, total_aed.
    """
    from apps.hr.models import Payroll
    from apps.hr.models_extended import AttendanceSummary, PayrollSettings
    from apps.hr.payroll_allowances import payrolls_for_company_entity

    qs = (
        Payroll.objects.filter(month=month_first, status='paid', is_active=True)
        .select_related('employee', 'company', 'employee__bank_detail', 'employee__uae_compliance')
    )
    qs = payrolls_for_company_entity(qs, company).order_by('employee__employee_code')

    ps = PayrollSettings.objects.filter(pk=1).first()
    default_wd = int(ps.working_days_in_month) if ps else 26

    edr_rows: list[dict[str, Any]] = []
    preview_rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    missing_iban: list[str] = []
    missing_mol: list[str] = []
    total_aed = Decimal('0')

    if not normalize_iban(company.bank_iban or ''):
        warnings.append('Company bank IBAN is missing — SCR may be incomplete.')
    if not (company.mol_number or '').strip():
        warnings.append('Company MOL (employer) number is missing — SCR may be incomplete.')
    if not (company.bank_routing_code or '').strip():
        warnings.append('Company bank routing code is missing — SCR may be incomplete.')

    for p in qs:
        emp = p.employee
        uc = getattr(emp, 'uae_compliance', None)
        bd = getattr(emp, 'bank_detail', None)
        mol = strip_mol_personal_number(emp.emirates_id or '')
        iban = normalize_iban(
            (uc.bank_iban if uc and uc.bank_iban else '') or (bd.iban if bd else '')
        )
        routing = (uc.bank_routing_code if uc and uc.bank_routing_code else '') or (
            bd.routing_bank_code if bd else ''
        )
        if not iban:
            missing_iban.append(emp.full_name or emp.employee_code)
        if not mol:
            missing_mol.append(emp.full_name or emp.employee_code)

        summ = AttendanceSummary.objects.filter(employee=emp, month=month_first).first()
        if summ and summ.total_working_days:
            wd = int(summ.total_working_days)
        else:
            wd = default_wd

        net = p.net_salary or Decimal('0')
        total_aed += net
        net_fils = aed_net_to_fils(net)
        period_start = month_first.isoformat()
        period_end = last_day_of_month(month_first).isoformat()
        name = f'{emp.first_name} {emp.last_name}'.strip()

        edr_rows.append(
            {
                'mol_personal': mol,
                'name': name or emp.employee_code,
                'routing': routing or '',
                'iban': iban,
                'period_start': period_start,
                'period_end': period_end,
                'working_days': wd,
                'net_fils': net_fils,
            }
        )

        status = 'OK' if (iban and mol) else 'Incomplete'
        preview_rows.append(
            {
                'employee': name or emp.employee_code,
                'iban': iban or '—',
                'working_days': wd,
                'net_aed': net,
                'status': status,
            }
        )

    if missing_iban:
        tail = ', '.join(missing_iban[:15])
        if len(missing_iban) > 15:
            tail += '…'
        warnings.append(f'{len(missing_iban)} employee(s) missing bank IBAN: {tail}')
    if missing_mol:
        tail = ', '.join(missing_mol[:15])
        if len(missing_mol) > 15:
            tail += '…'
        warnings.append(
            f'{len(missing_mol)} employee(s) missing Emirates ID / MOL personal number: {tail}'
        )

    return {
        'edr_rows': edr_rows,
        'preview_rows': preview_rows,
        'warnings': warnings,
        'missing_iban': missing_iban,
        'missing_mol': missing_mol,
        'employee_count': len(edr_rows),
        'total_aed': total_aed,
    }


def wps_sif_filename(company: Company, month_first: date) -> str:
    mol = (company.mol_number or 'UNKNOWN').strip() or 'UNKNOWN'
    mol_safe = re.sub(r'[^\w\-]', '_', mol)
    mmyyyy = f'{month_first.month:02d}{month_first.year}'
    return f'WPS_{mol_safe}_{mmyyyy}.SIF'


def wps_excel_filename(company: Company, month_first: date) -> str:
    mol = (company.mol_number or 'UNKNOWN').strip() or 'UNKNOWN'
    mol_safe = re.sub(r'[^\w\-]', '_', mol)
    mmyyyy = f'{month_first.month:02d}{month_first.year}'
    return f'WPS_{mol_safe}_{mmyyyy}.xlsx'


def build_wps_excel_bytes(company: Company, month_first: date, payload: dict[str, Any]) -> bytes:
    """
    Excel workbook: employee lines (review/edit) + raw SIF lines sheet + SCR summary.
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_emp = wb.active
    ws_emp.title = 'Employees'
    hdr_font = Font(bold=True)
    emp_headers = [
        'Type',
        'MOL personal',
        'Employee name',
        'Routing code',
        'IBAN',
        'Period start',
        'Period end',
        'Working days',
        'Net (AED)',
        'Net (fils)',
        'Notes',
    ]
    for col, h in enumerate(emp_headers, start=1):
        c = ws_emp.cell(row=1, column=col, value=h)
        c.font = hdr_font

    row_idx = 2
    for r in payload.get('edr_rows') or []:
        try:
            nf = int(r.get('net_fils') or 0)
        except (TypeError, ValueError):
            nf = 0
        net_aed = (Decimal(nf) / Decimal('100')).quantize(Decimal('0.01'))
        try:
            wd = int(r.get('working_days') or 0)
        except (TypeError, ValueError):
            wd = 0
        mol = str(r.get('mol_personal') or '').strip()
        iban = str(r.get('iban') or '').strip()
        notes = []
        if not iban:
            notes.append('missing IBAN')
        if not mol:
            notes.append('missing MOL personal')
        ws_emp.cell(row=row_idx, column=1, value='EDR')
        ws_emp.cell(row=row_idx, column=2, value=mol)
        ws_emp.cell(row=row_idx, column=3, value=str(r.get('name') or ''))
        ws_emp.cell(row=row_idx, column=4, value=str(r.get('routing') or ''))
        ws_emp.cell(row=row_idx, column=5, value=str(r.get('iban') or ''))
        ws_emp.cell(row=row_idx, column=6, value=str(r.get('period_start') or ''))
        ws_emp.cell(row=row_idx, column=7, value=str(r.get('period_end') or ''))
        ws_emp.cell(row=row_idx, column=8, value=wd)
        ws_emp.cell(row=row_idx, column=9, value=float(net_aed))
        ws_emp.cell(row=row_idx, column=10, value=nf)
        ws_emp.cell(row=row_idx, column=11, value='; '.join(notes) if notes else '')
        row_idx += 1

    ws_sif = wb.create_sheet('SIF_lines')
    sif_text = build_uae_central_bank_sif(company, month_first, payload.get('edr_rows') or [])
    ws_sif.cell(row=1, column=1, value='Line (pipe-delimited, for WPS upload)').font = hdr_font
    for i, line in enumerate(sif_text.strip().split('\n'), start=2):
        ws_sif.cell(row=i, column=1, value=line)

    ws_info = wb.create_sheet('Export_info')
    ws_info.cell(row=1, column=1, value='Company').font = hdr_font
    ws_info.cell(row=1, column=2, value=company.name)
    ws_info.cell(row=2, column=1, value='Payroll month').font = hdr_font
    ws_info.cell(row=2, column=2, value=month_first.strftime('%Y-%m'))
    ws_info.cell(row=3, column=1, value='Employees (paid)').font = hdr_font
    ws_info.cell(row=3, column=2, value=len(payload.get('edr_rows') or []))
    ws_info.cell(row=4, column=1, value='Total net AED').font = hdr_font
    ws_info.cell(row=4, column=2, value=float(payload.get('total_aed') or Decimal('0')))
    ws_info.cell(row=6, column=1, value='Warnings / incomplete data').font = hdr_font
    warn_row = 7
    for w in payload.get('warnings') or []:
        ws_info.cell(row=warn_row, column=1, value=str(w))
        warn_row += 1
    if not (payload.get('warnings') or []):
        ws_info.cell(row=warn_row, column=1, value='—')

    for ws in (ws_emp, ws_sif, ws_info):
        for idx in range(1, (ws.max_column or 1) + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 22

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_and_store_wps_for_month(month_first: date) -> str:
    """Cron: store first UAE company SIF with paid rows for the month in WPSMonthlyFile."""
    from apps.hr.models_extended import WPSMonthlyFile

    uae_companies = Company.objects.filter(is_active=True, country='uae').order_by('pk')
    content = ''
    all_paid = False
    for co in uae_companies:
        data = collect_wps_payload(co, month_first)
        if data['employee_count']:
            content = build_uae_central_bank_sif(co, month_first, data['edr_rows'])
            all_paid = True
            break
    WPSMonthlyFile.objects.update_or_create(
        month=month_first,
        defaults={'file_content': content, 'all_payrolls_paid': all_paid},
    )
    return content
