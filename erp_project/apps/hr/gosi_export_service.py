"""KSA GOSI export — payroll-driven rows, Excel (.xlsx), DB sync via GOSIRecord."""
from __future__ import annotations

import re
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Any

from django.utils import timezone

from apps.settings_app.models import Company


def gosi_contribution_rates(nationality: str) -> tuple[Decimal, Decimal]:
    """Employee rate, employer rate (decimal fractions)."""
    if (nationality or '').lower() == 'saudi':
        return Decimal('0.10'), Decimal('0.12')
    return Decimal('0'), Decimal('0.02')


def nationality_label(nationality: str) -> str:
    return 'Saudi' if (nationality or '').lower() == 'saudi' else 'Non-Saudi'


def rate_display(rate: Decimal) -> str:
    return f'{rate * 100:g}%'


def sanitize_filename_part(name: str) -> str:
    return re.sub(r'[^\w\-]+', '_', (name or 'company').strip())[:80] or 'company'


def gosi_xlsx_filename(company: Company, month_first: date) -> str:
    part = sanitize_filename_part(company.name)
    return f'GOSI_{part}_{month_first.month:02d}_{month_first.year}.xlsx'


def sync_gosi_record_for_payroll(payroll) -> bool:
    """
    Upsert GOSIRecord from payroll + KSA compliance (processed/paid payrolls).
    Does not alter deduction lines. Returns True if a record was written.
    """
    emp = payroll.employee
    if (getattr(emp, 'location', None) or '').strip().lower() != 'ksa':
        return False
    kc = getattr(emp, 'ksa_compliance', None)
    if not kc or not kc.gosi_applicable:
        return False
    basic = payroll.basic_salary or Decimal('0')
    nat = (kc.nationality or 'non_saudi').lower()
    emp_rate, er_rate = gosi_contribution_rates(nat)
    emp_c = (basic * emp_rate).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    er_c = (basic * er_rate).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    month_start = date(payroll.month.year, payroll.month.month, 1)
    _sync_gosi_record(
        payroll,
        kc=kc,
        basic=basic,
        emp_rate=emp_rate,
        er_rate=er_rate,
        emp_c=emp_c,
        er_c=er_c,
        month_start=month_start,
    )
    return True


def _sync_gosi_record(
    payroll,
    *,
    kc,
    basic: Decimal,
    emp_rate: Decimal,
    er_rate: Decimal,
    emp_c: Decimal,
    er_c: Decimal,
    month_start: date,
) -> None:
    from apps.hr.models_extended import GOSIRecord

    emp = payroll.employee
    total = (emp_c + er_c).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    co = payroll.company or getattr(emp, 'company', None)
    nat_code = (kc.nationality or 'non_saudi').lower()
    GOSIRecord.objects.update_or_create(
        payroll=payroll,
        defaults={
            'employee': emp,
            'company': co,
            'month': month_start,
            'basic_salary': basic,
            'nationality': nat_code,
            'gosi_number': (kc.gosi_number or '')[:80],
            'iqama_number': (kc.iqama_number or '')[:20],
            'employee_rate': emp_rate,
            'employer_rate': er_rate,
            'employee_contribution': emp_c,
            'employer_contribution': er_c,
            'total_contribution': total,
            'gross_up_basic_for_rates': basic,
        },
    )


def collect_gosi_payload(company: Company, month_first: date, *, sync_records: bool = True) -> dict[str, Any]:
    """
    Processed/paid KSA-location payrolls for the company entity.
    Skips employees without KSA compliance or with GOSI not applicable.
    Optionally upserts GOSIRecord per row.
    """
    from apps.hr.models import Payroll
    from apps.hr.payroll_allowances import payrolls_for_company_entity

    qs = (
        Payroll.objects.filter(
            month=month_first,
            status__in=['processed', 'paid'],
            is_active=True,
            employee__location='ksa',
        )
        .select_related('employee', 'company', 'employee__ksa_compliance')
    )
    qs = payrolls_for_company_entity(qs, company).order_by('employee__employee_code')

    warnings: list[str] = []
    rows: list[dict[str, Any]] = []
    missing_gosi: list[str] = []
    missing_iqama: list[str] = []

    payrolls_scanned = qs.count()
    excluded_not_applicable = 0

    total_saudi = 0
    total_non_saudi = 0
    total_saudi_gosi = Decimal('0')
    total_non_saudi_hazard = Decimal('0')
    sum_employee = Decimal('0')
    sum_employer = Decimal('0')
    grand_total = Decimal('0')

    for p in qs.iterator():
        emp = p.employee
        kc = getattr(emp, 'ksa_compliance', None)
        if not kc or not kc.gosi_applicable:
            excluded_not_applicable += 1
            continue

        nationality = (kc.nationality or 'non_saudi').lower()
        gosi_num = (kc.gosi_number or '').strip()
        iqama = (kc.iqama_number or '').strip()

        if not gosi_num:
            missing_gosi.append(emp.full_name or emp.employee_code or str(emp.pk))
        if not iqama:
            missing_iqama.append(emp.full_name or emp.employee_code or str(emp.pk))

        basic = p.basic_salary or Decimal('0')
        emp_rate, er_rate = gosi_contribution_rates(nationality)
        emp_c = (basic * emp_rate).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        er_c = (basic * er_rate).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_c = (emp_c + er_c).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        month_disp = f'{month_first.month:02d}/{month_first.year}'
        co_name = (emp.company.name if getattr(emp, 'company_id', None) and emp.company else '') or company.name

        if sync_records:
            sync_gosi_record_for_payroll(p)

        rows.append(
            {
                'employee_code': emp.employee_code or '',
                'employee_name': emp.full_name,
                'gosi_number': gosi_num,
                'nationality': nationality_label(nationality),
                'nationality_code': nationality,
                'id_number': iqama,
                'basic_salary': basic,
                'employee_rate': rate_display(emp_rate),
                'employee_contribution': emp_c,
                'employer_rate': rate_display(er_rate),
                'employer_contribution': er_c,
                'total_contribution': total_c,
                'month': month_disp,
                'company_name': co_name,
            }
        )

        if nationality == 'saudi':
            total_saudi += 1
            total_saudi_gosi += total_c
        else:
            total_non_saudi += 1
            total_non_saudi_hazard += er_c
        sum_employee += emp_c
        sum_employer += er_c
        grand_total += total_c

    if missing_gosi:
        warnings.append(
            f'{len(missing_gosi)} employee(s) missing GOSI number: '
            + ', '.join(missing_gosi[:12])
            + ('…' if len(missing_gosi) > 12 else '')
        )
    if missing_iqama:
        warnings.append(
            f'{len(missing_iqama)} employee(s) missing Iqama / ID number: '
            + ', '.join(missing_iqama[:12])
            + ('…' if len(missing_iqama) > 12 else '')
        )

    grand_total = grand_total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    return {
        'rows': rows,
        'warnings': warnings,
        'employee_count': len(rows),
        'payrolls_scanned': payrolls_scanned,
        'excluded_not_applicable': excluded_not_applicable,
        'total_saudi': total_saudi,
        'total_non_saudi': total_non_saudi,
        'total_saudi_gosi': total_saudi_gosi,
        'total_non_saudi_hazard': total_non_saudi_hazard,
        'total_employee_contributions': sum_employee,
        'total_employer_contributions': sum_employer,
        'grand_total': grand_total,
    }


def build_gosi_excel_bytes(payload: dict[str, Any], company: Company, month_first: date) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'GOSI Report'

    header_fill = PatternFill('solid', fgColor='1F7A4A')
    header_font = Font(bold=True, color='FFFFFF')
    money_fmt = '#,##0.00'

    headers = [
        'Employee Code',
        'Employee Name',
        'GOSI Number',
        'Nationality',
        'Iqama / National ID',
        'Basic Salary (SAR)',
        'Employee Rate',
        'Employee Contribution (SAR)',
        'Employer Rate',
        'Employer Contribution (SAR)',
        'Total Contribution (SAR)',
        'Month',
        'Company',
    ]
    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    grey_fill = PatternFill('solid', fgColor='F2F2F2')
    white_fill = PatternFill('solid', fgColor='FFFFFF')
    rows = payload.get('rows') or []
    for i, r in enumerate(rows):
        row_idx = i + 2
        fill = grey_fill if i % 2 else white_fill
        vals = [
            r['employee_code'],
            r['employee_name'],
            r['gosi_number'],
            r['nationality'],
            r['id_number'],
            float(r['basic_salary']),
            r['employee_rate'],
            float(r['employee_contribution']),
            r['employer_rate'],
            float(r['employer_contribution']),
            float(r['total_contribution']),
            r['month'],
            r['company_name'],
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.fill = fill
            if col in (6, 8, 10, 11):
                cell.number_format = money_fmt

    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=float(payload.get('total_employee_contributions') or 0)).font = Font(bold=True)
    ws.cell(row=total_row, column=8).number_format = money_fmt
    ws.cell(row=total_row, column=10, value=float(payload.get('total_employer_contributions') or 0)).font = Font(bold=True)
    ws.cell(row=total_row, column=10).number_format = money_fmt
    ws.cell(row=total_row, column=11, value=float(payload.get('grand_total') or 0)).font = Font(bold=True)
    ws.cell(row=total_row, column=11).number_format = money_fmt

    s1 = total_row + 2
    ws.cell(row=s1, column=1, value=(
        f"Saudi Employees: {payload.get('total_saudi', 0)} | "
        f"Total Saudi GOSI: SAR {float(payload.get('total_saudi_gosi') or 0):,.2f}"
    ))
    ws.cell(row=s1 + 1, column=1, value=(
        f"Non-Saudi Employees: {payload.get('total_non_saudi', 0)} | "
        f"Total Hazard (Employer): SAR {float(payload.get('total_non_saudi_hazard') or 0):,.2f}"
    ))
    ws.cell(row=s1 + 2, column=1, value=(
        f"Grand Total GOSI Liability: SAR {float(payload.get('grand_total') or 0):,.2f}"
    ))
    for rr in range(3):
        ws.cell(row=s1 + rr, column=1).font = Font(bold=True)

    ws.cell(row=s1 + 3, column=1, value=f'Generated (app TZ): {timezone.localtime().strftime("%Y-%m-%d %H:%M")}')

    for idx in range(1, len(headers) + 1):
        letter = get_column_letter(idx)
        max_len = len(headers[idx - 1])
        for r in range(1, min(total_row + 5, ws.max_row + 1)):
            v = ws.cell(row=r, column=idx).value
            if v is not None:
                max_len = max(max_len, min(len(str(v)), 50))
        ws.column_dimensions[letter].width = max_len + 2

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
